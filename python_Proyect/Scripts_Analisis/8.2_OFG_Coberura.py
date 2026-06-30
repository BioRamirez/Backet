#--------------## Cargar librerias necesarias------------------------------
import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from tabulate import tabulate
import openpyxl

# Carpeta donde guardarás los gráficos (solo una vez)
output_folder = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados"
os.makedirs(output_folder, exist_ok=True)
#--------------## Leer archivo y revisar columnas------------------------------

#--------------## Leer archivo y revisar columnas------------------------------
# Ruta del archivo
ruta = r"D:\Forestal Consultores\2026\FAUNA\BD\BD_SANROQUE.xlsx"
# Leer el archivo Excel
Registros = pd.read_excel(ruta)

# Mostrar las primeras filas
print(" Primeras filas del archivo:")
print(Registros.head())

# Mostrar nombres de las columnas
print("\n Columnas del DataFrame:")
print(Registros.columns)

# =========================================================
# FILTRAR UNA SOLA CLASE
# =========================================================

grupo = "AVES"   # AVES, MAMIFEROS, REPTILES, ANFIBIOS etc.

Registros = Registros[
    Registros["CLASE"].astype(str).str.upper() == grupo.upper()
].copy()

# Reiniciar índice
Registros.reset_index(drop=True, inplace=True)

# =========================================================
# VERIFICACIÓN
# =========================================================

print(f"\nGrupo seleccionado: {grupo}")
print(f"Número de registros: {len(Registros)}")

print("\nPrimeras filas:")
# --------------------------------------------------
# 2️⃣ Funciones para abreviar nombres de coberturas
# --------------------------------------------------

def generar_abreviacion(nombre):
    """
    Genera abreviaciones automáticas a partir de nombres de coberturas.
    Ejemplo: 'Bosque de galería y ripario' → 'BGR'
    """
    # Convertir a minúsculas y dividir en palabras
    palabras = nombre.lower().split()

    # Eliminar conectores comunes
    palabras = [p for p in palabras if p not in ['de', 'del', 'la', 'el', 'y', 'con', 'en', 'los', 'las']]

    # Tomar la primera letra de cada palabra y convertir a mayúsculas
    abreviacion = ''.join([p[0] for p in palabras]).upper()

    # Asegurar longitud mínima
    if len(abreviacion) < 3:
        abreviacion = abreviacion.ljust(3, ' ')

    return abreviacion


def abreviar_coberturas(df, columna='COBERTURA'):
    """
    Crea un diccionario de abreviaciones y reemplaza los nombres en el DataFrame.
    """
    coberturas_unicas = df[columna].dropna().unique()
    abreviaciones = {c: generar_abreviacion(c) for c in coberturas_unicas}

    print("\n Abreviaciones generadas automáticamente:")
    for original, abrev in abreviaciones.items():
        print(f"  {original} → {abrev}")

    # Reemplazar en el DataFrame
    df[columna] = df[columna].replace(abreviaciones)

    return df, abreviaciones


# --- Aplicar las abreviaciones ---
Registros, abreviaciones_cobertura = abreviar_coberturas(Registros, columna='COBERTURA')

import pandas as pd
import matplotlib.pyplot as plt
import numpy as np


# --- Verificar columnas clave ---
print(Registros.columns)


# --------------------------------------------------------------
# 3️⃣ Construir tabla de número de órdenes, familias, géneros
#     y especies por cada cobertura
# --------------------------------------------------------------

# Verificar que las columnas necesarias existan
columnas_necesarias = ["COBERTURA", "Orden", "Familia", "Genero", "ESPECIE"]
faltantes = [c for c in columnas_necesarias if c not in Registros.columns]

if faltantes:
    raise ValueError(f"Faltan columnas necesarias en la base de datos: {faltantes}")

# Agrupar por cobertura y calcular métricas
tabla_coberturas = (
    Registros.groupby("COBERTURA")
    .agg(
        Ordenes=("Orden", lambda x: x.nunique()),
        Familias=("Familia", lambda x: x.nunique()),
        Generos=("Genero", lambda x: x.nunique()),
        Especies=("ESPECIE", lambda x: x.nunique())
    )
)

# Crear fila TOTAL
fila_total = pd.DataFrame({
    "Ordenes": [Registros["Orden"].nunique()],
    "Familias": [Registros["Familia"].nunique()],
    "Generos": [Registros["Genero"].nunique()],
    "Especies": [Registros["ESPECIE"].nunique()]
}, index=["Total"])

# Unir tabla con fila TOTAL
tabla_final = pd.concat([tabla_coberturas, fila_total])

# Ordenar la tabla desde mayor a menor número de órdenes
tabla_final = tabla_final.sort_values(by="Especies", ascending=False)

print("\n Tabla final de diversidad por cobertura:")
print(tabla_final)

# --------------------------------------------------------------
# 4️⃣ Exportar tabla a archivo Excel en carpeta Resultados
# --------------------------------------------------------------
salida_excel = os.path.join(output_folder, "8.2_Diversidad_por_cobertura.xlsx")
tabla_final.to_excel(salida_excel, sheet_name="Coberturas", index=True)

print(f"\n Archivo exportado con éxito a:\n {salida_excel}")

#-------------------------------------------------------




#---------------------------------------------------
#   GRAFICO PROFESIONAL – TEMA MINIMALISTA + PAIRED
#---------------------------------------------------

import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd
import os

# ---- Preparar datos ----
tabla_plot = (
    tabla_final
    .reset_index()
    .rename(columns={"index": "COBERTURA"})
    .melt(
        id_vars="COBERTURA",
        value_vars=["Ordenes", "Familias", "Generos", "Especies"],
        var_name="Categoria",
        value_name="Cantidad"
    )
)

# ---- TEMA VISUAL IDENTICO AL GRÁFICO APILADO ----
sns.set_theme(style="whitegrid")
palette = sns.color_palette("Paired", 12)

fig, ax = plt.subplots(figsize=(14, 7))

# ---- Gráfico de barras ----
sns.barplot(
    data=tabla_plot,
    x="COBERTURA",
    y="Cantidad",
    hue="Categoria",
    palette=palette,
    edgecolor="black",
    linewidth=0.8,
    ax=ax
)

# ---- Etiquetas de valores ----
for p in ax.patches:
    height = p.get_height()
    if height > 0:
        ax.text(
            p.get_x() + p.get_width() / 2,
            height + 0.1,
            f"{int(height)}",
            ha='center', va='bottom',
            fontsize=9, fontweight='bold'
        )

# ---- Estética minimalista (idéntica al otro gráfico) ----
ax.set_facecolor("white")
ax.set_title("Riqueza Taxonómica por Cobertura",
             fontsize=16, fontweight="bold")

ax.set_xlabel("Cobertura", fontsize=12)
ax.set_ylabel("Cantidad", fontsize=12)

plt.xticks(rotation=45, ha="right", fontsize=11)
plt.yticks(fontsize=11)

# Bordes finos gris oscuro
for spine in ax.spines.values():
    spine.set_linewidth(0.8)
    spine.set_color("#444444")

# Grid suave como el gráfico apilado
ax.grid(True, axis="y", linestyle="--", alpha=0.35)
ax.grid(False, axis="x")

# ---- Leyenda minimalista ----
plt.legend(
    title="Categoría",
    title_fontsize=12,
    fontsize=11,
    frameon=False,   # igual al apilado
    loc="upper right"
)

plt.tight_layout()

# ---- Guardar ----
ruta_fig = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados"
os.makedirs(ruta_fig, exist_ok=True)

plt.savefig(
    os.path.join(ruta_fig, "8.2_Riqueza_Cobertura_PRO.png"),
    dpi=350,
    bbox_inches="tight"
)

plt.show()
plt.close()















#-----------------------------Por Municipio---------------

# ==============================================================
#  NUEVA SECCIÓN: GENERAR TABLAS DE DIVERSIDAD POR MUNICIPIO
# ==============================================================
import os

ruta_salida_graficos = r"D:\CORPONOR 2025\Backet\Resultados\Graficos"
os.makedirs(ruta_salida_graficos, exist_ok=True)

# ==============================================================
#  NUEVA SECCIÓN: GENERAR TABLAS DE DIVERSIDAD POR MUNICIPIO
# ==============================================================

print("\n=== GENERANDO TABLAS DE DIVERSIDAD POR MUNICIPIO ===")

ruta_salida_muni = os.path.join(output_folder, "DIVERSIDAD_POR_MUNICIPIO")
os.makedirs(ruta_salida_muni, exist_ok=True)

# Lista de municipios en los registros
municipios = sorted(Registros["MUNICIPIO"].dropna().unique())

# ------------------------------------------------------
# FUNCIÓN PARA APLICAR FORMATO (MISMO QUE EL ARCHIVO 8.2)
# ------------------------------------------------------
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

def aplicar_formato_excel(ruta_archivo):

    wb = load_workbook(ruta_archivo)
    ws = wb.active

    # Estilos
    header_fill = PatternFill(start_color='BFD8B8', end_color='BFD8B8', fill_type='solid')
    header_font = Font(bold=True, color='000000', name='Calibri')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Aplicar formato general
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or str(cell.value).strip() == "":
                cell.value = "-"
            cell.alignment = center_align
            cell.border = thin_border

    # Encabezado
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Ajuste de ancho de columnas
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                l = len(str(cell.value))
                if l > max_len:
                    max_len = l
        ws.column_dimensions[col_letter].width = max_len + 3

    # Altura filas
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 18

    wb.save(ruta_archivo)

# ==============================
#   GENERAR TABLAS POR MUNICIPIO
# ==============================

for muni in municipios:

    print(f"\n--- Procesando municipio: {muni} ---")

    # Filtrar registros del municipio
    df_muni = Registros[Registros["MUNICIPIO"] == muni]

    if df_muni.empty:
        print(f"⚠ Sin registros para el municipio {muni}, se omite.")
        continue

    # Construcción de tabla de diversidad
    tabla_muni = (
        df_muni.groupby("COBERTURA")
        .agg(
            Ordenes=("Orden", lambda x: x.nunique()),
            Familias=("Familia", lambda x: x.nunique()),
            Generos=("Genero", lambda x: x.nunique()),
            Especies=("ESPECIE", lambda x: x.nunique())
        )
    )

    # Fila TOTAL primero
    fila_total = pd.DataFrame({
        "Ordenes": [df_muni["Orden"].nunique()],
        "Familias": [df_muni["Familia"].nunique()],
        "Generos": [df_muni["Genero"].nunique()],
        "Especies": [df_muni["ESPECIE"].nunique()]
    }, index=[f"{muni}"])   # ← SIN “TOTAL” como pediste antes

    # Ordenar coberturas por número de especies
    tabla_muni_ordenada = tabla_muni.sort_values(by="Especies", ascending=False)

    # Unir TOTAL arriba + resto ordenado
    tabla_final = pd.concat([fila_total, tabla_muni_ordenada])

    # Guardar archivo por municipio
    archivo_muni = os.path.join(ruta_salida_muni, f"Diversidad_{muni}.xlsx")
    tabla_final.to_excel(archivo_muni, index=True)

    # Aplicar formato
    aplicar_formato_excel(archivo_muni)

    print(f"✔ Archivo generado y formateado: {archivo_muni}")

print("\n✔ PROCESO COMPLETADO PARA TODOS LOS MUNICIPIOS.")





#----------------------Graficar-----------------
#---------------------------------------------




































#------------------------------Grafico por municipio-----------------
for muni in municipios:

    archivo_muni = os.path.join(ruta_salida_muni, f"Diversidad_{muni}.xlsx")

    if not os.path.exists(archivo_muni):
        print(f"⚠ No hay tabla para {muni}, se omite gráfico.")
        continue

    # ------------------ Cargar tabla ------------------
    tabla_muni = pd.read_excel(archivo_muni)
    
    # Convertir SIEMPRE la primera columna a 'COBERTURA'
    tabla_muni = tabla_muni.rename(columns={tabla_muni.columns[0]: "COBERTURA"})

    # ------------------ ORDENAR: TOTAL PRIMERO ------------------
    if "TOTAL" in tabla_muni["COBERTURA"].values:
        tabla_muni["orden_temp"] = tabla_muni["COBERTURA"].apply(
            lambda x: 0 if x == "TOTAL" else 1
        )
        tabla_muni = tabla_muni.sort_values(by=["orden_temp", "COBERTURA"])
        tabla_muni = tabla_muni.drop(columns="orden_temp")

    # ------------------ Preparar formato long ------------------
    tabla_plot = (
        tabla_muni
        .melt(
            id_vars="COBERTURA",
            value_vars=["Ordenes", "Familias", "Generos", "Especies"],
            var_name="Categoria",
            value_name="Cantidad"
        )
    )

    # ------------------ Estilo del gráfico ------------------
    sns.set_theme(style="whitegrid")
    palette = sns.color_palette("Paired", 12)

    fig, ax = plt.subplots(figsize=(14, 7))

    sns.barplot(
        data=tabla_plot,
        x="COBERTURA",
        y="Cantidad",
        hue="Categoria",
        palette=palette,
        edgecolor="black",
        linewidth=0.8,
        ax=ax
    )

    # ---- Etiquetas ----
    for p in ax.patches:
        height = p.get_height()
        if height > 0:
            ax.text(
                p.get_x() + p.get_width() / 2,
                height + 0.01,
                f"{int(height)}",
                ha='center', va='bottom',
                fontsize=11, fontweight='bold'
            )

    # ---- Estética ----
    ax.set_facecolor("white")
    ax.set_title(f"Riqueza Taxonómica – Municipio de {muni}",
                 fontsize=16, fontweight="bold")
    ax.set_xlabel("Cobertura", fontsize=12)
    ax.set_ylabel("Cantidad", fontsize=12)

    plt.xticks(rotation=45, ha="right", fontsize=11)
    plt.yticks(fontsize=11)

    for spine in ax.spines.values():
        spine.set_linewidth(0.8)
        spine.set_color("#444444")

    ax.grid(True, axis="y", linestyle="--", alpha=0.35)
    ax.grid(False, axis="x")

    # ---- Colocar leyenda fuera del área del gráfico ----
    lgd = ax.legend(
        title="Categoría",
        title_fontsize=12,
        fontsize=11,
        frameon=False,
        bbox_to_anchor=(1.02, 1),   # Siempre fuera
        loc="upper left"
    )

    # Ajuste automático de espacio lateral
    fig.subplots_adjust(right=0.80)

    plt.tight_layout()

    # ------------------ Guardar gráfico ------------------
    salida_png = os.path.join(
        ruta_salida_graficos,
        f"Riqueza_Taxonomica_{muni}.png"
    )

    plt.savefig(salida_png, dpi=350, bbox_inches="tight")
    plt.close()

    print(f"✔ Gráfico generado para {muni}: {salida_png}")

























