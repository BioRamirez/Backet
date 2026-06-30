#--------------## Cargar librerias necesarias------------------------------
import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from tabulate import tabulate
import openpyxl

# Carpeta donde guardarás los gráficos (solo una vez)
output_folder = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados"
os.makedirs(output_folder, exist_ok=True)
#--------------## Leer archivo y revisar columnas------------------------------
# Ruta del archivo
ruta = r"D:\CORPONOR 2025\Backet\python_Proyect\data\POF_ZULIA_2025_BD_AVES_MAMIFEROS.xlsx"
ruta = r"D:\Aeropuerto Aguachica\OSO_PARDO_2025\Muestreo\GDB\ANALISIS\REPTILES\COMPARATIVO_R.xlsx"
# Leer el archivo Excel
Registros = pd.read_excel(ruta)

# Mostrar las primeras filas
print(" Primeras filas del archivo:")
print(Registros.head())

# Mostrar nombres de las columnas
print("\n Columnas del DataFrame:")
print(Registros.columns)


# ============================================================
# ANÁLISIS COMPARATIVO DE ABUNDANCIA Y COMPOSICIÓN DE ESPECIES
# ============================================================
# Autor: (tu nombre)
# Descripción:
# - Compara abundancia y composición de especies entre dos años
# - Funciona aunque los años tengan distinto número de especies
# - Usa métodos ecológicos estándar
# ============================================================


# ------------------------------------------------------------
# 1. IMPORTAR LIBRERÍAS NECESARIAS
# ------------------------------------------------------------
import pandas as pd
import numpy as np
from scipy.stats import wilcoxon
from scipy.spatial.distance import jaccard, braycurtis

# ============================================================
# ANÁLISIS COMPARATIVO DE ABUNDANCIA Y COMPOSICIÓN DE ESPECIES
# + PERMANOVA
# ============================================================

import pandas as pd
import numpy as np
import os
from scipy.stats import wilcoxon
from scipy.spatial.distance import jaccard, braycurtis, pdist, squareform
from skbio.stats.distance import DistanceMatrix, permanova


# ------------------------------------------------------------
# 1. PREPARACIÓN DE DATOS
# ------------------------------------------------------------

df = Registros.rename(columns={
    'Especies_2022': 'Especie_2022',
    'Individuos_2022': 'Abund_2022',
    'Especies_2025': 'Especie_2025',
    'Individuos_2025': 'Abund_2025'
})

df_2022 = df[['Especie_2022', 'Abund_2022']].dropna()
df_2025 = df[['Especie_2025', 'Abund_2025']].dropna()

df_2022.columns = ['Especie', 'Abundancia_2022']
df_2025.columns = ['Especie', 'Abundancia_2025']

df_total = pd.merge(df_2022, df_2025, on='Especie', how='outer')
df_total[['Abundancia_2022', 'Abundancia_2025']] = \
    df_total[['Abundancia_2022', 'Abundancia_2025']].fillna(0)


# ------------------------------------------------------------
# 2. MÉTRICAS BÁSICAS
# ------------------------------------------------------------

abund_2022 = df_total['Abundancia_2022'].sum()
abund_2025 = df_total['Abundancia_2025'].sum()

riqueza_2022 = (df_total['Abundancia_2022'] > 0).sum()
riqueza_2025 = (df_total['Abundancia_2025'] > 0).sum()


# ------------------------------------------------------------
# 3. WILCOXON
# ------------------------------------------------------------

stat, p_value = wilcoxon(
    df_total['Abundancia_2022'],
    df_total['Abundancia_2025']
)


# ------------------------------------------------------------
# 4. ÍNDICES DE SIMILITUD
# ------------------------------------------------------------

pres_2022 = df_total['Abundancia_2022'] > 0
pres_2025 = df_total['Abundancia_2025'] > 0

jaccard_sim = 1 - jaccard(pres_2022, pres_2025)
bray = braycurtis(df_total['Abundancia_2022'],
                   df_total['Abundancia_2025'])


# ------------------------------------------------------------
# 5. PERMANOVA (NÚCLEO DEL ANÁLISIS)
# ------------------------------------------------------------
# ------------------------------------------------------------
# 5. PERMANOVA (solo si hay réplicas)
# ------------------------------------------------------------

permanova_p = np.nan
permanova_msg = "No aplicable (una sola muestra por grupo)"

# Para PERMANOVA se requieren ≥ 2 muestras por grupo
if df_total.shape[0] > 2:

    matriz = df_total[['Abundancia_2022', 'Abundancia_2025']].T
    dist_matrix = squareform(pdist(matriz, metric='braycurtis'))

    dm = DistanceMatrix(dist_matrix, ids=['2022', '2025'])
    grupos = ['2022', '2025']

    try:
        permanova_result = permanova(dm, grupos, permutations=999)
        permanova_p = permanova_result['p-value']
        permanova_msg = "PERMANOVA ejecutado correctamente"
    except:
        permanova_msg = "PERMANOVA no aplicable (sin réplicas)"

else:
    permanova_msg = "PERMANOVA no aplicable (n < 3)"

# ------------------------------------------------------------
# 6. TABLA RESUMEN FINAL
# ------------------------------------------------------------
resumen = pd.DataFrame({
    'Métrica': [
        'Abundancia total',
        'Riqueza de especies',
        'Similitud Jaccard',
        'Disimilitud Bray-Curtis',
        'p-valor Wilcoxon',
        'p-valor PERMANOVA'
    ],
    '2022': [
        abund_2022,
        riqueza_2022,
        jaccard_sim,
        bray,
        p_value,
        permanova_p
    ],
    '2025': [
        abund_2025,
        riqueza_2025,
        jaccard_sim,
        bray,
        p_value,
        permanova_p
    ]
})

print("\n--- RESUMEN FINAL ---")
print(resumen)
print("\nPERMANOVA:", permanova_msg)



# ------------------------------------------------------------
# 7. EXPORTAR RESULTADOS
# ------------------------------------------------------------

output_path = os.path.join(output_folder, "15_tabla_Comparativo.xlsx")
resumen.to_excel(output_path, index=False)

print(f"\nArchivo guardado en: {output_path}")




#---------------------------------- Reparar y formatear archivo de tabla_sensibilidad -----------------------
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import os

# --- Rutas ---
ruta_original = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\15_tabla_Comparativo.xlsx"
ruta_limpia = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\15_tabla_Comparativo.xlsx"

# --- Verificar existencia ---
if not os.path.exists(ruta_original):
    raise FileNotFoundError(f"⚠️ No se encontró el archivo: {ruta_original}")

# --- Leer archivo dañado con pandas ---
try:
    df = pd.read_excel(ruta_original)
    print(" Archivo leído correctamente con pandas.")
except Exception as e:
    raise RuntimeError(f" No se pudo leer el archivo: {e}")

# --- Reescribir el archivo limpio ---
df.to_excel(ruta_limpia, index=False)
print(f" Archivo reparado y guardado como:\n{ruta_limpia}")

# --- Aplicar formato con openpyxl ---
from openpyxl import load_workbook

wb = load_workbook(ruta_limpia)
ws = wb.active

# --- Estilos base ---
header_fill = PatternFill(start_color='BFD8B8', end_color='BFD8B8', fill_type='solid')
header_font = Font(bold=True, color='000000', name='Calibri')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# --- Aplicar formato y reemplazar vacíos ---
for row in ws.iter_rows():
    for cell in row:
        if cell.value is None or str(cell.value).strip() == '':
            cell.value = '-'
        cell.alignment = center_align
        cell.border = thin_border

# --- Encabezado ---
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align

# --- Ajustar ancho de columnas ---
for col in ws.columns:
    max_length = 0
    column = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            length = len(str(cell.value))
            if length > max_length:
                max_length = length
    ws.column_dimensions[column].width = max_length + 3

# --- Ajustar altura de filas ---
for row in ws.iter_rows():
    ws.row_dimensions[row[0].row].height = 18

# --- Guardar cambios ---
wb.save(ruta_limpia)
print(f' Archivo formateado y reparado correctamente:\n{ruta_limpia}')


def interpretar_resultados(resumen):
    """
    Genera una interpretación ecológica automática a partir del DataFrame resumen,
    incorporando abundancia, riqueza, similitud, Bray-Curtis y PERMANOVA.
    """

    # -------------------------------------------------
    # EXTRACCIÓN DE VALORES
    # -------------------------------------------------
    abund_2022 = resumen.loc[resumen['Métrica'] == 'Abundancia total', '2022'].values[0]
    abund_2025 = resumen.loc[resumen['Métrica'] == 'Abundancia total', '2025'].values[0]

    riqueza_2022 = resumen.loc[resumen['Métrica'] == 'Riqueza de especies', '2022'].values[0]
    riqueza_2025 = resumen.loc[resumen['Métrica'] == 'Riqueza de especies', '2025'].values[0]

    jaccard = resumen.loc[resumen['Métrica'] == 'Similitud Jaccard', '2022'].values[0]
    bray = resumen.loc[resumen['Métrica'] == 'Disimilitud Bray-Curtis', '2022'].values[0]

    p_wilcoxon = resumen.loc[resumen['Métrica'] == 'p-valor Wilcoxon', '2022'].values[0]
    p_permanova = resumen.loc[resumen['Métrica'] == 'p-valor PERMANOVA', '2022'].values[0]

    # -------------------------------------------------
    # INTERPRETACIONES AUTOMÁTICAS
    # -------------------------------------------------

    # Abundancia
    if abund_2025 > abund_2022:
        abund_txt = "un incremento en la abundancia total"
    elif abund_2025 < abund_2022:
        abund_txt = "una disminución en la abundancia total"
    else:
        abund_txt = "una estabilidad en la abundancia total"

    # Riqueza
    if riqueza_2025 > riqueza_2022:
        riqueza_txt = "un aumento en la riqueza de especies"
    elif riqueza_2025 < riqueza_2022:
        riqueza_txt = "una reducción en la riqueza de especies"
    else:
        riqueza_txt = "una riqueza de especies estable"

    # Jaccard
    if jaccard < 0.3:
        jaccard_txt = "una baja similitud en la composición de especies"
    elif jaccard < 0.6:
        jaccard_txt = "una similitud moderada en la composición de especies"
    else:
        jaccard_txt = "una alta similitud en la composición de especies"

    # Bray–Curtis
    if bray > 0.6:
        bray_txt = "altas diferencias en la estructura de abundancia"
    elif bray > 0.3:
        bray_txt = "diferencias moderadas en la estructura de abundancia"
    else:
        bray_txt = "estructura de abundancia similar entre los años"

    # Wilcoxon
    if p_wilcoxon < 0.05:
        wilcoxon_txt = "diferencias estadísticamente significativas en la abundancia"
    else:
        wilcoxon_txt = "ausencia de diferencias estadísticamente significativas en la abundancia"

    # PERMANOVA
    if p_permanova < 0.05:
        permanova_txt = (
            "diferencias estadísticamente significativas en la composición de especies, "
            "lo que indica cambios estructurales en la comunidad entre los periodos evaluados"
        )
    else:
        permanova_txt = (
            "ausencia de diferencias estadísticamente significativas en la composición de especies, "
            "lo que sugiere una estructura comunitaria relativamente estable"
        )

    # -------------------------------------------------
    # TEXTO FINAL
    # -------------------------------------------------

    interpretacion = f"""
INTERPRETACIÓN ECOLÓGICA AUTOMÁTICA

El análisis comparativo entre los años evaluados evidencia {abund_txt}, 
junto con {riqueza_txt}.

La evaluación de la composición específica mediante el índice de Jaccard 
(J = {jaccard:.2f}) indica {jaccard_txt}, lo que sugiere un grado importante 
de recambio de especies entre los periodos analizados.

El índice de Bray–Curtis (BC = {bray:.2f}) refleja {bray_txt}, evidenciando 
cambios en la estructura y dominancia de la comunidad.

Desde el punto de vista estadístico, la prueba de Wilcoxon indicó 
{wilcoxon_txt} (p = {p_wilcoxon:.4f}).

Adicionalmente, el análisis PERMANOVA mostró {permanova_txt} 
(p = {p_permanova:.4f}), lo que permite evaluar de manera robusta 
las diferencias multivariadas en la composición biológica.

En conjunto, los resultados sugieren que, aunque pueden presentarse variaciones 
en la composición y estructura de la comunidad, estas responden a procesos 
ecológicos dinámicos propios del sistema evaluado.
"""

    return interpretacion


# Generar interpretación automática

texto_interpretacion = interpretar_resultados(resumen)
print(texto_interpretacion)











#-------------------GRAFICOS COMPARATIVOS DE ABUNDANCIA POR ESPECIE -----------------------

# ============================================================
# GRÁFICO WILCOXON – ABUNDANCIA 2022 vs 2025
# ============================================================

import matplotlib.pyplot as plt
import seaborn as sns

# Preparar datos en formato largo
df_plot = df_total.melt(
    id_vars='Especie',
    value_vars=['Abundancia_2022', 'Abundancia_2025'],
    var_name='Año',
    value_name='Abundancia'
)

# Renombrar para estética
df_plot['Año'] = df_plot['Año'].replace({
    'Abundancia_2022': '2022',
    'Abundancia_2025': '2025'
})

# ===============================
# CREAR FIGURA
# ===============================
plt.figure(figsize=(7, 5))

# Boxplot
sns.boxplot(
    data=df_plot,
    x='Año',
    y='Abundancia',
    width=0.5,
    showfliers=False,
    color="#D0D0D0"
)

# Puntos individuales
sns.stripplot(
    data=df_plot,
    x='Año',
    y='Abundancia',
    color='black',
    size=5,
    jitter=True
)

# Líneas pareadas
for i in range(len(df_total)):
    plt.plot(
        ['2022', '2025'],
        [df_total.loc[i, 'Abundancia_2022'],
         df_total.loc[i, 'Abundancia_2025']],
        color='gray',
        alpha=0.5,
        linewidth=1
    )

# Anotación del p-valor
plt.text(
    0.5,
    max(df_plot['Abundancia']) * 1.05,
    f"Wilcoxon p = {p_value:.4f}",
    ha='center',
    fontsize=11,
    fontweight='bold'
)

# Estética final
plt.title("Comparación de abundancia por especie (Wilcoxon pareado)", fontsize=13)
plt.ylabel("Abundancia")
plt.xlabel("")
sns.despine()
plt.tight_layout()

# Guardar
ruta_fig = os.path.join(output_folder, "Wilcoxon_Abundancia_2022_vs_2025.png")
plt.savefig(ruta_fig, dpi=300)
plt.show()

print(f"Gráfico guardado en: {ruta_fig}")








#--------------------SEGUNDO GRAFICO------------------------

# ============================================================
# GRÁFICO KDE SUPERPUESTO – WILCOXON
# ============================================================

import matplotlib.pyplot as plt
import seaborn as sns

plt.figure(figsize=(8, 5))

# KDE 2022
sns.kdeplot(
    df_total['Abundancia_2022'],
    fill=True,
    alpha=0.5,
    linewidth=2,
    label='2022'
)

# KDE 2025
sns.kdeplot(
    df_total['Abundancia_2025'],
    fill=True,
    alpha=0.5,
    linewidth=2,
    label='2025'
)

# Línea del valor central (mediana)
plt.axvline(df_total['Abundancia_2022'].median(),
            linestyle='--', linewidth=1.5, label='Mediana 2022')

plt.axvline(df_total['Abundancia_2025'].median(),
            linestyle='--', linewidth=1.5, label='Mediana 2025')

# Anotación estadística
plt.text(
    0.95,
    0.90,
    f"Wilcoxon p = {p_value:.4f}",
    transform=plt.gca().transAxes,
    ha='right',
    fontsize=11,
    fontweight='bold'
)

# Estética final
plt.title("Distribución de abundancia por especie (2022 vs 2025)")
plt.xlabel("Abundancia")
plt.ylabel("Densidad")
plt.legend()
sns.despine()
plt.tight_layout()

# Guardar figura
ruta_fig = os.path.join(output_folder, "Wilcoxon_KDE_Abundancia.png")
plt.savefig(ruta_fig, dpi=300)
plt.show()

print(f"Gráfico guardado en: {ruta_fig}")

#--------------------3 grafico--------------

# ============================================================
# GRÁFICO DE OLAS – ABUNDANCIA POR ESPECIE
# ============================================================

import matplotlib.pyplot as plt
import numpy as np

# Ordenar especies por abundancia promedio
df_plot = df_total.copy()
df_plot["Promedio"] = df_plot[["Abundancia_2022", "Abundancia_2025"]].mean(axis=1)
df_plot = df_plot.sort_values("Promedio", ascending=False)

x = np.arange(len(df_plot))

plt.figure(figsize=(10, 5))


# OLA 2025
plt.fill_between(
    x,
    df_plot["Abundancia_2025"],
    alpha=0.5,
    label="2025"
)

# OLA 2022
plt.fill_between(
    x,
    df_plot["Abundancia_2022"],
    alpha=0.5,
    label="2022"
)
# Estética
plt.xticks(x, df_plot["Especie"], rotation=90, fontsize=8)
plt.ylabel("Número de individuos")
plt.xlabel("Especies")
plt.title("Comparación de abundancia por especie (2022 vs 2025)")
plt.legend()
plt.tight_layout()

# Guardar
ruta_fig = os.path.join(output_folder, "Olas_Abundancia_Especies.png")
plt.savefig(ruta_fig, dpi=300)
plt.show()

print(f"Gráfico guardado en: {ruta_fig}")




#--------------GRAFICO 4--------------------


#--------------------

# ============================================================
# CURVA RANGO–ABUNDANCIA (OLAS) – SIN CEROS VISUALES
# ============================================================

import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.ticker import MaxNLocator

colors = sns.color_palette("Paired")

abund_2022 = np.sort(df_total['Abundancia_2022'])[::-1]
abund_2025 = np.sort(df_total['Abundancia_2025'])[::-1]

# Eliminar ceros solo para visualización
abund_2022 = np.where(abund_2022 == 0, np.nan, abund_2022)
abund_2025 = np.where(abund_2025 == 0, np.nan, abund_2025)

x_2022 = np.arange(1, len(abund_2022) + 1)
x_2025 = np.arange(1, len(abund_2025) + 1)

plt.figure(figsize=(8, 5))

plt.fill_between(
    x_2025,
    abund_2025,
    color=colors[1],
    alpha=0.6,
    label="2025"
)

plt.fill_between(
    x_2022,
    abund_2022,
    color=colors[7],
    alpha=0.6,
    label="2022"
)

# 🔧 FORZAR ENTEROS EN EJE X
plt.gca().xaxis.set_major_locator(MaxNLocator(integer=True))

plt.xlabel("Rango de especies")
plt.ylabel("Número de individuos")
plt.title("Curva rango–abundancia (Comparativo temporal)")
plt.legend(frameon=False)
sns.despine()
plt.tight_layout()

# Guardar figura
ruta_fig = os.path.join(output_folder, "Curva_Rango_Abundancia_Sin_Ceros.png")
plt.savefig(ruta_fig, dpi=300)
plt.show()







#--------------------

# ============================================================
# CURVA RANGO–ABUNDANCIA (OLAS) – SIN CEROS VISUALES
# + LEYENDA CON PRUEBAS ESTADÍSTICAS
# ============================================================

import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.ticker import MaxNLocator

colors = sns.color_palette("Paired")

# Ordenar abundancias
abund_2022 = np.sort(df_total['Abundancia_2022'])[::-1]
abund_2025 = np.sort(df_total['Abundancia_2025'])[::-1]

# Eliminar ceros solo para visualización
abund_2022 = np.where(abund_2022 == 0, np.nan, abund_2022)
abund_2025 = np.where(abund_2025 == 0, np.nan, abund_2025)

x_2022 = np.arange(1, len(abund_2022) + 1)
x_2025 = np.arange(1, len(abund_2025) + 1)

plt.figure(figsize=(8, 5))

# Olas con etiquetas enriquecidas
plt.fill_between(
    x_2025,
    abund_2025,
    color=colors[1],
    alpha=0.6,
    label="2025"
)

plt.fill_between(
    x_2022,
    abund_2022,
    color=colors[7],
    alpha=0.6,
    label="2022"
)

# 🔧 Forzar enteros en eje X
plt.gca().xaxis.set_major_locator(MaxNLocator(integer=True))

plt.xlabel("Rango de especies")
plt.ylabel("Número de individuos")
plt.title("Curva rango–abundancia (comparativo temporal)")

# ========================
# LEYENDA ESTADÍSTICA
# ========================

plt.legend(
    title=legend_text,
    frameon=False,
    loc="upper right"
)



legend_text = (
    f"Wilcoxon p = {p_value:.4f}\n"
    f"Jaccard = {jaccard_sim:.3f}\n"
    f"Bray–Curtis = {bray:.3f}"
)


sns.despine()
plt.tight_layout()

# Guardar figura
ruta_fig = os.path.join(output_folder, "Curva_Rango_Abundancia_Sin_Ceros.png")
plt.savefig(ruta_fig, dpi=300)
plt.show()









# ============================================================
# CURVA RANGO–ABUNDANCIA (OLAS)
# LEYENDA ALINEADA Y TEXTO CENTRADO
# ============================================================

import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.ticker import MaxNLocator
from matplotlib.lines import Line2D

colors = sns.color_palette("Paired")

# Ordenar abundancias
abund_2022 = np.sort(df_total['Abundancia_2022'])[::-1]
abund_2025 = np.sort(df_total['Abundancia_2025'])[::-1]

# Eliminar ceros solo para visualización
abund_2022 = np.where(abund_2022 == 0, np.nan, abund_2022)
abund_2025 = np.where(abund_2025 == 0, np.nan, abund_2025)

x_2022 = np.arange(1, len(abund_2022) + 1)
x_2025 = np.arange(1, len(abund_2025) + 1)

plt.figure(figsize=(8, 5))

plt.fill_between(
    x_2025, abund_2025,
    color=colors[1], alpha=0.6, label="2025"
)

plt.fill_between(
    x_2022, abund_2022,
    color=colors[7], alpha=0.6, label="2022"
)

plt.gca().xaxis.set_major_locator(MaxNLocator(integer=True))

plt.xlabel("Rango de especies")
plt.ylabel("Número de individuos")
plt.title("Curva rango–abundancia (comparativo temporal)")

# ========================
# LEYENDA ALINEADA
# ========================
handles = [
    Line2D([0], [0], color=colors[1], lw=8, label="2025"),
    Line2D([0], [0], color=colors[7], lw=8, label="2022"),
    Line2D([0], [0], color='none', label=""),  # separador
    Line2D([0], [0], color='none', label=
           f"Wilcoxon p = {p_value:.4f}\n"
           f"Jaccard = {jaccard_sim:.3f}\n"
           f"Bray–Curtis = {bray:.3f}")
]

legend = plt.legend(
    handles=handles,
    frameon=False,
    loc="upper right",
    ncol=1,
    handlelength=1,
    handletextpad=0.6
)

# 🔧 Centrar texto estadístico
for text in legend.get_texts():
    text.set_ha("center")

sns.despine()
plt.tight_layout()

# Guardar
ruta_fig = os.path.join(output_folder, "Curva_Rango_Abundancia_Leyenda_Centrada.png")
plt.savefig(ruta_fig, dpi=300)
plt.show()
