#--------------## Cargar librerias necesarias------------------------------
import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from tabulate import tabulate
import openpyxl

# Carpeta donde guardarás los gráficos (solo una vez)
output_folder = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados"
os.makedirs(output_folder, exist_ok=True)
#--------------## Leer archivo y revisar columnas------------------------------
# Ruta del archivo
ruta = r"D:\Aeropuerto Aguachica\OSO_PARDO_2025\Muestreo\GDB\ANALISIS\MAMIFEROS\MAMIFEROS_OSO_PARDO_2025.xlsx"

# Leer el archivo Excel
Registros = pd.read_excel(ruta)

# Mostrar las primeras filas
print(" Primeras filas del archivo:")
print(Registros.head())

# Mostrar nombres de las columnas
print("\n Columnas del DataFrame:")
print(Registros.columns)

# --------------------------------------------------
# 2️⃣ Funciones para abreviar nombres de coberturas
# --------------------------------------------------

def generar_abreviacion(nombre):
    """
    Genera abreviaciones automáticas a partir de nombres de coberturas.
    Ejemplo: 'Bosque de galería y ripario' → 'BGR'
    """
    # Convertir a minúsculas y dividir en palabras
    palabras = nombre.lower().split()

    # Eliminar conectores comunes
    palabras = [p for p in palabras if p not in ['de', 'del', 'la', 'el', 'y', 'con', 'en', 'los', 'las']]

    # Tomar la primera letra de cada palabra y convertir a mayúsculas
    abreviacion = ''.join([p[0] for p in palabras]).upper()

    # Asegurar longitud mínima
    if len(abreviacion) < 3:
        abreviacion = abreviacion.ljust(3, ' ')

    return abreviacion



def abreviar_coberturas(df, columna='COBERTURA'):
    """
    Crea un diccionario de abreviaciones y reemplaza los nombres en el DataFrame.
    """
    coberturas_unicas = df[columna].dropna().unique()
    abreviaciones = {c: generar_abreviacion(c) for c in coberturas_unicas}

    print("\n Abreviaciones generadas automáticamente:")
    for original, abrev in abreviaciones.items():
        print(f"  {original} → {abrev}")

    # Reemplazar en el DataFrame
    df[columna] = df[columna].replace(abreviaciones)

    return df, abreviaciones


# --- Aplicar las abreviaciones ---
Registros, abreviaciones_cobertura = abreviar_coberturas(Registros, columna='COBERTURA')

import pandas as pd
import matplotlib.pyplot as plt
import numpy as np


# --- Verificar columnas clave ---
print(Registros.columns)

Registros


# ============================================
#  Análisis del uso de hábitat por cobertura
# ============================================

import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns

# --- Cargar los datos (ya los tienes en Registros)
# Asegurar que no haya valores nulos en columnas clave
df = Registros.dropna(subset=['ESPECIE', 'COBERTURA']).copy()

# ===============================
# 1️⃣ Riqueza total de especies por cobertura
# ===============================
riqueza_por_cobertura = (
    df.groupby('COBERTURA')['ESPECIE']
    .nunique()
    .reset_index()
    .rename(columns={'ESPECIE': 'Riqueza_total'})
)

# ===============================
# 2️⃣ Especies exclusivas por cobertura
# ===============================
# Calcular cuántas coberturas tiene cada especie
coberturas_por_especie = (
    df[['ESPECIE', 'COBERTURA']].drop_duplicates()
    .groupby('ESPECIE')['COBERTURA']
    .nunique()
    .reset_index()
    .rename(columns={'COBERTURA': 'Num_coberturas'})
)

# Filtrar especies presentes en una sola cobertura
especies_exclusivas = coberturas_por_especie.loc[
    coberturas_por_especie['Num_coberturas'] == 1, 'ESPECIE'
]

# Contar cuántas exclusivas hay por cobertura
exclusivas_por_cobertura = (
    df[df['ESPECIE'].isin(especies_exclusivas)]
    .groupby('COBERTURA')['ESPECIE']
    .nunique()
    .reset_index()
    .rename(columns={'ESPECIE': 'Especies_exclusivas'})
)

# ===============================
# 3️⃣ Combinar ambos resultados
# ===============================
uso_habitat = pd.merge(
    riqueza_por_cobertura,
    exclusivas_por_cobertura,
    on='COBERTURA',
    how='left'
).fillna(0)

# Calcular el porcentaje de especies exclusivas
uso_habitat['%_Exclusivas'] = (
    uso_habitat['Especies_exclusivas'] / uso_habitat['Riqueza_total'] * 100
).round(2)

# ===============================
# 4️⃣ Mostrar tabla resumen
# ===============================
print("\n Resumen del uso de hábitat por cobertura:")
print(uso_habitat.sort_values(by='Riqueza_total', ascending=False))

# ===============================
#  GUARDAR RESULTADOS
# ===============================

# 1️⃣ Guardar resumen como archivo Excel
resumen_path = os.path.join(output_folder, "10_Resumen_Uso_Habitat.xlsx")
uso_habitat.to_excel(resumen_path, index=False)
print(f" Resumen guardado en: {resumen_path}")

#---------------------------------- Reparar y formatear archivo de Resumen_Uso_Habitat -----------------------
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import os

# --- Rutas ---
ruta_original = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\10_Resumen_Uso_Habitat.xlsx"
ruta_limpia = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\10_Resumen_Uso_Habitat.xlsx"

# --- Verificar existencia ---
if not os.path.exists(ruta_original):
    raise FileNotFoundError(f"⚠️ No se encontró el archivo: {ruta_original}")

# --- Leer archivo dañado con pandas ---
try:
    df = pd.read_excel(ruta_original)
    print(" Archivo leído correctamente con pandas.")
except Exception as e:
    raise RuntimeError(f" No se pudo leer el archivo: {e}")

# --- Reescribir el archivo limpio ---
df.to_excel(ruta_limpia, index=False)
print(f" Archivo reparado y guardado como:\n{ruta_limpia}")

# --- Aplicar formato con openpyxl ---
from openpyxl import load_workbook

wb = load_workbook(ruta_limpia)
ws = wb.active

# --- Estilos base ---
header_fill = PatternFill(start_color='BFD8B8', end_color='BFD8B8', fill_type='solid')
header_font = Font(bold=True, color='000000', name='Calibri')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# --- Aplicar formato y reemplazar vacíos ---
for row in ws.iter_rows():
    for cell in row:
        if cell.value is None or str(cell.value).strip() == '':
            cell.value = '-'
        cell.alignment = center_align
        cell.border = thin_border

# --- Encabezado ---
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align

# --- Ajustar ancho de columnas ---
for col in ws.columns:
    max_length = 0
    column = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            length = len(str(cell.value))
            if length > max_length:
                max_length = length
    ws.column_dimensions[column].width = max_length + 3

# --- Ajustar altura de filas ---
for row in ws.iter_rows():
    ws.row_dimensions[row[0].row].height = 18

# --- Guardar cambios ---
wb.save(ruta_limpia)
print(f' Archivo formateado y reparado correctamente:\n{ruta_limpia}')



# ===============================
# 5️⃣ Gráfico comparativo
# ===============================
plt.figure(figsize=(10,6))
sns.barplot(
    data=uso_habitat.sort_values('Riqueza_total', ascending=False),
    x='COBERTURA',
    y='Riqueza_total',
    color='skyblue',
    label='Riqueza total'
)
sns.barplot(
    data=uso_habitat.sort_values('Riqueza_total', ascending=False),
    x='COBERTURA',
    y='Especies_exclusivas',
    color='steelblue',
    label='Exclusivas'
)

plt.title('Uso de hábitat: riqueza y especies exclusivas por cobertura')
plt.ylabel('Número de especies')
plt.xlabel('Cobertura')
plt.xticks(rotation=45, ha='right')
plt.legend()
plt.tight_layout()
plt.show()










from matplotlib import cm
pa = cm.get_cmap("Paired").colors
print(pa)


# --- Colores tomados de la paleta Paired ---
color_excl = "#FF7F00"   # Naranja Paired
color_comp = "#1F78B4"   # Azul Paired






#----------------------------------------------
# 6️⃣ Gráfico de barras apiladas


import matplotlib.pyplot as plt
import numpy as np

# --- Datos ordenados ---
data_plot = uso_habitat.sort_values('Riqueza_total', ascending=False)
x = np.arange(len(data_plot))
width = 0.6

# --- Colores ---
color_excl = '#FF7F00'  # naranja
color_comp = '#1F78B4'  # azul

# --- Calcular compartidas ---
data_plot['Especies_compartidas'] = data_plot['Riqueza_total'] - data_plot['Especies_exclusivas']

# --- Crear barras apiladas ---
plt.figure(figsize=(10,6))
plt.bar(x, data_plot['Especies_exclusivas'], color=color_excl, label='Exclusivas')
plt.bar(x, data_plot['Especies_compartidas'], 
        bottom=data_plot['Especies_exclusivas'], color=color_comp, label='Compartidas')

# --- Etiquetas ---
for i, row in data_plot.iterrows():
    exclusivas = row['Especies_exclusivas']
    compartidas = row['Especies_compartidas']
    total = row['Riqueza_total']
    
    # Exclusivas (centro de la barra naranja)
    plt.text(i, exclusivas/2, f"{int(exclusivas)}", ha='center', va='center', color='black', fontsize=9)
    
    # Compartidas (centro de la barra azul)
    plt.text(i, exclusivas + compartidas/2, f"{int(compartidas)}", ha='center', va='center', color='black', fontsize=9)
    
    # Total (encima)
    plt.text(i, total + 1, f"Total: {int(total)}", ha='center', va='bottom', fontsize=9, color='black')


# --- Personalización ---
plt.xticks(x, data_plot['COBERTURA'], rotation=45, ha='right')
plt.ylabel('Número de especies')
plt.xlabel('Cobertura')
#Uso de hábitat
plt.title('', fontsize=13, weight='bold')
plt.legend(
    title="Categorías",
    fontsize=11,       #  Tamaño del texto de la leyenda
    title_fontsize=12, #  Tamaño del título de la leyenda (opcional)
    loc='upper right'  #  Puedes moverla (ej: 'upper left', 'lower right', etc.)
)
plt.tight_layout()
plt.grid(
    True,           # activa la grilla
    axis='y',       # solo líneas horizontales (eje Y)
    linestyle='--', # tipo de línea (puede ser '-', '--', ':', '-.')
    alpha=0.4,      # transparencia
    zorder=0        # asegura que quede detrás de las barras
)


# 2️⃣ Guardar la figura apilada como imagen
fig_path = os.path.join(output_folder, "10.1_Grafico_Uso_Habitat_Apilado.png")
plt.savefig(fig_path, dpi=300, bbox_inches='tight')
print(f" Gráfico guardado en: {fig_path}")

plt.show()













































# ====================================================
#   GRÁFICO APILADO – ESTILO MINIMALISTA + PAIRED
# ====================================================

import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
import os

# --- PALETA PAIRED (misma que usas en índices) ---
set2 = sns.color_palette("Paired", 12)

color_excl = set2[3]   # naranja 7
color_comp = set2[1]   # azul intenso

# Activar estilo general
sns.set_theme(style="whitegrid")

# --- Datos ordenados ---
data_plot = uso_habitat.sort_values("Riqueza_total", ascending=False)
x = np.arange(len(data_plot))

# --- Calcular compartidas ---
data_plot["Especies_compartidas"] = data_plot["Riqueza_total"] - data_plot["Especies_exclusivas"]

# --- Crear figura ---
fig, ax = plt.subplots(figsize=(10, 6))

# Barras apiladas
ax.bar(
    x, data_plot["Especies_exclusivas"],
    color=color_excl,
    edgecolor="black",
    linewidth=0.8,
    label="Exclusivas"
)

ax.bar(
    x, data_plot["Especies_compartidas"],
    bottom=data_plot["Especies_exclusivas"],
    color=color_comp,
    edgecolor="black",
    linewidth=0.8,
    label="Compartidas"
)
# --- Etiquetas corregidas ---
for i, row in enumerate(data_plot.itertuples()):
    excl = row.Especies_exclusivas
    comp = row.Especies_compartidas
    tot = row.Riqueza_total

    ax.text(i, excl / 2, f"{int(excl)}",
            ha="center", va="center",
            color="white", fontsize=9, fontweight="bold")

    ax.text(i, excl + comp / 2, f"{int(comp)}",
            ha="center", va="center",
            color="white", fontsize=9, fontweight="bold")

    ax.text(i, tot + 0.3, f"{int(tot)}",
            ha="center", fontsize=10, fontweight="bold")


# --- Estética minimalista ---
ax.set_xticks(x)
ax.set_xticklabels(data_plot["COBERTURA"], rotation=45, ha="right", fontsize=11)

ax.set_ylabel("Número de especies", fontsize=12)
ax.set_xlabel("Cobertura", fontsize=12)
ax.set_title("Riqueza por cobertura (Exclusivas vs Compartidas)", fontsize=15, fontweight="bold")

# Bordes minimalistas
for spine in ax.spines.values():
    spine.set_linewidth(0.8)
    spine.set_color("#444444")

# Grid suave
ax.grid(True, axis="y", linestyle="--", alpha=0.4)

# --- Leyenda minimalista ---
leg = ax.legend(
    title="Categorías",
    fontsize=11,
    title_fontsize=12,
    loc="upper right",
    frameon=False
)

plt.tight_layout()

# Guardado
fig_path = os.path.join(output_folder, "10.1_Grafico_Uso_Habitat_Apilado.png")
plt.savefig(fig_path, dpi=300, bbox_inches="tight")

plt.show()

print("✔ Estilo aplicado y gráfico guardado")



























#-------------------------INTERPRETACION-------------------
#-----------------------------------------------------------


# ==========================================================
# FUNCIÓN AUTOMÁTICA DE INTERPRETACIÓN DEL USO DE HÁBITAT
# ==========================================================
import os

def interpretar_uso_habitat(df, output_folder):
    """
    df debe contener las columnas:
    ['COBERTURA', 'Riqueza_total', 'Especies_exclusivas', '%_Exclusivas']
    
    El resultado es un archivo TXT con la interpretación completa.
    """
    
    # Ordenamientos automáticos
    df_riqueza = df.sort_values("Riqueza_total", ascending=False)
    df_exclusivas = df.sort_values("%_Exclusivas", ascending=False)

    cobertura_max_riqueza = df_riqueza.iloc[0]
    cobertura_min_riqueza = df_riqueza.iloc[-1]

    cobertura_max_exclusivas = df_exclusivas.iloc[0]
    cobertura_min_exclusivas = df_exclusivas.iloc[-1]

    # -------------------------------------------------------------
    # 1. Interpretación general automática
    # -------------------------------------------------------------
    texto_general = f"""
==========================================================
 INTERPRETACIÓN AUTOMÁTICA DEL USO DE HÁBITAT POR COBERTURA
==========================================================

1. --- Riqueza total de especies ---
La cobertura con mayor riqueza registrada es **{cobertura_max_riqueza['COBERTURA']}** 
con **{cobertura_max_riqueza['Riqueza_total']} especies**, indicando una comunidad diversa
y estructuralmente compleja.

La cobertura con menor riqueza es **{cobertura_min_riqueza['COBERTURA']}**, con 
**{cobertura_min_riqueza['Riqueza_total']} especies**, lo que refleja condiciones de menor 
complejidad estructural o efectos de perturbación.

2. --- Especies exclusivas ---
El mayor número de especies exclusivas se observó en **{cobertura_max_exclusivas['COBERTURA']}**,
con **{cobertura_max_exclusivas['Especies_exclusivas']} especies** exclusivas
( {cobertura_max_exclusivas['%_Exclusivas']}% del total ), lo que indica alta singularidad
ecológica y presencia de especies especializadas.

La menor exclusividad se registró en **{cobertura_min_exclusivas['COBERTURA']}**, con 
solo **{cobertura_min_exclusivas['Especies_exclusivas']} especies exclusivas**
( {cobertura_min_exclusivas['%_Exclusivas']}% ), característico de hábitats con
predominio de especies generalistas.

"""
    # -------------------------------------------------------------
    # 2. Descripción por cobertura automática
    # -------------------------------------------------------------
    texto_por_cobertura = "3. --- Descripción ecológica por cobertura ---\n\n"

    for _, row in df.iterrows():
        texto_por_cobertura += f"""• **{row['COBERTURA']}**
   - Riqueza total: {row['Riqueza_total']} especies
   - Especies exclusivas: {row['Especies_exclusivas']} ({row['%_Exclusivas']}%)
   - Interpretación automática: {interpretacion_cobertura(row)}

"""

    # -------------------------------------------------------------
    # 3. Síntesis ecológica automática
    # -------------------------------------------------------------
    texto_sintesis = f"""
4. --- Síntesis ecológica integrada ---

Las coberturas con alta riqueza y alta exclusividad, como **{cobertura_max_riqueza['COBERTURA']}**
y **{cobertura_max_exclusivas['COBERTURA']}**, representan núcleos ecológicos fundamentales
para la conservación debido a su alta singularidad y complejidad funcional.

Las coberturas con valores intermedios mantienen comunidades funcionales estables,
pero requieren monitoreo para evitar procesos de simplificación ecológica.

Las coberturas con menores valores tanto en riqueza como en exclusividad, como 
**{cobertura_min_riqueza['COBERTURA']}**, reflejan sistemas en estados más tempranos
de sucesión ecológica o sometidos a perturbaciones, siendo candidatos ideales para
procesos de restauración ecológica y conectividad del paisaje.

"""

    texto_final = texto_general + texto_por_cobertura + texto_sintesis

    # ==========================================================
    # GUARDAR RESULTADO EN TXT
    # ==========================================================
    os.makedirs(output_folder, exist_ok=True)
    output_path = os.path.join(output_folder, "10.2_Interpretacion_Uso_Habitat.txt")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(texto_final)

    print(f"\n Interpretación generada y guardada en:\n {output_path}\n")
    return texto_final


# ==========================================================
# FUNCIÓN AUXILIAR: interpretación automática por cobertura
# ==========================================================
def interpretacion_cobertura(row):
    """Genera una frase ecológica automática para cada cobertura."""
    r = row["Riqueza_total"]
    e = row["Especies_exclusivas"]
    p = row["%_Exclusivas"]

    # Reglas automáticas
    if p > 30:
        exclus = "alta singularidad ecológica y presencia de especies especializadas"
    elif p > 15:
        exclus = "una comunidad con nivel intermedio de especialización"
    else:
        exclus = "predominio de especies generalistas y baja singularidad ecológica"

    if r > 200:
        ric = "alta complejidad estructural y diversidad elevada"
    elif r > 100:
        ric = "riqueza moderadamente alta, típica de bosques en buen estado"
    elif r > 60:
        ric = "riqueza media, representativa de coberturas secundarias"
    else:
        ric = "riqueza baja, asociada a hábitats simplificados o perturbados"

    return f"presenta {ric}, con {exclus}."

#-----------------------------------------------

interpretar_uso_habitat(
    uso_habitat,
    r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados"
)


























# ============================================
#  TABLA COMPARATIVA POR COBERTURA
#  Individuos – Abundancia – Especies
# ============================================

import pandas as pd

# Asegurar datos válidos
df = Registros.dropna(subset=['COBERTURA', 'ESPECIE', 'INDIVIDUOS']).copy()

# -------------------------------
# 1️⃣ Número total de individuos
# -------------------------------
individuos_por_cobertura = (
    df.groupby('COBERTURA')['INDIVIDUOS']
    .sum()
    .reset_index()
    .rename(columns={'INDIVIDUOS': 'Individuos'})
)

# -------------------------------
# 2️⃣ Abundancia relativa (%)
# -------------------------------
total_individuos = individuos_por_cobertura['Individuos'].sum()

individuos_por_cobertura['Abundancia_%'] = (
    individuos_por_cobertura['Individuos'] / total_individuos * 100
).round(2)

# -------------------------------
# 3️⃣ Número de especies
# -------------------------------
especies_por_cobertura = (
    df.groupby('COBERTURA')['ESPECIE']
    .nunique()
    .reset_index()
    .rename(columns={'ESPECIE': 'Especies'})
)

# -------------------------------
# 4️⃣ Unir todo en una sola tabla
# -------------------------------
tabla_resumen = pd.merge(
    individuos_por_cobertura,
    especies_por_cobertura,
    on='COBERTURA'
)

# Ordenar por número de individuos (descendente)
tabla_resumen = tabla_resumen.sort_values(by='Individuos', ascending=False)

# -------------------------------
# 5️⃣ Mostrar resultado
# -------------------------------
print("\n📊 RESUMEN POR COBERTURA")
print(tabla_resumen)

# -------------------------------
# 6️⃣ Guardar en Excel
# -------------------------------
ruta_salida = os.path.join(
    output_folder,
    "11_Comparacion_Individuos_Abundancia_Especies.xlsx"
)

tabla_resumen.to_excel(ruta_salida, index=False)

print(f"\n✅ Tabla guardada en:\n{ruta_salida}")
