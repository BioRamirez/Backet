# --------------------------------------------------------
# 1. Cargar librerías
# --------------------------------------------------------
import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from tabulate import tabulate
import openpyxl
import re

# --------------------------------------------------------
# 2. Crear carpeta de salida
# --------------------------------------------------------
output_folder = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados"
os.makedirs(output_folder, exist_ok=True)


# --------------------------------------------------------
# 3. Cargar archivo principal de registros
# --------------------------------------------------------
ruta_registros = r"D:\Forestal Consultores\2026\FAUNA\BD\BD_SANROQUE.xlsx"
Registros = pd.read_excel(ruta_registros)

print("\n Registros:")
print(Registros.head())
print(Registros.columns)


# --------------------------------------------------------
# 4. Cargar archivo SIG (puntos + coberturas)
# --------------------------------------------------------
ruta_sig = r"D:\Forestal Consultores\2026\FAUNA\BD\Puntos_Transectos_SIG_SanRoque.xlsx"
datosSIG = pd.read_excel(ruta_sig)

print("\n datos SIG:")
print(datosSIG.head())
print(datosSIG.columns)


# --------------------------------------------------------
# 5. Cargar archivo OPERADOR
# --------------------------------------------------------
ruta_oper = r"D:\Forestal Consultores\2026\FAUNA\BD\Operador_Info_SanRoque.xlsx"
OPER_INFO = pd.read_excel(ruta_oper)

print("\n Operador:")
print(OPER_INFO.head())
print(OPER_INFO.columns)



# --------------------------------------------------------
# 6. Unir SIG a los registros originales
# --------------------------------------------------------
df_registros_sig = Registros.merge(
    datosSIG[["ID", "COBERTURA"]],
    on="ID",
    how="left"
)


# --------------------------------------------------------
# 7. Crear estructura ANLA para MuestreoFaunaResultadosTB
# --------------------------------------------------------
estructura_muestreo_fauna = {
    "EXPEDIENTE": str,
    "PROYECTO": str,
    "N_COBERT": str,
    "NOMENCLAT": "float64",
    "DIVISION": str,
    "CLASE": str,
    "ORDEN": str,
    "FAMILIA": str,
    "GENERO": str,
    "ESPECIE": str,
    "N_COMUN": str,
    "CATEG_CIT": "float64",
    "CATEG_UICN": "float64",
    "CATE_MINIS": "float64",
    "T_DISTRIB": "float64",
    "MIGRACION": "float64",
    "TIPO_MIGR": "Int64",
    "VEDA": "float64",
    "RESOLUCION": str,
    "ENTID_VEDA": "float64",
    "VIGEN_VEDA": "float64",
    "ABUND_ABS": "float64",
    "ABUND_REL": "float64",
    "USO": "Int64",
    "DIETA": "Int64",
    "DISTR_ALT": str,
    "FECHA_IMUE": "datetime64[ns]",
    "FECHA_FMUE": "datetime64[ns]",
    "OBSERV": str
}

df_result = pd.DataFrame(columns=estructura_muestreo_fauna.keys())
df_result = df_result.astype(estructura_muestreo_fauna)



# --------------------------------------------------------
# 8. Agregar los registros por especie + cobertura
# --------------------------------------------------------

print(OPER_INFO.columns)
print(datosSIG.columns)
print(df_result.columns)
print(df_result)



print("Columnas de df_registros_sig:")
print(df_registros_sig.columns.tolist())

df_registros_sig[["COBERTURA_x", "COBERTURA_y"]].head()


# 1. Eliminar la columna duplicada
df_registros_sig.drop(columns=["COBERTURA_y"], inplace=True)

# 2. Renombrar la columna principal
df_registros_sig.rename(columns={"COBERTURA_x": "COBERTURA"}, inplace=True)

# 3. Mostrar estructura actualizada
print(df_registros_sig.columns.tolist())


# --------------------------------------------------------
# AGRUPAR y CALCULAR ABUNDANCIAS (corrección al uso de .size())
# --------------------------------------------------------
# Agrupamos por las columnas relevantes y:
# - ABUND_ABS: suma real de INDIVIDUOS
# - FECHA_IMUE: primera fecha (min)
# - FECHA_FMUE: última fecha (max)
# Luego calculamos abundancia relativa total y relativa por cobertura.
# --------------------------------------------------------

df_agg = (
    df_registros_sig
    .groupby([
        "COBERTURA",
        "DIVISION", "CLASE", "Orden", "Familia",
        "Genero", "Especie", "N. comun",
        "CITES", "IUCN", "MADS (Resol 0126)",  "Dist_Geo",
        "Tipo_Migra", "Uso", "Gremio", "Dist_Alt", "OBSERVACION", "Veda"
    ], dropna=False)
    .agg(
        ABUND_ABS=("INDIVIDUOS", "sum"),      # <-- suma de individuos (abundancia absoluta)
        FECHA_IMUE=("FECHA", "min"),
        FECHA_FMUE=("FECHA", "max")
    )
    .reset_index()
)

# --------------------------------------------------------
# Abundancia relativa sobre el total del dataset
# --------------------------------------------------------
total_individuos = df_agg["ABUND_ABS"].sum()
if total_individuos == 0:
    df_agg["ABUND_REL_TOTAL"] = 0.0
else:
    df_agg["ABUND_REL_TOTAL"] = df_agg["ABUND_ABS"] / total_individuos * 100

# --------------------------------------------------------
# Abundancia relativa dentro de cada COBERTURA
# --------------------------------------------------------
# Calculamos el total por cobertura y luego la proporción por fila
total_por_cobertura = df_agg.groupby("COBERTURA")["ABUND_ABS"].transform("sum")
# Evitar división por cero
df_agg["ABUND_REL_POR_COBERT"] = np.where(
    total_por_cobertura > 0,
    df_agg["ABUND_ABS"] / total_por_cobertura,
    0.0
)

# --------------------------------------------------------
# Revisar resultado
# --------------------------------------------------------
print("\nPreview de df_agg con abundancias (absoluta y relativas):")
print(df_agg.head())




print(df_agg.columns)
print(df_result)
print(df_result.columns)



#######################------------------

# ===========================================================
# 1. Asegurar que df_result tenga las MISMAS FILAS que df_agg
#    pero SIN perder su estructura de columnas
# ===========================================================

if len(df_result) == 0:    # si está vacío pero con columnas
    # Crear tantas filas como df_agg
    df_result = pd.DataFrame(
        {col: [None] * len(df_agg) for col in df_result.columns}
    )

# ===========================================================
# 2. Mapeo de columnas df_agg → df_result
# ===========================================================
map_cols = {
    "COBERTURA": "N_COBERT",
    "DIVISION": "DIVISION",
    "CLASE": "CLASE",
    "Orden": "ORDEN",
    "Familia": "FAMILIA",
    "Genero": "GENERO",
    "Especie": "ESPECIE",
    "N. comun": "N_COMUN",
    "Veda": "VEDA",
    "ABUND_ABS": "ABUND_ABS",
    "FECHA_IMUE": "FECHA_IMUE",
    "FECHA_FMUE": "FECHA_FMUE",
    "OBSERVACION": "OBSERV",
    "Dist_Alt": "DISTR_ALT",
    "ABUND_REL_TOTAL": "ABUND_REL"
}

# ===========================================================
# 3. Rellenar columnas existentes SIN borrar nada
# ===========================================================
for col_agg, col_res in map_cols.items():

    # Validar que la columna exista en df_result
    if col_res not in df_result.columns:
        df_result[col_res] = None

    # Rellenar valores
    df_result[col_res] = df_agg[col_agg].values

# ===========================================================
# 4. Verificar
# ===========================================================
print(df_result.head())
print(df_result.shape)


#---------------------------

















df_agg["Dist_Geo"].unique()












































# ---------------------------------------------------------
# Rellenar/actualizar df_result a partir de df_agg
# - no limpia dataframes
# - reemplaza siempre los valores existentes
# - usa descripciones (no códigos)
# - aplica misma conversión para CATE_MINIS (MADS)
# - rellena RESOLUCION/ENTID_VEDA/VIGEN_VEDA según VEDA
# ---------------------------------------------------------
import pandas as pd
import numpy as np

# -------------------------
# Diccionarios (descripciones)
# -------------------------
MAP_CITES = {
    "Apendice I": "Apendice I",
    "Apendice II": "Apendice II",
    "Apendice III": "Apendice III",
    "No aplica": "No aplica",
}

MAP_IUCN = {
    "Preocupación Menor (LC)": "Preocupación Menor (LC)",
    "Preocupacin Menor (LC)": "Preocupación Menor (LC)",  # corregir sin tilde
    "Casi Amenazado (NT)": "Casi Amenazada (NT)",
    "Casi Amenazada (NT)": "Casi Amenazada (NT)",
    "Vulnerable (VU)": "Vulnerable (VU)",
    "Peligro (EN)": "Peligro (EN)",
    "Peligro Crítico (CR)": "Peligro Crítico (CR)",
    "Extinto en estado silvestre (EW)": "Extinto en estado silvestre (EW)",
    "Extinto (EX)": "Extinto (EX)",
    "Datos insuficientes (DD)": "Datos insuficientes (DD)",
    "No Evaluado (NE)": "No Evaluado (NE)",
    "No aplica": "No aplica"
}

# Usamos la misma lógica (texto) para CATE_MINIS (MADS)
MAP_MADS = {
    "NL": "NL",
    "Preocupación Menor (LC)": "Preocupación Menor (LC)",
    "Preocupacin Menor (LC)": "Preocupación Menor (LC)",
    "Vulnerable (VU)": "Vulnerable (VU)",
    "Peligro Crítico (CR)": "Peligro Crítico (CR)",
    "No aplica": "No aplica"
}

# Uso: normalizar a descripciones del dominio (Dom_Uso_Fauna)
MAP_USO = {
    "Sin uso conocido": "Otro",
    "Medicinal": "Otro",
    "Mascotas": "Mascotas",
    "Mascota": "Mascotas",
    "Cultural": "Uso Cultural",
    "Uso Cultural": "Uso Cultural",
    "Subsistencia": "Subsistencia",
    "Otro": "Otro",
    "Mascotas, Subsistencia": "Mascotas",
    "Subsistencia, Mascotas": "Mascotas",
    "Medicinal, Cultural": "Uso Cultural",
    "Cultural, Mascotas": "Mascotas"
}

# Gremio -> Dieta (texto)
MAP_GREMIO = {
    "Nectarívoro": "Otro",
    "Frugívoro": "Frugívoro",
    "Granívoro": "Granívoro",
    "Insectívoro": "Insectívoro",
    "Omnívoro": "Omnívoro",
    "Carnívoro": "Carnívoro",
    "Carroñero": "Otro",
    "Herbivoro": "Herbívoro",
    "Herbívoro": "Herbívoro"
}

# Tipo de migración: mapeo a descripciones del dominio
MAP_TIPO_MIGRA = {
    "Res": None,
    "Lat": "Latitudinal",
    "Latitudinal": "Latitudinal",
    "Lat-Trans": "Latitudinal",
    "Alt-Loc": "Altitudinal",
    "Lat-Alt-Trans-Loc": "Altitudinal",  # decisión: agrupar en altitudinal
    "Estacional": "Estacional",
    "Loc": "Nomadismo",
    "Nomadismo": "Nomadismo"
}

# Veda: mantener texto tal cual (tu df_agg tiene 'No')
# Las columnas RESOLUCION/ENTID_VEDA/VIGEN_VEDA se llenarán según regla
# -------------------------

# -------------------------
# Asegurar filas en df_result (pero sin perder columnas)
# -------------------------
n_rows = len(df_agg)
if len(df_result) == 0:
    # crear n_rows filas con None manteniendo columnas de df_result
    df_result = pd.DataFrame({col: [None] * n_rows for col in df_result.columns})

# -------------------------
# Mapeo directo de columnas entre df_agg --> df_result
# (reemplaza siempre)
# -------------------------
col_map = {
    "COBERTURA": "N_COBERT",
    "DIVISION": "DIVISION",
    "CLASE": "CLASE",
    "Orden": "ORDEN",
    "Familia": "FAMILIA",
    "Genero": "GENERO",
    "Especie": "ESPECIE",
    "N. comun": "N_COMUN",
    "Veda": "VEDA",
    "ABUND_ABS": "ABUND_ABS",
    "FECHA_IMUE": "FECHA_IMUE",
    "FECHA_FMUE": "FECHA_FMUE",
    # observación ya está en df_agg bajo 'OBSERVACION' -> queremos 'OBSERV'
    "OBSERVACION": "OBSERV",
    # campos para conversión con diccionarios
    "CITES": "CATEG_CIT",       # texto (Apendice ...)
    "IUCN": "CATEG_UICN",       # texto (Preocupación Menor (LC), ...)
    "MADS (Resol 0126)": "CATE_MINIS",
    "Tipo_Migra": "TIPO_MIGR",
    "Uso": "USO",
    "Gremio": "DIETA",
    "Dist_Alt": "DISTR_ALT"
}

# Recorremos columnas y rellenamos (reemplazando siempre)
for src_col, dst_col in col_map.items():
    # Si columna fuente no existe en df_agg, saltar
    if src_col not in df_agg.columns:
        # print(f"Fuente ausente en df_agg: {src_col} -> se salta")
        continue

    # Obtener valores de df_agg como lista (en el mismo orden de índices)
    vals = df_agg[src_col].tolist()

    # Si df_result no tiene la columna, crearla
    if dst_col not in df_result.columns:
        df_result[dst_col] = [None] * len(df_result)

    # Asignar (reemplazando todo)
    df_result.loc[:, dst_col] = vals

# -------------------------
# Aplicar diccionarios (reemplazar siempre)
# -------------------------
# CITES -> CATEG_CIT (texto)
if "CATEG_CIT" in df_result.columns:
    df_result["CATEG_CIT"] = df_result["CATEG_CIT"].apply(
        lambda v: MAP_CITES.get(v, None) if pd.notna(v) and str(v) != "nan" else None
    )

# IUCN -> CATEG_UICN (texto normalizado)
if "CATEG_UICN" in df_result.columns:
    df_result["CATEG_UICN"] = df_result["CATEG_UICN"].apply(
        lambda v: MAP_IUCN.get(v, None) if pd.notna(v) and str(v) != "nan" else None
    )

# MADS -> CATE_MINIS (misma conversión que IUCN-style: texto estandarizado)
if "CATE_MINIS" in df_result.columns:
    df_result["CATE_MINIS"] = df_result["CATE_MINIS"].apply(
        lambda v: MAP_MADS.get(v, None) if pd.notna(v) and str(v) != "nan" else None
    )

# Uso
if "USO" in df_result.columns:
    df_result["USO"] = df_result["USO"].apply(
        lambda v: MAP_USO.get(v, v) if pd.notna(v) and str(v) != "nan" else None
    )

# Gremio -> DIETA
if "DIETA" in df_result.columns:
    df_result["DIETA"] = df_result["DIETA"].apply(
        lambda v: MAP_GREMIO.get(v, v) if pd.notna(v) and str(v) != "nan" else None
    )

# Tipo migracion
if "TIPO_MIGR" in df_result.columns:
    df_result["TIPO_MIGR"] = df_result["TIPO_MIGR"].apply(
        lambda v: MAP_TIPO_MIGRA.get(v, v) if pd.notna(v) and str(v) != "nan" else None
    )

# -------------------------
# ABUND_REL: calcular en porcentaje (0-100) a partir de ABUND_ABS
# -------------------------
if "ABUND_ABS" in df_result.columns:
    total_individuos = df_result["ABUND_ABS"].astype(float).sum()
    if total_individuos == 0:
        df_result["ABUND_REL"] = 0.0
    else:
        df_result["ABUND_REL"] = (df_result["ABUND_ABS"].astype(float) / total_individuos) * 100

# -------------------------
# RESOLUCION / ENTID_VEDA / VIGEN_VEDA según VEDA
# - si VEDA == 'No' -> dejar vacíos ""
# - si VEDA es distinto de 'No' y no nulo -> "RELLENAR"
# -------------------------
for idx in df_result.index:
    v = df_result.at[idx, "VEDA"] if "VEDA" in df_result.columns else None
    # Normalizar string si viene con NaN
    if pd.isna(v) or str(v).strip() == "" or str(v).lower() == "nan":
        # no hay veda: dejar vacíos
        if "RESOLUCION" in df_result.columns:
            df_result.at[idx, "RESOLUCION"] = ""
        if "ENTID_VEDA" in df_result.columns:
            df_result.at[idx, "ENTID_VEDA"] = ""
        if "VIGEN_VEDA" in df_result.columns:
            df_result.at[idx, "VIGEN_VEDA"] = ""
    else:
        # hay algún valor en VEDA; si el texto es exactamente 'No' considerar como no
        if str(v).strip().lower() == "no":
            if "RESOLUCION" in df_result.columns:
                df_result.at[idx, "RESOLUCION"] = ""
            if "ENTID_VEDA" in df_result.columns:
                df_result.at[idx, "ENTID_VEDA"] = ""
            if "VIGEN_VEDA" in df_result.columns:
                df_result.at[idx, "VIGEN_VEDA"] = ""
        else:
            # VEDA != 'No' => marcar para rellenar
            if "RESOLUCION" in df_result.columns:
                df_result.at[idx, "RESOLUCION"] = "RELLENAR"
            if "ENTID_VEDA" in df_result.columns:
                df_result.at[idx, "ENTID_VEDA"] = "RELLENAR"
            if "VIGEN_VEDA" in df_result.columns:
                df_result.at[idx, "VIGEN_VEDA"] = "RELLENAR"

# -------------------------
# Resultado: mostrar resumen
# -------------------------
print("\nRelleno completado. Resumen rápido:")
to_show = ["N_COBERT","DIVISION","CLASE","ORDEN","FAMILIA","GENERO","ESPECIE","N_COMUN",
           "VEDA","RESOLUCION","ENTID_VEDA","VIGEN_VEDA","ABUND_ABS","ABUND_REL","CATEG_CIT","CATEG_UICN","CATE_MINIS"]
present = [c for c in to_show if c in df_result.columns]
print(df_result[present].head(10))
print("\nShape df_result:", df_result.shape)

























codigo_a_cobertura = {
    "1": "TERRITORIOS ARTIFICIALIZADOS",
    "11": "Zonas urbanizadas",
    "111": "Tejido urbano continuo",
    "112": "Tejido urbano discontinuo",
    "12": "Zonas industriales o comerciales y redes de comunicación",
    "121": "Zonas industriales o comerciales",
    "1211": "Zonas industriales",
    "1212": "Zonas comerciales",
    "122": "Red vial, ferroviaria y terrenos asociados",
    "1221": "Red vial y terrenos asociados",
    "1222": "Red ferroviaria y terrenos asociados",
    "123": "Zonas portuarias",
    "1231": "Zonas portuarias fluviales",
    "1232": "Zonas portuarias marítimas",
    "124": "Aeropuertos",
    "1241": "Aeropuerto con infraestructura asociada",
    "1242": "Aeropuerto sin infraestructura asociada",
    "125": "Obras hidráulicas",
    "13": "Zonas de extracción minera y escombreras",
    "131": "Zonas de extracción minera",
    "1311": "Otras explotaciones mineras",
    "1312": "Explotación de hidrocarburos",
    "1313": "Explotación de carbón",
    "1314": "Explotación de oro",
    "1315": "Explotación de materiales de construcción",
    "1316": "Explotación de sal",
    "1317": "Explotación de esmeraldas",
    "132": "Zonas de disposición de residuos",
    "1321": "Otros sitios de disposición de residuos a cielo abierto",
    "1322": "Escombreras",
    "1323": "Vertederos",
    "1324": "Relleno sanitario",
    "14": "Zonas verdes artificializadas, no agrícolas",
    "141": "Zonas verdes urbanas",
    "1411": "Otras zonas verdes urbanas",
    "1412": "Parques cementerios",
    "1413": "Jardines botánicos",
    "1414": "Zoológicos",
    "1415": "Parques urbanos",
    "1416": "Rondas de cuerpos de agua de zonas urbanas",
    "142": "Instalaciones recreativas",
    "1421": "Áreas culturales",
    "1422": "Áreas deportivas",
    "1423": "Áreas turísticas",

    "2": "TERRITORIOS AGRÍCOLAS",
    "21": "Cultivos transitorios",
    "211": "Otros cultivos transitorios",
    "212": "Cereales",
    "2121": "Arroz",
    "2122": "Maíz",
    "2123": "Sorgo",
    "2124": "Cebada",
    "2125": "Trigo",
    "213": "Oleaginosas y leguminosas",
    "2131": "Algodón",
    "2132": "Ajonjolí",
    "2133": "Fríjol",
    "2134": "Soya",
    "2135": "Maní",
    "214": "Hortalizas",
    "2141": "Cebolla",
    "2142": "Zanahoria",
    "2143": "Remolacha",
    "215": "Tubérculos",
    "2151": "Papa",
    "2152": "Yuca",

    "22": "Cultivos permanentes",
    "221": "Cultivos permanentes herbáceos",
    "2211": "Otros cultivos permanentes herbáceos",
    "2212": "Caña",
    "22121": "Caña de azúcar",
    "22122": "Caña panelera",
    "2213": "Plátano y banano",
    "2214": "Tabaco",
    "2215": "Papaya",
    "2216": "Amapola",
    "222": "Cultivos permanentes arbustivos",
    "2221": "Otros cultivos permanentes arbustivos",
    "2222": "Café",
    "2223": "Cacao",
    "2224": "Viñedos",
    "2225": "Coca",
    "223": "Cultivos permanentes arbóreos",
    "2231": "Otros cultivos permanentes arbóreos",
    "2232": "Palma de aceite",
    "2233": "Cítricos",
    "2234": "Mango",
    "224": "Cultivos agroforestales",
    "2241": "Pastos y árboles plantados",
    "2242": "Cultivos y árboles plantados",
    "225": "Cultivos confinados",

    "23": "Pastos",
    "231": "Pastos limpios",
    "232": "Pastos arbolados",
    "233": "Pastos enmalezados",

    "24": "Áreas agrícolas heterogéneas",
    "241": "Mosaico de cultivos",
    "242": "Mosaico de pastos y cultivos",
    "243": "Mosaico de cultivos, pastos y espacios naturales",
    "244": "Mosaico de pastos con espacios naturales",
    "245": "Mosaico de cultivos y espacios naturales",

    "3": "BOSQUES Y ÁREAS SEMINATURALES",
    "31": "Bosques",
    "311": "Bosque denso",
    "3111": "Bosque denso alto",
    "31111": "Bosque denso alto de tierra firme",
    "31112": "Bosque denso alto inundable",
    "311121": "Bosque denso alto inundable heterogéneo",
    "311122": "Manglar denso alto",
    "311123": "Palmares",
    "3112": "Bosque denso bajo",
    "31121": "Bosque denso bajo de tierra firme",
    "311211": "Caatingas",
    "311212": "Bosque enano del Caribe",
    "311213": "Bosque denso altoandino",
    "31122": "Bosque denso bajo inundable",

    "312": "Bosque abierto",
    "3121": "Bosque abierto alto",
    "31211": "Bosque abierto alto de tierra firme",
    "31212": "Bosque abierto alto inundable",
    "3122": "Bosque abierto bajo",
    "31221": "Bosque abierto bajo de tierra firme",
    "31222": "Bosque abierto bajo inundable",

    "313": "Bosque fragmentado",
    "3131": "Bosque fragmentado con pastos y cultivos",
    "3132": "Bosque fragmentado con vegetación secundaria",

    "314": "Bosque de galería y ripario",
    "315": "Plantación forestal",
    "3151": "Plantación de coníferas",
    "3152": "Plantación de latifoliadas",

    "32": "Áreas con vegetación herbácea y/o arbustiva",
    "3211": "Herbazal denso",
    "32111": "Herbazal denso de tierra firme",
    "321111": "Herbazal denso de tierra firme no arbolado",
    "3211111": "Herbazal denso alto de tierra firme no arbolado",
    "3211112": "Herbazal denso bajo de tierra firme no arbolado",
    "321112": "Herbazal denso de tierra firme arbolado",
    "321113": "Herbazal denso de tierra firme con arbustos",
    "32112": "Herbazal denso inundable",
    "321121": "Herbazal denso inundable no arbolado",
    "321122": "Herbazal denso inundable arbolado",
    "321123": "Arracachal",
    "321124": "Helechal",

    "3212": "Herbazal abierto",
    "32121": "Herbazal abierto arenoso",
    "32122": "Herbazal abierto rocoso",

    "3221": "Arbustal denso",
    "3222": "Arbustal abierto",
    "32221": "Arbustal abierto esclerófilo",
    "32222": "Arbustal abierto mesófilo",

    "323": "Vegetación secundaria o en transición",
    "3231": "Vegetación secundaria alta",
    "3232": "Vegetación secundaria baja",

    "33": "Áreas abiertas sin o con poca vegetación",
    "331": "Zonas arenosas naturales",
    "3311": "Playas",
    "3312": "Arenales",
    "3313": "Campos de dunas",
    "332": "Afloramientos rocosos",
    "333": "Tierras desnudas y degradadas",
    "334": "Zonas quemadas",
    "335": "Zonas glaciares y nivales",
    "3351": "Zonas glaciares",
    "3352": "Zonas nivales",

    "4": "ÁREAS HÚMEDAS",
    "41": "Áreas húmedas continentales",
    "411": "Zonas pantanosas",
    "412": "Turberas",
    "413": "Vegetación acuática sobre cuerpos de agua",

    "42": "Áreas húmedas costeras",
    "421": "Pantanos costeros",
    "422": "Salitral",
    "423": "Sedimentos expuestos en bajamar",

    "5": "SUPERFICIES DE AGUA",
    "51": "Aguas continentales",
    "511": "Ríos (50 m)",
    "512": "Lagunas, lagos y ciénagas naturales",
    "513": "Canales",
    "514": "Cuerpos de agua artificiales",
    "5141": "Embalses",
    "5142": "Lagunas de oxidación",
    "5143": "Estanques para acuicultura continental",

    "52": "Aguas marítimas",
    "521": "Lagunas costeras",
    "522": "Mares y océanos",
    "5221": "Otros fondos",
    "5222": "Fondos coralinos someros",
    "5223": "Praderas de pastos marinos someras",
    "5224": "Fondos someros de arenas y cascajo",
    "523": "Estanques para acuicultura marina"
}




























# =============================
# Rellenar EXPEDIENTE, PROYECTO, NOMENCLAT, T_DISTRIB, MIGRACION
# =============================
import unicodedata

# -----------------------------
# Helper: normalizar texto (quita tildes, minuscula, strip)
# -----------------------------
def norm(text):
    if pd.isna(text):
        return ""
    if not isinstance(text, str):
        text = str(text)
    text = text.strip()
    # quitar acentos
    text = unicodedata.normalize('NFKD', text).encode('ASCII', 'ignore').decode('ASCII')
    return text.lower()

# -----------------------------
# 1) Asegurar filas en df_result sin perder columnas
# -----------------------------
n_rows = len(df_agg)
if len(df_result) == 0:
    df_result = pd.DataFrame({col: [None] * n_rows for col in df_result.columns})

# -----------------------------
# 2) EXPEDIENTE y PROYECTO desde OPER_INFO (primera fila)
# -----------------------------
if 'EXPEDIENTE' in df_result.columns and 'PROYECTO' in df_result.columns:
    if 'EXPEDIENTE' in OPER_INFO.columns and 'PROYECTO' in OPER_INFO.columns:
        exped = OPER_INFO['EXPEDIENTE'].iloc[0]
        proyecto = OPER_INFO['PROYECTO'].iloc[0]
    else:
        exped = None
        proyecto = None

    # Reemplazar siempre
    df_result['EXPEDIENTE'] = [exped] * len(df_result)
    df_result['PROYECTO'] = [proyecto] * len(df_result)

# -----------------------------
# 3) Preparar reverse mapping codigo_a_cobertura: descripcion -> codigo
#    (normalizamos descripciones para facilitar la busqueda)
# -----------------------------
codigo_a_cobertura_local = {
    # (usa el dict que ya tienes; lo incluyo parcialmente para ejemplo)
    # Asegúrate de tener todo el dict completo en tu entorno con el nombre codigo_a_cobertura
    **codigo_a_cobertura
}

# crear mapping descripcion_normalizada -> codigo (si hay duplicados, mantiene el primero)
desc_to_code = {}
for code, desc in codigo_a_cobertura_local.items():
    if desc is None:
        continue
    key = norm(desc)
    if key not in desc_to_code:
        desc_to_code[key] = code

# -----------------------------
# 4) Función para obtener NOMENCLAT desde N_COBERT
#    Estrategia:
#      - intento coincidencia exacta (normalizada)
#      - si no, intento buscar si N_COBERT está contenido en alguna descripcion (y viceversa)
#      - si no se encuentra, devuelve None (puedes cambiar a "")
# -----------------------------
def find_nomenclat_from_cobertura(cob_text):
    nc = norm(cob_text)
    if nc == "":
        return None
    # coincidencia exacta
    if nc in desc_to_code:
        return desc_to_code[nc]
    # coincidencia parcial: buscar en keys si alguna contiene nc
    for k, code in desc_to_code.items():
        if nc in k:
            return code
    # intentar que la descripcion del mapa esté contenida en nc
    for k, code in desc_to_code.items():
        if k in nc:
            return code
    return None

# -----------------------------
# 5) Rellenar N_COBERT (si no existe en df_result), luego NOMENCLAT
#    Observación: df_agg tiene 'COBERTURA' -> df_result 'N_COBERT' (ya hecho anteriormente),
#    pero lo validamos aquí y lo reemplazamos desde df_agg (si existe)
# -----------------------------
if 'COBERTURA' in df_agg.columns:
    # asegurarnos de asignar N_COBERT (reemplazando)
    df_result['N_COBERT'] = df_agg['COBERTURA'].values

# ahora crear/llenar NOMENCLAT
if 'NOMENCLAT' in df_result.columns:
    nomenclat_vals = []
    for i in df_result.index:
        cob = df_result.at[i, 'N_COBERT'] if 'N_COBERT' in df_result.columns else None
        code_found = find_nomenclat_from_cobertura(cob)
        nomenclat_vals.append(code_found)
    df_result['NOMENCLAT'] = nomenclat_vals

# -----------------------------
import pandas as pd
import numpy as np
import unicodedata

# Helper (si ya existe, es inofensivo volver a definir)
def norm(x):
    if pd.isna(x):
        return ""
    x = str(x).strip().lower()
    x = unicodedata.normalize("NFD", x)
    x = "".join(c for c in x if unicodedata.category(c) != "Mn")
    return x

# -------------------------
# 1) Re-mapeo T_DISTRIB: asegurar que 'none' -> "" y que use Dist_Geo correctamente
# -------------------------
# Reconstruimos lookup por seguridad (asegúrate que este dict coincida con tu lógica)
distrib_dict = {
    "cosmopolita": "Cosmopolita",
    "restringida": "Restringida",
    "casi endemica": "Casi endémica",
    "endemica": "Endémica",
    "neotropical": "",                    # pasar a vacio
    "nearctica, neotropical": "",         # vacio
    "introducida": ""                     # vacio
}
distrib_lookup = {k: v for k, v in distrib_dict.items()}

if "Dist_Geo" in df_agg.columns and "T_DISTRIB" in df_result.columns:
    tdist_list = []
    for i in df_agg.index:
        raw = df_agg.at[i, "Dist_Geo"]
        if pd.isna(raw) or str(raw).strip() == "":
            mapped = ""
        else:
            mapped = distrib_lookup.get(norm(raw), "")
        tdist_list.append(mapped)
    df_result["T_DISTRIB"] = tdist_list

# Si previamente quedaron valores literales 'none' por ejecuciones antiguas, limpiarlos:
if "T_DISTRIB" in df_result.columns:
    df_result["T_DISTRIB"] = df_result["T_DISTRIB"].replace({"none": ""})
    df_result["T_DISTRIB"] = df_result["T_DISTRIB"].fillna("")

# -------------------------
import pandas as pd
import numpy as np

# -----------------------
# MIGRACION basado SOLO en Tipo_Migra
# -----------------------

# Diccionario oficial
MAP_TIPO_MIGRA = {
    "Res": None,                   # residente → no migra
    "Lat": "Latitudinal",
    "Latitudinal": "Latitudinal",
    "Lat-Trans": "Latitudinal",
    "Alt-Loc": "Altitudinal",
    "Lat-Alt-Trans-Loc": "Altitudinal",
    "Estacional": "Estacional",
    "Loc": "Nomadismo",
    "Nomadismo": "Nomadismo",
    None: None,
    "": None,
    np.nan: None
}

# si la columna existe en df_agg:
if "Tipo_Migra" in df_agg.columns:

    # tomar la columna
    tipo_series = df_agg["Tipo_Migra"].astype(object)

    # 1) convertir a tipo migratorio ANLA usando el diccionario
    tipo_mapeado = tipo_series.map(MAP_TIPO_MIGRA)

    # asignarlo a df_result por posición
    df_result["TIPO_MIGR"] = tipo_mapeado.fillna("")

    # 2) calcular MIGRACION según TIPO_MIGR
    def decidir_migr(val):
        if val is None:
            return "No"
        s = str(val).strip()
        if s == "":
            return "No"
        return "Sí"

    df_result["MIGRACION"] = df_result["TIPO_MIGR"].apply(decidir_migr)

else:
    # si no existe Tipo_Migra en df_agg, nos basamos en df_result
    tipo_series = df_result["TIPO_MIGR"].astype(object).fillna("")

    def decidir_migr2(val):
        s = str(val).strip()
        if s == "":
            return "No"
        return "Sí"

    df_result["MIGRACION"] = tipo_series.apply(decidir_migr2)

# -------------------------
# 3) Resumen de control
# -------------------------
print("Resumen MIGRACION (value_counts):")
print(df_result["MIGRACION"].value_counts(dropna=False))

print("\nPrimeras filas de TIPO_MIGR y MIGRACION:")
cols_show = [c for c in ["TIPO_MIGR", "MIGRACION"] if c in df_result.columns]
print(df_result[cols_show].head(12))



# -------------------------
# LIMPIAR FECHA_IMUE y FECHA_FMUE PARA QUITAR LA HORA
# -------------------------

# Convertir a datetime (por si acaso) y luego formatear solo la fecha
for col in ["FECHA_IMUE", "FECHA_FMUE"]:
    if col in df_result.columns:
        df_result[col] = pd.to_datetime(df_result[col], errors="coerce").dt.strftime("%Y-%m-%d")





#-------------------Guardar df_punto_muestreo para inspección----------------------


# Guardar el resultado en un nuevo archivo Excel

# Archivo final
output_file = os.path.join(output_folder, "14.3_M_fauna_Result_TB.xlsx")

# Guardar resultado
df_result.to_excel(output_file, index=False)
print(f"\n Archivo guardado en: {output_file}")


#---------------------------------- Reparar y formatear archivo de M_fauna_Result_TB -----------------------
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import os

# --- Rutas ---
ruta_original = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\14.3_M_fauna_Result_TB.xlsx"
ruta_limpia = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\14.3_M_fauna_Result_TB.xlsx"

# --- Verificar existencia ---
if not os.path.exists(ruta_original):
    raise FileNotFoundError(f"⚠️ No se encontró el archivo: {ruta_original}")

# --- Leer archivo dañado con pandas ---
try:
    df = pd.read_excel(ruta_original)
    print(" Archivo leído correctamente con pandas.")
except Exception as e:
    raise RuntimeError(f" No se pudo leer el archivo: {e}")

# --- Reescribir el archivo limpio ---
df.to_excel(ruta_limpia, index=False)
print(f" Archivo reparado y guardado como:\n{ruta_limpia}")

# --- Aplicar formato con openpyxl ---
from openpyxl import load_workbook

wb = load_workbook(ruta_limpia)
ws = wb.active

# --- Estilos base ---
header_fill = PatternFill(start_color='BFD8B8', end_color='BFD8B8', fill_type='solid')
header_font = Font(bold=True, color='000000', name='Calibri')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# --- Aplicar formato y reemplazar vacíos ---
for row in ws.iter_rows():
    for cell in row:
        if cell.value is None or str(cell.value).strip() == '':
            cell.value = '-'
        cell.alignment = center_align
        cell.border = thin_border

# --- Encabezado ---
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align

# --- Ajustar ancho de columnas ---
for col in ws.columns:
    max_length = 0
    column = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            length = len(str(cell.value))
            if length > max_length:
                max_length = length
    ws.column_dimensions[column].width = max_length + 3

# --- Ajustar altura de filas ---
for row in ws.iter_rows():
    ws.row_dimensions[row[0].row].height = 18

# --- Guardar cambios ---
wb.save(ruta_limpia)
print(f' Archivo formateado y reparado correctamente:\n{ruta_limpia}')











