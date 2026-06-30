import pandas as pd
import openpyxl

# Ruta del archivo
ruta = r"D:\Forestal Consultores\2026\FAUNA\BD\BD_SANROQUE.xlsx"
# Leer el archivo Excel
Registros = pd.read_excel(ruta)



# =========================================================
# FILTRAR UNA SOLA CLASE
# =========================================================

grupo = "MAMIFEROS"   # AVES, MAMIFEROS, REPTILES, ANFIBIOS etc.

Registros = Registros[
    Registros["CLASE"].astype(str).str.upper() == grupo.upper()
].copy()

# Reiniciar índice
Registros.reset_index(drop=True, inplace=True)

# =========================================================
# VERIFICACIÓN
# =========================================================

print(f"\nGrupo seleccionado: {grupo}")
print(f"Número de registros: {len(Registros)}")

print("\nPrimeras filas:")
print(Registros.head())



# =========================================================
# TABLA DE ABUNDANCIA POR INTERVALOS DE DÍAS
# =========================================================

import pandas as pd

# ---------------------------------------------------------
# CONFIGURACIÓN MANUAL
# ---------------------------------------------------------

dias_intervalo = 1   # Cambiar a 1, 5, 10, 15, 30, etc.

# ---------------------------------------------------------
# ASEGURAR FORMATO FECHA
# ---------------------------------------------------------

Registros['FECHA'] = pd.to_datetime(
    Registros['FECHA'],
    errors='coerce'
)

# ---------------------------------------------------------
# ORDENAR POR FECHA
# ---------------------------------------------------------

Registros = Registros.sort_values('FECHA')

# ---------------------------------------------------------
# CREAR INTERVALOS PERSONALIZADOS
# ---------------------------------------------------------

fecha_inicio = Registros['FECHA'].min()

Registros['GRUPO_DIAS'] = (
    (Registros['FECHA'] - fecha_inicio).dt.days // dias_intervalo
) + 1

# ---------------------------------------------------------
# NOMBRE DEL RANGO
# ---------------------------------------------------------

Registros['RANGO_FECHA'] = Registros.groupby('GRUPO_DIAS')['FECHA'].transform(
    lambda x: f"{x.min().strftime('%Y-%m-%d')} a {x.max().strftime('%Y-%m-%d')}"
)

# ---------------------------------------------------------
# TABLA DE ABUNDANCIA
# ---------------------------------------------------------

tabla_abundancia = (
    Registros
    .groupby(['ESPECIE', 'RANGO_FECHA'])['INDIVIDUOS']
    .sum()
    .unstack(fill_value=0)
)

# ---------------------------------------------------------
# EXPORTAR
# ---------------------------------------------------------

ruta_salida = (
    f'D:/CORPONOR 2025/Backet/python_Proyect/Resultados/'
    f'3_Tabla_Abundancia_Semanal.xlsx'
)

with pd.ExcelWriter(ruta_salida, engine='openpyxl') as writer:

    tabla_abundancia.to_excel(
        writer,
        sheet_name='Tabla_Abundancia'
    )

# ---------------------------------------------------------
# VERIFICACIÓN
# ---------------------------------------------------------

print(f'\nTabla creada por intervalos de {dias_intervalo} días')

print('\nRuta:')
print(ruta_salida)

print('\nVista previa:')
print(tabla_abundancia.head())

#------------------Fin Tabla de Abundancia Semanal------------------#
#-----------------Dar formato al archivo de Tabla_Abundancia_Semanal.xlsx------------------
#---------------------------------- Reparar y formatear archivo de Tabla_Abundancia_Semanal -----------------------
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import os

# --- Rutas ---
ruta_original = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\3_Tabla_Abundancia_Semanal.xlsx"
ruta_limpia = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\3_Tabla_Abundancia_Semanal.xlsx"

# --- Verificar existencia ---
if not os.path.exists(ruta_original):
    raise FileNotFoundError(f"⚠️ No se encontró el archivo: {ruta_original}")

# --- Leer archivo dañado con pandas ---
try:
    df = pd.read_excel(ruta_original)
    print(" Archivo leído correctamente con pandas.")
except Exception as e:
    raise RuntimeError(f" No se pudo leer el archivo: {e}")

# --- Reescribir el archivo limpio ---
df.to_excel(ruta_limpia, index=False)
print(f" Archivo reparado y guardado como:\n{ruta_limpia}")

# --- Aplicar formato con openpyxl ---
from openpyxl import load_workbook

wb = load_workbook(ruta_limpia)
ws = wb.active

# --- Estilos base ---
header_fill = PatternFill(start_color='BFD8B8', end_color='BFD8B8', fill_type='solid')
header_font = Font(bold=True, color='000000', name='Calibri')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# --- Aplicar formato y reemplazar vacíos ---
for row in ws.iter_rows():
    for cell in row:
        if cell.value is None or str(cell.value).strip() == '':
            cell.value = '-'
        cell.alignment = center_align
        cell.border = thin_border

# --- Encabezado ---
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align

# --- Ajustar ancho de columnas ---
for col in ws.columns:
    max_length = 0
    column = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            length = len(str(cell.value))
            if length > max_length:
                max_length = length
    ws.column_dimensions[column].width = max_length + 3

# --- Ajustar altura de filas ---
for row in ws.iter_rows():
    ws.row_dimensions[row[0].row].height = 18

# --- Guardar cambios ---
wb.save(ruta_limpia)
print(f' Archivo formateado y reparado correctamente:\n{ruta_limpia}')
#------------------Fin Formaterar tabla de Tabla_Abundancia_Semanal------------------#