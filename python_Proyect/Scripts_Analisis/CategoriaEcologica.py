


# --------------------------------------------------
# 1. LIBRERÍAS
# --------------------------------------------------
import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import textwrap

# --------------------------------------------------
# 2. CONFIGURACIÓN GENERAL
# --------------------------------------------------
ruta = r"D:\CORPONOR 2025\Backet\python_Proyect\data\SRF_LAM_5235_AVES_SAMORE_AVES.xlsx"
output_folder = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados"
os.makedirs(output_folder, exist_ok=True)

# --------------------------------------------------
# 3. CARGA DE DATOS
# --------------------------------------------------
Registros = pd.read_excel(ruta)

print("Columnas disponibles:")
print(Registros.columns)

# --------------------------------------------------
# 4. ABREVIAR COBERTURAS (AUTOMÁTICO)
# --------------------------------------------------
def generar_abreviacion(nombre):
    palabras = nombre.lower().split()
    palabras = [p for p in palabras if p not in
                ['de', 'del', 'la', 'el', 'y', 'con', 'en', 'los', 'las']]
    abrev = ''.join(p[0] for p in palabras)
    return abrev.upper()

def abreviar_coberturas(df, columna="COBERTURA"):
    dic = {c: generar_abreviacion(c) for c in df[columna].dropna().unique()}
    df[columna] = df[columna].replace(dic)
    return df, dic

Registros, abreviaciones = abreviar_coberturas(Registros)

# --------------------------------------------------
# 5. TABLA RESUMEN POR CATEGORÍA ECOLÓGICA Y COBERTURA
# --------------------------------------------------
df_cat = Registros.dropna(
    subset=["CECOLOGICA", "COBERTURA", "INDIVIDUOS"]
).copy()

cat_cobertura = (
    df_cat.groupby(["COBERTURA", "CECOLOGICA"])["INDIVIDUOS"]
    .sum()
    .reset_index()
)

total_cobertura = (
    cat_cobertura.groupby("COBERTURA")["INDIVIDUOS"]
    .sum()
    .reset_index()
    .rename(columns={"INDIVIDUOS": "Total_individuos"})
)

cat_cobertura = cat_cobertura.merge(total_cobertura, on="COBERTURA")

cat_cobertura["Abund_relativa_%"] = (
    cat_cobertura["INDIVIDUOS"] /
    cat_cobertura["Total_individuos"] * 100
).round(2)

# Guardar Excel
excel_path = os.path.join(output_folder, "11_Resumen_Categorias_Ecologicas.xlsx")
cat_cobertura.to_excel(excel_path, index=False)

# --------------------------------------------------
# 6. FORMATO PROFESIONAL EXCEL
# --------------------------------------------------
wb = load_workbook(excel_path)
ws = wb.active

header_fill = PatternFill("solid", fgColor="BFD8B8")
header_font = Font(bold=True)
center = Alignment(horizontal="center", vertical="center")
border = Border(*(Side(style="thin") for _ in range(4)))

for row in ws.iter_rows():
    for c in row:
        c.alignment = center
        c.border = border
        if c.value is None:
            c.value = "-"

for c in ws[1]:
    c.fill = header_fill
    c.font = header_font

for col in ws.columns:
    ws.column_dimensions[get_column_letter(col[0].column)].width = 18

wb.save(excel_path)

# --------------------------------------------------
# 7. ANÁLISIS GLOBAL POR CATEGORÍA ECOLÓGICA
# --------------------------------------------------
total_global = cat_cobertura["INDIVIDUOS"].sum()

cat_total = (
    cat_cobertura.groupby("CECOLOGICA")["INDIVIDUOS"]
    .sum()
    .reset_index()
)

cat_total["Abund_relativa_%"] = (
    cat_total["INDIVIDUOS"] / total_global * 100
).round(2)

cat_total = cat_total.sort_values("Abund_relativa_%", ascending=False)

# --------------------------------------------------
# 8. GRÁFICO DE TORTA GLOBAL
# --------------------------------------------------
colors = sns.color_palette("Paired", len(cat_total))

fig, ax = plt.subplots(figsize=(9, 8))
wedges, _ = ax.pie(
    cat_total["Abund_relativa_%"],
    startangle=90,
    colors=colors,
    wedgeprops={"edgecolor": "white"}
)

for i, w in enumerate(wedges):
    ang = (w.theta2 + w.theta1) / 2
    x = 0.6 * np.cos(np.deg2rad(ang))
    y = 0.6 * np.sin(np.deg2rad(ang))
    ax.text(x, y, f"{cat_total.iloc[i]['Abund_relativa_%']:.1f}%",
            ha="center", va="center", fontweight="bold")

ax.legend(
    wedges,
    cat_total["CECOLOGICA"],
    title="Categorías ecológicas",
    loc="center left",
    bbox_to_anchor=(1, 0.5)
)

ax.set_title("Distribución global de categorías ecológicas")
ax.axis("equal")

plt.tight_layout()
plt.savefig(os.path.join(output_folder, "11.1_Torta_Categorias_Ecologicas.png"),
            dpi=350)
plt.close()

# --------------------------------------------------
# 9. GRÁFICO DE BARRAS APILADAS POR COBERTURA
# --------------------------------------------------
pivot = cat_cobertura.pivot(
    index="COBERTURA",
    columns="CECOLOGICA",
    values="Abund_relativa_%"
).fillna(0)

pivot.plot(
    kind="bar",
    stacked=True,
    figsize=(11, 6),
    color=colors,
    edgecolor="black"
)

plt.ylabel("Abundancia relativa (%)")
plt.title("Categorías ecológicas por cobertura")
plt.legend(title="Categorías ecológicas",
           bbox_to_anchor=(1.05, 1),
           loc="upper left")
plt.grid(axis="y", linestyle="--", alpha=0.4)
plt.tight_layout()

plt.savefig(os.path.join(output_folder,
                         "11.2_Barras_Categorias_Ecologicas_Cobertura.png"),
            dpi=350)
plt.close()

# --------------------------------------------------
# 10. DESCRIPCIÓN AUTOMÁTICA GLOBAL
# --------------------------------------------------
desc = []
dominante = cat_total.iloc[0]

desc.append(
    f"La estructura ecológica del ensamblaje de aves se encuentra dominada por la "
    f"categoría ecológica {dominante['CECOLOGICA']}, la cual representa el "
    f"{dominante['Abund_relativa_%']:.1f}% del total de individuos registrados "
    f"({total_global} individuos)."
)

sec = cat_total.iloc[1:3]
if not sec.empty:
    lista = ", ".join(
        f"{r['CECOLOGICA']} ({r['Abund_relativa_%']:.1f}%)"
        for _, r in sec.iterrows()
    )
    desc.append(
        f"Otras categorías ecológicas con una participación relevante corresponden a "
        f"{lista}, reflejando una heterogeneidad funcional asociada a la diversidad de "
        f"coberturas evaluadas."
    )

menores = cat_total[cat_total["Abund_relativa_%"] < 5]
if not menores.empty:
    desc.append(
        "Las categorías ecológicas con baja representatividad (<5 %) incluyen: " +
        ", ".join(menores["CECOLOGICA"]) +
        ", las cuales, aunque menos abundantes, cumplen funciones ecológicas específicas."
    )

descripcion_final = "\n\n".join(desc)

# Guardar texto
txt_path = os.path.join(output_folder,
                        "11.3_Descripcion_Global_Categorias_Ecologicas.txt")

with open(txt_path, "w", encoding="utf-8") as f:
    f.write(textwrap.fill(descripcion_final, 95))

print("✔ Análisis completo generado correctamente.")







# --------------------------------------------------
# 11. INTERPRETACIÓN ECOLÓGICA POR COBERTURA
# --------------------------------------------------

interpretaciones = []

for cobertura in cat_cobertura["COBERTURA"].unique():

    sub = cat_cobertura[cat_cobertura["COBERTURA"] == cobertura].copy()

    total_ind = sub["INDIVIDUOS"].sum()
    total_cat = sub.shape[0]

    sub = sub.sort_values("Abund_relativa_%", ascending=False)

    texto = []
    texto.append(f"Cobertura {cobertura}")
    texto.append("-" * 75)

    texto.append(
        f"En la cobertura {cobertura} se registraron {total_ind} individuos de aves, "
        f"distribuidos en {total_cat} categorías ecológicas. La estructura ecológica "
        f"del ensamblaje "
        f"evidencia una distribución relativamente particular de los roles funcionales."
    )

    # Categorías con mayor representación
    principales = sub[sub["Abund_relativa_%"] >= 10]

    if not principales.empty:
        lista_principales = ", ".join(
            f"{r['CECOLOGICA']} ({r['Abund_relativa_%']:.1f}%)"
            for _, r in principales.iterrows()
        )
        texto.append(
            f"Las categorías ecológicas con mayor representatividad corresponden a "
            f"{lista_principales}, las cuales concentran una proporción significativa "
            f"del total de individuos registrados en esta cobertura."
        )

    # Categorías intermedias
    intermedias = sub[
        (sub["Abund_relativa_%"] >= 5) &
        (sub["Abund_relativa_%"] < 10)
    ]

    if not intermedias.empty:
        texto.append(
            "Adicionalmente, se identifican categorías ecológicas con participación "
            "intermedia, lo que sugiere una oferta variada de recursos tróficos y "
            "microhábitats disponibles."
        )

    # Categorías de baja representatividad
    bajas = sub[sub["Abund_relativa_%"] < 5]

    if not bajas.empty:
        texto.append(
            "Las categorías ecológicas de baja representatividad (<5 %) incluyen: " +
            ", ".join(bajas["CECOLOGICA"]) +
            ", las cuales, aunque menos abundantes, cumplen funciones ecológicas "
            "específicas y contribuyen a la complejidad funcional del ecosistema."
        )

    texto.append(
        "Este patrón es característico de ambientes con heterogeneidad estructural, "
        "donde no se favorece la dominancia funcional, sino la coexistencia de especies "
        "con diferentes estrategias ecológicas."
    )

    interpretaciones.append("\n".join(texto))


# --------------------------------------------------
# 12. EXPORTAR INTERPRETACIÓN A TXT
# --------------------------------------------------

txt_cobertura = os.path.join(
    output_folder,
    "11.4_Interpretacion_Categorias_Ecologicas_por_Cobertura.txt"
)

with open(txt_cobertura, "w", encoding="utf-8") as f:
    for bloque in interpretaciones:
        f.write(textwrap.fill(bloque, 95))
        f.write("\n\n")

print("✔ Interpretación ecológica por cobertura generada correctamente.")









#-----------------------------------------------------------
#   GRÁFICO DE TORTA – Categorías ecológicas (Estilo Profesional)
#-----------------------------------------------------------

import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
import os

# -----------------------------------------
# 🎨 PALETA PROFESIONAL: Paired
# -----------------------------------------
colors = sns.color_palette("Paired", len(cat_total))

# -----------------------------------------
# 🔧 FIGURA
# -----------------------------------------
fig, ax = plt.subplots(figsize=(9, 8))

wedges, texts = ax.pie(
    cat_total["Abund_relativa_%"],
    startangle=90,
    colors=colors,
    wedgeprops={'edgecolor': 'white', 'linewidth': 1}
)

# -----------------------------------------------------
# 🔢 ETIQUETAS INTERNAS (sin superposición)
# -----------------------------------------------------
prev_positions = []

for i, w in enumerate(wedges):
    porcentaje = cat_total["Abund_relativa_%"].iloc[i]

    ang = (w.theta2 - w.theta1) / 2 + w.theta1
    x = np.cos(np.deg2rad(ang))
    y = np.sin(np.deg2rad(ang))

    text_x, text_y = 0.55 * x, 0.55 * y

    for px, py in prev_positions:
        if abs(text_y - py) < 0.08:
            text_y += 0.1 if text_y > py else -0.1

    prev_positions.append((text_x, text_y))

    ax.text(
        text_x, text_y,
        f"{porcentaje:.1f}%",
        ha='center', va='center',
        fontsize=11, fontweight='bold',
        color="black"
    )

# -----------------------------------------
# 📝 LEYENDA PROFESIONAL
# -----------------------------------------
ax.legend(
    wedges,
    cat_total["CECOLOGICA"],
    title="Categoría ecológica",
    loc="center left",
    bbox_to_anchor=(1, 0.5),
    fontsize=12,
    title_fontsize=14,
    frameon=True,
    fancybox=True,
    edgecolor="gray"
)

# -----------------------------------------
# 🏷 TÍTULO LIMPIO
# -----------------------------------------
ax.set_title(
    "Distribución global de categorías ecológicas",
    fontsize=16,
    fontweight='bold',
    pad=20
)

ax.axis('equal')
plt.tight_layout()

# -----------------------------------------
# 💾 GUARDAR FIGURA
# -----------------------------------------
ruta_torta = os.path.join(
    output_folder,
    "11.1_Torta_Categorias_Ecologicas_PRO.png"
)

plt.savefig(ruta_torta, dpi=350, bbox_inches='tight')
plt.show()

print(f"✔ Gráfico guardado en: {ruta_torta}")












#-----------------------------------------------------------
#   GRÁFICO DE BARRAS APILADAS – Categorías ecológicas por cobertura
#-----------------------------------------------------------

# -----------------------------------------
# 🔄 TABLA PIVOTE
# -----------------------------------------
pivot = cat_cobertura.pivot(
    index="COBERTURA",
    columns="CECOLOGICA",
    values="Abund_relativa_%"
).fillna(0)

# -----------------------------------------
# 🔧 FIGURA
# -----------------------------------------
fig, ax = plt.subplots(figsize=(12, 6))

pivot.plot(
    kind="bar",
    stacked=True,
    ax=ax,
    color=colors,
    edgecolor="black",
    linewidth=0.7
)


# -----------------------------------------
# 🎨 ELIMINAR BORDES SUPERIOR Y DERECHO
# -----------------------------------------
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)

# -----------------------------------------
# 🏷 ETIQUETAS Y ESTILO
# -----------------------------------------
ax.set_ylabel("Abundancia relativa (%)", fontsize=12)
ax.set_xlabel("Cobertura", fontsize=12)

ax.set_title(
    "Distribución de categorías ecológicas por cobertura",
    fontsize=16,
    fontweight="bold",
    pad=15
)

ax.tick_params(axis='x', rotation=45, labelsize=11)
ax.tick_params(axis='y', labelsize=11)

# Grid profesional
ax.grid(True, axis="y", linestyle="--", alpha=0.4)
ax.grid(False, axis="x")

# Bordes sobrios
for spine in ax.spines.values():
    spine.set_linewidth(0.8)
    spine.set_color("#444444")

# -----------------------------------------
# 📝 LEYENDA PROFESIONAL
# -----------------------------------------
ax.legend(
    title="Categoría ecológica",
    title_fontsize=12,
    fontsize=11,
    frameon=True,
    fancybox=True,
    edgecolor="gray",
    bbox_to_anchor=(1.05, 1),
    loc="upper left"
)

plt.tight_layout()

# -----------------------------------------
# 💾 GUARDAR FIGURA
# -----------------------------------------
ruta_barras = os.path.join(
    output_folder,
    "11.2_Barras_Categorias_Ecologicas_Cobertura_PRO.png"
)

plt.savefig(ruta_barras, dpi=350, bbox_inches='tight')
plt.show()

print(f"✔ Gráfico guardado en: {ruta_barras}")
