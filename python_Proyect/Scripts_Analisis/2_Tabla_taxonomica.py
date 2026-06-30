
#--------------## Cargar librerias necesarias------------------------------

# Si no las tienes instaladas, ejecuta esta celda una vez:
# Salir del interprete con: exit() exit() python   pip install tabulate pandas numpy scipy scikit-bio openpyxl
#
# !pip install pandas numpy matplotlib tabulate openpyxl

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from tabulate import tabulate
import openpyxl

#--------------## Leer archivo y revisar columnas------------------------------
# Ruta del archivo
ruta = r"D:\Forestal Consultores\2026\FAUNA\BD\MAMIFEROS\Mamiferos_Secundario_San_Roque.xlsx"
# Leer el archivo Excel
Registros = pd.read_excel(ruta)



# =========================================================
# FILTRAR UNA SOLA CLASE
# =========================================================

grupo = "MAMIFEROS"   # AVES, MAMIFEROS, REPTILES, ANFIBIOS etc.

Registros = Registros[
    Registros["CLASE"].astype(str).str.upper() == grupo.upper()
].copy()

# Reiniciar índice
Registros.reset_index(drop=True, inplace=True)

# =========================================================
# VERIFICACIÓN
# =========================================================

print(f"\nGrupo seleccionado: {grupo}")
print(f"Número de registros: {len(Registros)}")

print("\nPrimeras filas:")
print(Registros.head())

# Mostrar las primeras filas
print(" Primeras filas del archivo:")
print(Registros.head())

# Mostrar nombres de las columnas
print("\n Columnas del DataFrame:")
print(Registros.columns)

#------------------------Tabla general grupo taxonomico-------------------------

# Mostrar las primeras filas
print(Registros.info())


import pandas as pd

# --- Copiar el DataFrame base ---
df = Registros.copy()

# --- Normalizar texto ---
for col in ['CLASE', 'Orden', 'Familia', 'Genero', 'Epiteto', 'N. comun', 'Gremio', 'COBERTURA', 'METODOLOGIA']:
    df[col] = df[col].astype(str).str.strip().str.title()

# --- Crear nombre científico completo ---
df['Especie_cientifica'] = df['Genero'] + ' ' + df['Epiteto']

# --- Diccionario de abreviaciones de Metodologia ---
abreviaciones_metodo = {
    'Auditivo': 'Aud',
    'Fotografia': 'Fot',
    'Fotografia ': 'Fot',
    'Marcas De Presencia': 'MP',
    'Avistamiento': 'Obs',
    'Observacion': 'Obs',
    'Entrevista': 'Ent',
    'Captura': 'Cap',
    'Rastros': 'Ras',
    'Huellas': 'Hue',
    'Cueva': 'Cuv',
    'Heces': 'Hec',
    'Video': 'Vid',
    'Informacion Mcnup': 'MCNUP'
}

# --- Diccionario de abreviaciones de cobertura ---
abreviaciones_cobertura = {
    'Bosque De Galería Y Ripario': 'Bgr',
    'Bosque Denso Alto De Tierra Firme': 'Bda',
    'Bosque Denso Bajo De Tierra Firme': 'Bdb',
    'Bosque Fragmentado Con Vegetación Secundaria': 'Bfvs',
    'Bosque Fragmentado': 'BF',
    'Pastos Limpios': 'PL','pastos Limpios': 'PL',
    'Pastos Enmalezados': 'PE',
    'Pastos Arbolados': 'PA',
    'Cultivos Permanentes Arboreos': 'CPA',
    'Mosaico De Cultivos, Pastos Y Espacios Naturales': 'MCPEN',
    'Zonas De Extracción Minera': 'ZEM',
    'Pastos Enmalezados': 'PE'
}

# --- Diccionario de abreviaciones de gremio ---
abreviaciones_gremio = {
    'Carnívoro': 'Car',
    'Nectarívoro': 'Nec',
    'Carroñero': 'Crr',
    'Granívoro': 'Gra',
    'Frugívoro': 'Fru',
    'Insectívoro': 'Ins',
    'Omnívoro': 'Omn',
    'Piscívoro': 'Psc',
    'Insectivoro': 'Ins',
    'Insectívoro': 'Ins',
    'Omnivoro': 'Omn',
    'Herbívoro': 'Her',
    'Herbivoro': 'Her',
    'Nan': 'NA'
}

# --- Reemplazar nombres por abreviaciones ---
df['METODOLOGIA'] = df['METODOLOGIA'].replace(abreviaciones_metodo)
df['COBERTURA'] = df['COBERTURA'].replace(abreviaciones_cobertura)
df['Gremio'] = df['Gremio'].replace(abreviaciones_gremio)

# --- Agrupar registros únicos por especie ---
tabla = (
    df.groupby(['CLASE', 'Orden', 'Familia', 'ESPECIE', 'N. comun', 'Gremio'], dropna=False)
      .agg({
          'COBERTURA': lambda x: ', '.join(sorted(set(x.dropna()))),
          'INDIVIDUOS': 'sum',
          'METODOLOGIA': lambda x: ', '.join(sorted(set(x.dropna())))
      })
      .reset_index()
)

# --- Crear tabla pivote con coberturas como columnas ---
pivot = (
    df.groupby(['ESPECIE', 'COBERTURA'], as_index=False)['INDIVIDUOS'].sum()
      .pivot(index='ESPECIE', columns='COBERTURA', values='INDIVIDUOS')
      .fillna(0)
      .reset_index()
)

# --- Unir tabla pivote con la tabla principal ---
tabla = tabla.merge(pivot, on='ESPECIE', how='left')

# --- Renombrar columnas ---
tabla = tabla.rename(columns={
    'CLASE': 'Clase',
    'Orden': 'Orden',
    'Familia': 'Familia',
    'ESPECIE': 'Especie',
    'N. comun': 'Nombre comun',
    'Gremio': 'Gremio trófico',
    'COBERTURA': 'Cobertura(s)',
    'INDIVIDUOS': 'Abundancia',
    'METODOLOGIA': 'Tipo de registro'
})

# --- Ordenar clases ---
orden_clase = ['Aves', 'Mammalia']
tabla['Clase'] = pd.Categorical(tabla['Clase'], categories=orden_clase + sorted(set(tabla['Clase']) - set(orden_clase)), ordered=True)

# --- Ordenar por Clase, Orden y Familia ---
tabla = tabla.sort_values(['Clase', 'Orden', 'Familia', 'Especie']).reset_index(drop=True)

# ---  Agregar conteo reiniciado por Clase ---
tabla['N°'] = tabla.groupby('Clase').cumcount() + 1

# ---  Insertar fila con nombres de columnas justo antes de Mammalia ---
# ---  Insertar fila con nombres de columnas justo antes de Mammalia ---
idx_mam = tabla.index[tabla['Clase'] == 'Mammalia']
if len(idx_mam) > 0:
    insert_pos = idx_mam[0]
    fila_header = pd.DataFrame([{col: str(col) for col in tabla.columns}])  #  mantiene texto
    tabla = pd.concat([tabla.iloc[:insert_pos], fila_header, tabla.iloc[insert_pos:]], ignore_index=True)


# ---  Eliminar columnas duplicadas ---
tabla = tabla.loc[:, ~tabla.columns.duplicated()]

# ---  Reordenar columnas ---
columnas_orden = ['N°', 'Clase', 'Orden', 'Familia', 'Especie', 'Nombre comun',
                  'Gremio trófico', 'MCPEN', 'CPA', 'PA', 'Bgr', 'BF', 'PL', 'ZEM',
                  'Abundancia', 'Tipo de registro']
tabla = tabla[[col for col in columnas_orden if col in tabla.columns]]



# --- Exportar a Excel a una ruta específica ---
import os

# Definir la ruta exacta donde guardar el archivo
output_path = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados"
output_file = os.path.join(output_path, "2_tabla_composicion_taxonomica.xlsx")

# Exportar el DataFrame a Excel
tabla.to_excel(output_file, index=False)

# Confirmar la ubicación del archivo guardado
print(f" Archivo exportado correctamente en:\n{output_file}")



#---------------Dar formato a archivo generato o tabla---------------


from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import os

# --- Nombre del archivo a formatear ---
output_file = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\2_tabla_composicion_taxonomica.xlsx"

# --- Verificar que el archivo existe ---
if not os.path.exists(output_file):
    raise FileNotFoundError(f"⚠️ No se encontró el archivo: {output_file}")

# --- Cargar el archivo ---
wb = load_workbook(output_file)
ws = wb.active

# --- Estilos base ---
header_fill = PatternFill(start_color='BFD8B8', end_color='BFD8B8', fill_type='solid')
header_font = Font(bold=True, color='000000', name='Calibri')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

# --- Bordes finos para toda la tabla ---
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# --- Aplicar formato y reemplazar vacíos ---
for row in ws.iter_rows():
    for cell in row:
        # Reemplazar vacíos o None por guion
        if cell.value is None or str(cell.value).strip() == '':
            cell.value = '-'
        # Aplicar formato general
        cell.border = thin_border
        cell.alignment = center_align

# --- Aplicar formato al encabezado ---
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align

# --- Ajustar ancho de columnas automáticamente ---
for col in ws.columns:
    max_length = 0
    column = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            length = len(str(cell.value))
            if length > max_length:
                max_length = length
    ws.column_dimensions[column].width = max_length + 3

# --- Ajustar altura de filas automáticamente ---
for row in ws.iter_rows():
    max_height = 15
    for cell in row:
        if cell.value and "\n" in str(cell.value):
            lines = str(cell.value).count('\n') + 1
            if lines > 1:
                max_height = 15 * lines
    ws.row_dimensions[cell.row].height = max_height

# --- Guardar cambios ---
wb.save(output_file)
print(f' Archivo formateado con éxito:\n{output_file}')







#--------------------------------------------
#--------------------------------------------
#--------------------------------------------
# ============================================================
# === GENERAR TABLA TAXONÓMICA PARA CADA MUNICIPIO ===========
# ============================================================

municipios = Registros['MUNICIPIO'].dropna().unique()

print("\n=== Generando tablas por municipio ===\n")

for muni in municipios:
    print(f" → Procesando municipio: {muni}")

    # --- Filtrar datos del municipio ---
    df_muni = Registros[Registros["MUNICIPIO"] == muni].copy()

    # ========================================================
    # --- Normalizar texto (versión robusta) ---
    # ========================================================

    columnas_normalizar = [
        'CLASE', 'Orden', 'Familia', 'Genero', 'Epiteto',
        'N. comun', 'Gremio', 'COBERTURA', 'METODOLOGIA'
    ]

    for col in columnas_normalizar:
        if col in df_muni.columns:
            df_muni[col] = (
                df_muni[col]
                .astype(str)
                .str.replace("None", "", regex=False)
                .str.replace("nan", "", regex=False)
                .str.replace("NaN", "", regex=False)
                .str.strip()
                .str.title()
            )

    # ========================================================
    # --- Nombre científico ---
    # ========================================================
    df_muni['Especie_cientifica'] = df_muni['Genero'] + ' ' + df_muni['Epiteto']

    # ========================================================
    # --- Reemplazar abreviaciones ---
    # ========================================================

    df_muni['METODOLOGIA'] = df_muni['METODOLOGIA'].replace(abreviaciones_metodo)
    df_muni['COBERTURA'] = df_muni['COBERTURA'].replace(abreviaciones_cobertura)
    df_muni['Gremio'] = df_muni['Gremio'].replace(abreviaciones_gremio)

    # ========================================================
    # --- Construcción tabla base ---
    # ========================================================

    tabla_muni = (
        df_muni.groupby(
            ['CLASE', 'Orden', 'Familia', 'ESPECIE', 'N. comun', 'Gremio'],
            dropna=False
        )
        .agg({
            'COBERTURA': lambda x: ', '.join(sorted(set(x.dropna()))),
            'INDIVIDUOS': 'sum',
            'METODOLOGIA': lambda x: ', '.join(sorted(set(x.dropna())))
        })
        .reset_index()
    )

    # ========================================================
    # --- Matriz de coberturas ---
    # ========================================================

    pivot_muni = (
        df_muni.groupby(['ESPECIE', 'COBERTURA'], as_index=False)['INDIVIDUOS']
        .sum()
        .pivot(index='ESPECIE', columns='COBERTURA', values='INDIVIDUOS')
        .fillna(0)
        .reset_index()
    )

    # Unir matriz con la tabla base
    tabla_muni = tabla_muni.merge(pivot_muni, on='ESPECIE', how='left')

    # ========================================================
    # --- Renombrar columnas ---
    # ========================================================

    tabla_muni = tabla_muni.rename(columns={
        'CLASE': 'Clase',
        'Orden': 'Orden',
        'Familia': 'Familia',
        'ESPECIE': 'Especie',
        'N. comun': 'Nombre comun',
        'Gremio': 'Gremio trofico',
        'COBERTURA': 'Cobertura(s)',
        'INDIVIDUOS': 'Abundancia',
        'METODOLOGIA': 'Tipo de registro'
    })

    # ========================================================
    # --- Ordenamiento ---
    # ========================================================

    orden_clase = ['Aves', 'Mammalia']
    tabla_muni['Clase'] = pd.Categorical(tabla_muni['Clase'], categories=orden_clase, ordered=True)

    tabla_muni = tabla_muni.sort_values(
        ['Clase', 'Orden', 'Familia', 'Especie']
    ).reset_index(drop=True)

    tabla_muni['N°'] = tabla_muni.groupby('Clase').cumcount() + 1

    # Reordenar columnas en orden estándar
    columnas_ordenadas = [
        'N°', 'Clase', 'Orden', 'Familia', 'Especie', 'Nombre comun',
        'Gremio trofico', 'Abundancia', 'Tipo de registro'
    ]

    # Añadir coberturas si existen
    coberturas_existentes = [c for c in ['Bda', 'MCPEN', 'CPA', 'PA', 'Bgr', 'BF', 'PL', 'PE'] if c in tabla_muni.columns]
    columnas_finales = columnas_ordenadas[:7] + coberturas_existentes + columnas_ordenadas[7:]

    # Conservar solo las columnas existentes
    columnas_finales = [c for c in columnas_finales if c in tabla_muni.columns]

    tabla_muni = tabla_muni[columnas_finales]

    # ========================================================
    # === EXPORTAR ARCHIVO ===================================
    # ========================================================

    archivo_salida = os.path.join(
        output_path,
        f"2_tabla_composicion_taxonomica_{muni.replace(' ', '_')}.xlsx"
    )

    tabla_muni.to_excel(archivo_salida, index=False)

    # ========================================================
    # === FORMATEO ===========================================
    # ========================================================

    wb = load_workbook(archivo_salida)
    ws = wb.active

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or str(cell.value).strip() == '':
                cell.value = "-"
            cell.border = thin_border
            cell.alignment = center_align

    # Formato encabezados
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Ajustar ancho columnas
    for col in ws.columns:
        max_len = max(len(str(c.value)) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 3

    wb.save(archivo_salida)

    print(f"   ✔ Archivo generado: {archivo_salida}")

print("\n=== PROCESO COMPLETADO ===")
