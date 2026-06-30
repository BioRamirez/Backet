#--------------## Cargar librerias necesarias------------------------------
import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from tabulate import tabulate
import openpyxl

# Carpeta donde guardarás los gráficos (solo una vez)
output_folder = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados"
os.makedirs(output_folder, exist_ok=True)
#--------------## Leer archivo y revisar columnas------------------------------

#--------------## Leer archivo y revisar columnas------------------------------
# Ruta del archivo
ruta = r"D:\Forestal Consultores\2026\FAUNA\BD\BD_SANROQUE.xlsx"
# Leer el archivo Excel
Registros = pd.read_excel(ruta)

# Mostrar las primeras filas
print(" Primeras filas del archivo:")
print(Registros.head())

# Mostrar nombres de las columnas
print("\n Columnas del DataFrame:")
print(Registros.columns)

# =========================================================
# FILTRAR UNA SOLA CLASE
# =========================================================

grupo = "MAMIFEROS"   # AVES, MAMIFEROS, REPTILES, ANFIBIOS etc.

Registros = Registros[
    Registros["CLASE"].astype(str).str.upper() == grupo.upper()
].copy()

# Reiniciar índice
Registros.reset_index(drop=True, inplace=True)


# --------------------------------------------------
# 2️⃣ Funciones para abreviar nombres de coberturas
# --------------------------------------------------

def generar_abreviacion(nombre):
    """
    Genera abreviaciones automáticas a partir de nombres de coberturas.
    Ejemplo: 'Bosque de galería y ripario' → 'Bgr'
    """
    # Convertir a minúsculas y dividir en palabras
    palabras = nombre.lower().split()

    # Eliminar conectores comunes
    palabras = [p for p in palabras if p not in ['de', 'del', 'la', 'el', 'y', 'con', 'en', 'los', 'las']]

    # Tomar la primera letra de cada palabra
    abreviacion = ''.join([p[0] for p in palabras])

    # Asegurar que tenga al menos 3 caracteres
    if len(abreviacion) < 3:
        abreviacion = abreviacion.ljust(3, ' ')

    return abreviacion.upper()
#return abreviacion.capitalize()

def abreviar_coberturas(df, columna='COBERTURA'):
    """
    Crea un diccionario de abreviaciones y reemplaza los nombres en el DataFrame.
    """
    coberturas_unicas = df[columna].dropna().unique()
    abreviaciones = {c: generar_abreviacion(c) for c in coberturas_unicas}

    print("\n Abreviaciones generadas automáticamente:")
    for original, abrev in abreviaciones.items():
        print(f"  {original} → {abrev}")

    # Reemplazar en el DataFrame
    df[columna] = df[columna].replace(abreviaciones)

    return df, abreviaciones


# --- Aplicar las abreviaciones ---
Registros, abreviaciones_cobertura = abreviar_coberturas(Registros, columna='COBERTURA')

import pandas as pd
import matplotlib.pyplot as plt
import numpy as np


# --- Verificar columnas clave ---
print(Registros.columns)

Registros
# --------------------------------------------------

# ==============================================
#  Análisis de gremios tróficos
# ==============================================

# Filtrar registros válidos
df_gremios = Registros.dropna(subset=['Gremio', 'COBERTURA', 'INDIVIDUOS']).copy()

# Agrupar por gremio y cobertura
gremio_cobertura = (
    df_gremios.groupby(['COBERTURA', 'Gremio'])['INDIVIDUOS']
    .sum()
    .reset_index()
)

# Calcular abundancia total por cobertura
total_por_cobertura = (
    gremio_cobertura.groupby('COBERTURA')['INDIVIDUOS']
    .sum()
    .reset_index()
    .rename(columns={'INDIVIDUOS': 'Total_individuos'})
)

# Unir y calcular abundancia relativa (%)
gremio_cobertura = gremio_cobertura.merge(total_por_cobertura, on='COBERTURA')
gremio_cobertura['Abund_relativa_%'] = (
    gremio_cobertura['INDIVIDUOS'] / gremio_cobertura['Total_individuos'] * 100
).round(2)

# Mostrar tabla resumen
print("\n Abundancia relativa por gremio y cobertura:")
print(gremio_cobertura.sort_values(['COBERTURA', 'Abund_relativa_%'], ascending=[True, False]))

# Guardar tabla en Excel
output_path = os.path.join(output_folder, "11_Resumen_Gremios_Troficos.xlsx")
gremio_cobertura.to_excel(output_path, index=False)
print(f"\n Archivo guardado en: {output_path}")

#---------------------------------- Reparar y formatear archivo de Resumen_Uso_Habitat -----------------------
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import os

# --- Rutas ---
ruta_original = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\11_Resumen_Gremios_Troficos.xlsx"
ruta_limpia = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\11_Resumen_Gremios_Troficos.xlsx"

# --- Verificar existencia ---
if not os.path.exists(ruta_original):
    raise FileNotFoundError(f"⚠️ No se encontró el archivo: {ruta_original}")

# --- Leer archivo dañado con pandas ---
try:
    df = pd.read_excel(ruta_original)
    print(" Archivo leído correctamente con pandas.")
except Exception as e:
    raise RuntimeError(f" No se pudo leer el archivo: {e}")

# --- Reescribir el archivo limpio ---
df.to_excel(ruta_limpia, index=False)
print(f" Archivo reparado y guardado como:\n{ruta_limpia}")

# --- Aplicar formato con openpyxl ---
from openpyxl import load_workbook

wb = load_workbook(ruta_limpia)
ws = wb.active

# --- Estilos base ---
header_fill = PatternFill(start_color='BFD8B8', end_color='BFD8B8', fill_type='solid')
header_font = Font(bold=True, color='000000', name='Calibri')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# --- Aplicar formato y reemplazar vacíos ---
for row in ws.iter_rows():
    for cell in row:
        if cell.value is None or str(cell.value).strip() == '':
            cell.value = '-'
        cell.alignment = center_align
        cell.border = thin_border

# --- Encabezado ---
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align

# --- Ajustar ancho de columnas ---
for col in ws.columns:
    max_length = 0
    column = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            length = len(str(cell.value))
            if length > max_length:
                max_length = length
    ws.column_dimensions[column].width = max_length + 3

# --- Ajustar altura de filas ---
for row in ws.iter_rows():
    ws.row_dimensions[row[0].row].height = 18

# --- Guardar cambios ---
wb.save(ruta_limpia)
print(f' Archivo formateado y reparado correctamente:\n{ruta_limpia}')
























import pandas as pd

# ================================================================
# CONFIGURACIÓN
# ================================================================
output_txt = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\11.4_Analisis_Gremios_TOP5.txt"

# ================================================================
# 1. FUNCIÓN PARA EXTRAER TOP 5 FAMILIAS Y ESPECIES POR GREMIO
# ================================================================
def top5_familias(df):
    return (
        df.groupby("Familia")["INDIVIDUOS"]
        .sum()
        .sort_values(ascending=False)
        .head(5)
    )

def top5_especies(df):
    return (
        df.groupby("Especie")["INDIVIDUOS"]
        .sum()
        .sort_values(ascending=False)
        .head(5)
    )

# ================================================================
# 2. ANÁLISIS GENERAL POR GREMIO
# ================================================================
texto_general = []
texto_general.append("ANÁLISIS FUNCIONAL POR GREMIO – DESCRIPCIÓN GENERAL\n")

gremios = sorted(Registros["Gremio"].dropna().unique())

for gremio in gremios:

    df_g = Registros[Registros["Gremio"] == gremio]

    # Top 5 familias y especies por abundancia
    familias_top = top5_familias(df_g)
    especies_top = top5_especies(df_g)

    familias_txt = ", ".join([f"{fam} ({int(num)})" for fam, num in familias_top.items()])
    especies_txt = ", ".join([f"{sp} ({int(num)})" for sp, num in especies_top.items()])

    texto = (
        f"\n► **Gremio {gremio}**\n"
        f"Las familias más representativas son: {familias_txt}. "
        f"Las especies con mayor abundancia fueron: {especies_txt}. "
        f"Este gremio desempeña funciones ecológicas clave dentro del ensamblaje, "
        f"reflejando una disponibilidad diferencial de recursos en el paisaje."
    )

    texto_general.append(texto)

# ================================================================
# 3. ANÁLISIS POR COBERTURA Y GREMIO (TOP 5)
# ================================================================
texto_coberturas = []
texto_coberturas.append("\n\nANÁLISIS FUNCIONAL POR COBERTURA\n")

coberturas = sorted(Registros["COBERTURA"].dropna().unique())

for cob in coberturas:

    texto_coberturas.append(f"\n\n=== COBERTURA: {cob} ===\n")

    df_c = Registros[Registros["COBERTURA"] == cob]

    gremios_cob = sorted(df_c["Gremio"].dropna().unique())

    for gremio in gremios_cob:

        df_g = df_c[df_c["Gremio"] == gremio]

        familias_top = top5_familias(df_g)
        especies_top = top5_especies(df_g)

        familias_txt = ", ".join([f"{fam} ({int(num)})" for fam, num in familias_top.items()])
        especies_txt = ", ".join([f"{sp} ({int(num)})" for sp, num in especies_top.items()])

        texto = (
            f"\n• **Gremio {gremio}**: "
            f"Las familias mejor representadas fueron {familias_txt}. "
            f"Entre las especies dominantes destacan {especies_txt}. "
            f"Este patrón refleja la oferta de recursos y la estructura vegetal "
            f"particular de la cobertura {cob}, determinando la composición funcional."
        )

        texto_coberturas.append(texto)

# ================================================================
# 4. UNIR TODO EL TEXTO
# ================================================================
texto_final = "\n".join(texto_general) + "\n\n" + "\n".join(texto_coberturas)

# ================================================================
# 5. GUARDAR
# ================================================================
with open(output_txt, "w", encoding="utf-8") as f:
    f.write(texto_final)

print("\n✔ Archivo de análisis funcional generado en:")
print(output_txt)






































#  ===============================
#  Gráfico 1: barras apiladas por cobertura
#  ===============================

fig, ax = plt.subplots(figsize=(10,6))

# ---------------------------------------------
# 1. Crear tabla dinámica (pivot)
# ---------------------------------------------
pivot_data = gremio_cobertura.pivot(
    index='COBERTURA',
    columns='Gremio',
    values='Abund_relativa_%'
).fillna(0)

# ---------------------------------------------
# 2. CONFIGURAR PALETA TAB10 REORDENADA
# ---------------------------------------------
import matplotlib.cm as cm

# ===========================================
#  Índices y colores de TAB20 (Matplotlib)
# ===========================================
# 0  - azul fuerte
# 1  - azul claro
# 2  - naranja fuerte
# 3  - naranja claro
# 4  - verde fuerte
# 5  - verde claro
# 6  - rojo fuerte
# 7  - rojo claro
# 8  - morado fuerte
# 9  - morado claro
# 10 - marrón fuerte
# 11 - marrón claro
# 12 - rosa/magenta fuerte
# 13 - rosa/magenta claro
# 14 - gris fuerte
# 15 - gris claro
# 16 - oliva fuerte
# 17 - oliva claro
# 18 - turquesa fuerte
# 19 - turquesa claro
# ===========================================
#  ORDEN PERSONALIZADO PARA TAB20
#  Modifica la lista para cambiar el orden
# ===========================================

orden_colores = [
    
    19,   # turquesa claro
    16,   # oliva fuerte
    3,  # naranja claro
    10,   # marrón fuerte
    0,   # azul fuerte
    13,   # rosa/magenta claro
    6,  # rojo fuerte
    10,  # marrón fuerte
    14,  # gris fuerte
    12   # magenta fuerte
]

# ⬆ Cambia el orden como desees

# Seleccionar solo los colores necesarios según número de gremios
colors = [cm.tab20.colors[i] for i in orden_colores[:len(pivot_data.columns)]]

# ---------------------------------------------
# 3. GRAFICAR usando los colores reordenados
# ---------------------------------------------
pivot_data.plot(
    kind='bar',
    stacked=True,
    color=colors,      # <-- aquí aplican tus colores
    ax=ax
)

# ---------------------------------------------
# 4. Estilo y etiquetas
# ---------------------------------------------
ax.set_ylabel('Abundancia relativa (%)')
#Distribución de gremios tróficos por cobertura
ax.set_title('', weight='bold')

ax.legend(
    title='Gremio trófico',
    fontsize=8,
    title_fontsize=9,
    bbox_to_anchor=(1.05, 1),
    loc='upper left'
)

ax.grid(axis='y', linestyle='--', alpha=0.4)

plt.tight_layout()

# ---------------------------------------------
# 5. Guardar gráfico
# ---------------------------------------------
grafico_path = os.path.join(output_folder, "11.1_Grafico_Gremios_Troficos.png")
plt.savefig(grafico_path, dpi=300)
plt.show()

print(f" Gráfico guardado en: {grafico_path}")


#---------------------------------------------------------------------------------------



gremio_total = (
    Registros.groupby("Gremio")["INDIVIDUOS"]
    .sum()
    .reset_index()
    .sort_values("INDIVIDUOS", ascending=False)
)


import seaborn as sns

# Crear paleta igual a la del gráfico de torta
paired_colors = sns.color_palette("Paired", len(gremio_total))

# Diccionario {gremio: color}
color_dict = {
    gremio_total["Gremio"].iloc[i]: paired_colors[i]
    for i in range(len(gremio_total))
}

# Puedes imprimirlo para verificar
for k, v in color_dict.items():
    print(k, v)


# ===============================
#  Gráfico 1: barras apiladas por cobertura (COLORES IGUALES A LA TORTA)
# ===============================

fig, ax = plt.subplots(figsize=(10, 6))

# ---------------------------------------------
# 1. Crear tabla dinámica (pivot)
# ---------------------------------------------
pivot_data = gremio_cobertura.pivot(
    index='COBERTURA',
    columns='Gremio',
    values='Abund_relativa_%'
).fillna(0)

# ---------------------------------------------
# 2. Aplicar colores del diccionario
# ---------------------------------------------
# En el mismo orden en que aparecen las columnas
colors = [color_dict[g] for g in pivot_data.columns]

# ---------------------------------------------
# 3. Graficar con colores consistentes
# ---------------------------------------------
pivot_data.plot(
    kind='bar',
    stacked=True,
    color=colors,
    ax=ax,
    edgecolor="black",
    linewidth=0.5
)

# ---------------------------------------------
# 4. Estilo profesional
# ---------------------------------------------
ax.set_ylabel('Abundancia relativa (%)', fontsize=12)
ax.set_title('Distribución de gremios tróficos por cobertura', fontsize=14, weight='bold')

ax.legend(
    title='Gremio trófico',
    fontsize=9,
    title_fontsize=10,
    bbox_to_anchor=(1.05, 1),
    loc='upper left',
    frameon=True,
    edgecolor="gray",
    fancybox=True
)

ax.grid(axis='y', linestyle='--', alpha=0.4)

# Borde inferior más fuerte
ax.spines['bottom'].set_linewidth(1.4)
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)

plt.xticks(rotation=25, fontsize=10)

plt.tight_layout()

# ---------------------------------------------
# 5. Guardar gráfico
# ---------------------------------------------
grafico_path = os.path.join(output_folder, "11.1_Grafico_Gremios_Troficos_PRO.png")
plt.savefig(grafico_path, dpi=350, bbox_inches='tight')
plt.show()

print(f"✔ Gráfico guardado en: {grafico_path}")



















































































#---------------------------Calculo general de individuos por gremio--------------------------


# ===============================
#  CÁLCULO CORRECTO DE PORCENTAJE GLOBAL POR GREMIO
# ===============================

# 1. Total de individuos global
total_individuos_global = gremio_cobertura["INDIVIDUOS"].sum()

# 2. Total de individuos por gremio (global)
gremio_total = (
    gremio_cobertura.groupby("Gremio")["INDIVIDUOS"]
    .sum()
    .reset_index()
)

# 3. Calcular abundancia relativa REAL
gremio_total["Abund_relativa_%"] = (
    gremio_total["INDIVIDUOS"] / total_individuos_global * 100
)

# 4. Ordenar
gremio_total = gremio_total.sort_values(by="Abund_relativa_%", ascending=False)






#-------------------Grafico general de individuos por gremio--------------------------



# --- Generar colores ---
num_gremios = len(gremio_total)
colors = plt.cm.tab20(np.linspace(0, 1, num_gremios))

fig, ax = plt.subplots(figsize=(9, 8))

wedges, texts = ax.pie(
    gremio_total["Abund_relativa_%"],
    startangle=90,
    colors=colors,
    wedgeprops={'edgecolor': 'white'}
)

# === Etiquetas internas ajustadas ===
prev_positions = []

for i, w in enumerate(wedges):
    porcentaje = gremio_total["Abund_relativa_%"].iloc[i]
    ang = (w.theta2 - w.theta1)/2 + w.theta1
    x = np.cos(np.deg2rad(ang))
    y = np.sin(np.deg2rad(ang))
    text_x, text_y = 0.6*x, 0.6*y

    for px, py in prev_positions:
        if abs(text_y - py) < 0.08:
            text_y += 0.1 if text_y > py else -0.1

    prev_positions.append((text_x, text_y))

    ax.text(text_x, text_y, f"{porcentaje:.1f}%", ha='center',
            va='center', fontsize=10, fontweight='bold')

ax.legend(wedges, gremio_total["Gremio"],
          title="Gremio trófico",
          loc="center left", bbox_to_anchor=(1, 0.5),
          fontsize=12, title_fontsize=14)
#Distribución general de gremios tróficos
ax.set_title("",
             fontsize=14, fontweight='bold', pad=20)

ax.axis('equal')
plt.tight_layout()

grafico_torta_path = os.path.join(output_folder,
                                  "11.2_Grafico_Torta_Gremios_Troficos.png")
plt.savefig(grafico_torta_path, dpi=300, bbox_inches='tight')
plt.show()

print(f"Gráfico guardado en: {grafico_torta_path}")











#-----------------------------------------------------------
#   GRÁFICO DE TORTA – Gremios tróficos (Estilo Profesional)
#-----------------------------------------------------------

import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
import os

# -----------------------------------------
# 🎨 PALETA PROFESIONAL: Paired (12 colores)
# -----------------------------------------
colors = sns.color_palette("Paired", len(gremio_total))

# -----------------------------------------
# 🔧 FIGURA
# -----------------------------------------
fig, ax = plt.subplots(figsize=(9, 8))

wedges, texts = ax.pie(
    gremio_total["Abund_relativa_%"],
    startangle=90,
    colors=colors,
    wedgeprops={'edgecolor': 'white', 'linewidth': 1}
)

# -----------------------------------------------------
# 🔢 Etiquetas internas profesionales y sin superposición
# -----------------------------------------------------
prev_positions = []

for i, w in enumerate(wedges):
    porcentaje = gremio_total["Abund_relativa_%"].iloc[i]

    ang = (w.theta2 - w.theta1)/2 + w.theta1
    x = np.cos(np.deg2rad(ang))
    y = np.sin(np.deg2rad(ang))

    text_x, text_y = 0.55 * x, 0.55 * y

    for px, py in prev_positions:
        if abs(text_y - py) < 0.08:
            text_y += 0.1 if text_y > py else -0.1

    prev_positions.append((text_x, text_y))

    ax.text(
        text_x, text_y,
        f"{porcentaje:.1f}%",
        ha='center', va='center',
        fontsize=11, fontweight='bold',
        color="black"
    )

# -----------------------------------------
# 📝 LEYENDA PROFESIONAL
# -----------------------------------------
ax.legend(
    wedges,
    gremio_total["Gremio"],
    title="Gremio Trófico",
    loc="center left",
    bbox_to_anchor=(1, 0.5),
    fontsize=12,
    title_fontsize=14,
    frameon=True,
    fancybox=True,
    edgecolor="gray"
)

# -----------------------------------------
# 🏷 TÍTULO LIMPIO (igual estilo)
# -----------------------------------------
ax.set_title(
    "",
    fontsize=16,
    fontweight='bold',
    pad=20
)

ax.axis('equal')
plt.tight_layout()

# -----------------------------------------
# 💾 GUARDAR FIGURA
# -----------------------------------------
grafico_torta_path = os.path.join(
    output_folder,
    "11.2_Grafico_Torta_Gremios_Troficos_PRO.png"
)

plt.savefig(grafico_torta_path, dpi=350, bbox_inches='tight')
plt.show()

print(f"✔ Gráfico guardado en: {grafico_torta_path}")

























































#----------------------------------------------------------------------

# ============================================
# GENERAR DESCRIPCIÓN AUTOMÁTICA GLOBAL POR GREMIO
# ============================================

desc_lines = []

total_ind = total_individuos_global

# Ordenar datos (ya vienen ordenados, pero aseguramos consistencia)
df_desc = gremio_total.copy().reset_index(drop=True)

# Gremio dominante
g1 = df_desc.iloc[0]["Gremio"]
p1 = df_desc.iloc[0]["Abund_relativa_%"]

# Segundo y tercero
g2 = df_desc.iloc[1]["Gremio"]
p2 = df_desc.iloc[1]["Abund_relativa_%"]

g3 = df_desc.iloc[2]["Gremio"]
p3 = df_desc.iloc[2]["Abund_relativa_%"]

# Gremios minoritarios (menos del 5%)
gremios_poco_frecuentes = df_desc[df_desc["Abund_relativa_%"] < 5]
n_menores = len(gremios_poco_frecuentes)

# Construcción del párrafo
descripcion_global = (
    f"El análisis general de los gremios tróficos evidencia una comunidad estructurada "
    f"por la dominancia del gremio {g1.lower()}, que representa el {p1:.1f}% del total "
    f"de individuos registrados ({total_ind} individuos). Este patrón indica que las "
    f"condiciones ambientales favorecen fuertemente las estrategias alimenticias propias "
    f"de este gremio.\n\n"
    f"En menor proporción, pero también con una contribución destacada, se encuentran los "
    f"gremios {g2.lower()} ({p2:.1f}%) y {g3.lower()} ({p3:.1f}%), los cuales desempeñan "
    f"un papel importante dentro de la dinámica ecológica del paisaje evaluado.\n\n"
)

# Agregar gremios minoritarios solo si hay alguno
if n_menores > 0:
    lista_menores = ", ".join(gremios_poco_frecuentes["Gremio"].str.lower())
    descripcion_global += (
        f"Por otra parte, se identificaron gremios con baja representatividad relativa "
        f"(<5%), entre ellos: {lista_menores}. Aunque su aporte cuantitativo es reducido, "
        f"estos gremios cumplen funciones ecológicas específicas que contribuyen al "
        f"equilibrio funcional del ensamblaje de aves.\n"
    )
else:
    descripcion_global += (
        f"No se registran gremios con representaciones inferiores al 5%, indicando una "
        f"estructura más uniforme entre los grupos funcionales.\n"
    )

# Imprimir descripción final
print("\nDESCRIPCIÓN ECOLÓGICA GLOBAL POR GREMIO:\n")
print(descripcion_global)



# --------------------------------------------
# IMPRIMIR DESCRIPCIÓN FINAL EN CONSOLA
# --------------------------------------------
print("\nDESCRIPCIÓN ECOLÓGICA GLOBAL POR GREMIO:\n")
print(descripcion_global)

# --------------------------------------------
# GUARDAR DESCRIPCIÓN EN ARCHIVO .TXT
# --------------------------------------------
descripcion_path = os.path.join(output_folder, "11.3_Descripcion_Gremios_Troficos.txt")

import textwrap

# Ajustar el contenido a un máximo de 90 caracteres por línea
contenido_líneas_90 = textwrap.fill(descripcion_path, width=90)

with open(descripcion_path, "w", encoding="utf-8") as f:
    f.write("DESCRIPCIÓN ECOLÓGICA GLOBAL POR GREMIO\n")
    f.write("="*60 + "\n\n")
    f.write(descripcion_global)
    f.write(contenido_líneas_90)

print(f"\n Descripción guardada en: {descripcion_path}")


#----------------------------------------------------------------------
#----------------------------------------------------------------------










#----------------Desccripcion por cobertura-----------------------


# ==============================================================
# GENERACIÓN AUTOMÁTICA DE DESCRIPCIONES POR COBERTURA
# ==============================================================

descripciones = []

# Validar que coberturas exista
try:
    coberturas = gremio_cobertura["COBERTURA"].unique().tolist()
except:
    print("❌ ERROR: La variable 'uso_habitat' no contiene la columna COBERTURA.")
    raise SystemExit()

# Validar que gremio_cobertura existe
if "gremio_cobertura" not in globals():
    print("❌ ERROR: La tabla gremio_cobertura no está definida.")
    raise SystemExit()


for cob in coberturas:
    df_c = gremio_cobertura[gremio_cobertura["COBERTURA"] == cob]

    if df_c.empty:
        print(f"⚠ Advertencia: Sin datos para cobertura {cob}")
        continue

    # Total individuos
    total_ind = df_c["INDIVIDUOS"].sum()

    # Ordenar por abundancia (descendente)
    df_sorted = df_c.sort_values("INDIVIDUOS", ascending=False).reset_index(drop=True)

    # Construcción segura de descripciones
    texto = []
    texto.append(f"COBERTURA {cob}")
    texto.append(f"Total individuos registrados: {total_ind}.")

    # Gremio dominante
    gremio_dom = df_sorted.iloc[0]
    texto.append(
        f"El gremio dominante fue **{gremio_dom['Gremio']}**, con "
        f"{gremio_dom['INDIVIDUOS']} individuos "
        f"({gremio_dom['Abund_relativa_%']:.2f}%)."
    )

    # Segundo gremio (solo si existe)
    if len(df_sorted) > 1:
        gremio_seg = df_sorted.iloc[1]
        texto.append(
            f"El segundo gremio más representativo fue {gremio_seg['Gremio']}, "
            f"con {gremio_seg['INDIVIDUOS']} individuos "
            f"({gremio_seg['Abund_relativa_%']:.2f}%)."
        )
    else:
        texto.append("No se registraron otros gremios representativos.")

    # Gremios menores
    if len(df_sorted) > 2:
        gremios_menores = df_sorted.iloc[2:]
        lista_menores = ", ".join(
            f"{row['Gremio']} ({row['Abund_relativa_%']:.2f}%)"
            for _, row in gremios_menores.iterrows()
        )
        texto.append(f"Los gremios minoritarios fueron: {lista_menores}.")
    else:
        texto.append("No se registraron gremios minoritarios adicionales.")

    descripciones.append("\n".join(texto) + "\n" + "-"*80)


# ==============================================================
# GUARDAR TEXTO
# ==============================================================

output_txt = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\11.4_Descripcion_gremios_por_cobertura.txt"

with open(output_txt, "w", encoding="utf-8") as f:
    f.write("\n\n".join(descripciones))

print(f"\n✔ Archivo de descripciones guardado en:\n{output_txt}")


#--------------------------Graficos por cobertura----------------

import matplotlib.pyplot as plt
import numpy as np
import os

# Carpeta de salida
output_folder = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados"

# Obtener coberturas únicas
coberturas = gremio_cobertura["COBERTURA"].unique()

# ==========================================================
#   GENERAR UN GRÁFICO DE TORTA POR CADA COBERTURA
# ==========================================================

for cob in coberturas:

    df_c = gremio_cobertura[gremio_cobertura["COBERTURA"] == cob].copy()

    # Calcular abundancia relativa dentro de la cobertura
    total_cob = df_c["INDIVIDUOS"].sum()
    df_c["Abund_relativa_cob_%"] = (df_c["INDIVIDUOS"] / total_cob) * 100

    # Ordenar
    df_c = df_c.sort_values(by="Abund_relativa_cob_%", ascending=False)

    # Generar colores
    num_gremios = len(df_c)
    colors = plt.cm.tab20(np.linspace(0, 1, num_gremios))

    # Crear figura
    fig, ax = plt.subplots(figsize=(9, 8))

    # Crear gráfico de torta
    wedges, texts = ax.pie(
        df_c["Abund_relativa_cob_%"],
        startangle=90,
        colors=colors,
        wedgeprops={'edgecolor': 'white'}
    )

    # ==========================================================
    #   Etiquetas internas con anti-solapamiento
    # ==========================================================
    prev_positions = []

    for i, w in enumerate(wedges):
        porcentaje = df_c["Abund_relativa_cob_%"].iloc[i]
        ang = (w.theta2 - w.theta1) / 2 + w.theta1
        x = np.cos(np.deg2rad(ang))
        y = np.sin(np.deg2rad(ang))

        text_x, text_y = 0.6 * x, 0.6 * y

        # Evitar solapamiento vertical
        for px, py in prev_positions:
            if abs(text_y - py) < 0.08:
                text_y += 0.1 if text_y > py else -0.1

        prev_positions.append((text_x, text_y))

        # Línea si la etiqueta queda fuera
        if abs(text_x) > 0.75 or abs(text_y) > 0.75:
            ax.plot([0.8 * x, text_x], [0.8 * y, text_y], color='gray', lw=0.8)

        ax.text(
            text_x, text_y,
            f"{porcentaje:.1f}%",
            ha="center", va="center",
            fontsize=10, fontweight="bold", color="black"
        )

    # Leyenda
    ax.legend(
        wedges,
        df_c["Gremio"],
        title="Gremio trófico",
        loc="center left",
        bbox_to_anchor=(1, 0.5),
        fontsize=12,
        title_fontsize=14,
        frameon=True
    )

    # Título
    ax.set_title(
        f"Distribución de gremios tróficos – Cobertura {cob}",
        fontsize=14, fontweight='bold', pad=20
    )

    ax.axis("equal")
    plt.tight_layout()

    # Guardar gráfico
    name = f"11.5_Gra_Torta_Gremios_Indv_{cob}.png"
    grafico_path = os.path.join(output_folder, name)
    plt.savefig(grafico_path, dpi=300, bbox_inches='tight')
    plt.close()

    print(f"✔ Gráfico guardado: {grafico_path}")



































