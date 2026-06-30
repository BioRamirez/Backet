#--------------## Cargar librerias necesarias------------------------------

# Si no las tienes instaladas, ejecuta esta celda una vez:
# Salir del interprete con: exit() exit() python   pip install tabulate pandas numpy scipy scikit-bio openpyxl
#
# !pip install pandas numpy matplotlib tabulate openpyxl

import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from tabulate import tabulate
import openpyxl


# Carpeta donde guardarás los gráficos (solo una vez)
output_folder = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados"
os.makedirs(output_folder, exist_ok=True)

#--------------## Leer archivo y revisar columnas------------------------------
# Ruta del archivo
ruta = r"D:\Forestal Consultores\2026\FAUNA\BD\BD_SANROQUE.xlsx"
# Leer el archivo Excel
Registros = pd.read_excel(ruta)

# Mostrar las primeras filas
print(" Primeras filas del archivo:")
print(Registros.head())

# Mostrar nombres de las columnas
print("\n Columnas del DataFrame:")
print(Registros.columns)

# =========================================================
# FILTRAR UNA SOLA CLASE
# =========================================================

grupo = "MAMIFEROS"   # AVES, MAMIFEROS, REPTILES, ANFIBIOS etc.

Registros = Registros[
    Registros["CLASE"].astype(str).str.upper() == grupo.upper()
].copy()

# Reiniciar índice
Registros.reset_index(drop=True, inplace=True)


#------------------------Tabla general grupo taxonomico-------------------------

# Ruta del archivo
ruta = r"D:\Forestal Consultores\2026\FAUNA\BD\BD_SANROQUE.xlsx"
# Leer el archivo Excel
Registros = pd.read_excel(ruta)


# Mostrar las primeras filas
print(Registros.info())


import pandas as pd

# --- Copiar el DataFrame base ---
df = Registros.copy()

# --- Normalizar texto ---
for col in ['CLASE', 'Orden', 'Familia', 'Genero', 'N. comun', 'Gremio', 'COBERTURA', 'METODOLOGIA']:
    df[col] = df[col].astype(str).str.strip().str.title()

# --- Crear nombre científico completo ---
df['Especie_cientifica'] = df['Genero'] + ' ' + df['Epiteto']

# --- Diccionario de abreviaciones de Metodologia ---
abreviaciones_metodo = {
    'Auditivo': 'Aud',
    'Fotografia': 'Fot',
    'Fotografia ': 'Fot',
    'Marcas De Presencia': 'MP',
    'Avistamiento': 'Obs',
    'Observacion': 'Obs',
    'Entrevista': 'Ent',
    'Captura': 'Cap',
    'Rastros': 'Ras',
    'Huellas': 'Hue',
    'Cueva': 'Cuv',
    'Heces': 'Hec',
    'Video': 'Vid',
    'Informacion Mcnup': 'MCNUP'
}

# --- Diccionario de abreviaciones de cobertura ---
abreviaciones_cobertura = {
    'Bosque De Galería Y Ripario': 'Bgr',
    'Bosque Denso Alto De Tierra Firme': 'Bda',
    'Bosque Denso alto con vegetacion secundaria': 'Bdavs',
    'Bosque Denso Bajo De Tierra Firme': 'Bdb',
    'Bosque Fragmentado Con Vegetación Secundaria': 'Bfvs',
    'bosque de galería y ripario': 'Bgr',
    'bosque de galeria y ripario': 'Bgr',
    'bosque denso alto de tierra firme': 'Bda',
    'bosque denso bajo de tierra firme': 'Bdb',
    'bosque fragmentado con vegetación secundaria': 'Bfvs',
    'Bosque denso alto con vegetacion secundaria': 'Bdavs',
    'bosque fragmentado': 'BF',
    'pastos limpios': 'PL',
    'pastos enmalezados': 'PE',
    'pastos arbolados': 'PA',
    'cultivos permanentes arboreos': 'CPA',
    'mosaico de cultivos, pastos y espacios naturales': 'MEN',
    'zonas de extracción minera': 'ZEM',
    'sin dato': 'NA'
}

# --- Diccionario de abreviaciones de gremio ---
abreviaciones_gremio = {
    'Carnívoro': 'Car',
    'Nectarívoro': 'Nec',
    'Carroñero': 'Crr',
    'Granívoro': 'Gra',
    'Frugívoro': 'Fru',
    'Insectívoro': 'Ins',
    'Omnívoro': 'Omn',
    'Herbívoro': 'Her',
    'Herbivoro': 'Her',
    'Nan': 'NA'
}

# --- Reemplazar nombres por abreviaciones ---
df['METODOLOGIA'] = df['METODOLOGIA'].replace(abreviaciones_metodo)
df['COBERTURA'] = df['COBERTURA'].replace(abreviaciones_cobertura)
df['Gremio'] = df['Gremio'].replace(abreviaciones_gremio)


# Mostrar nombres de las columnas
print("\n Columnas del DataFrame:")
print(df.columns)

# ==========================
#  TABLA DE SENSIBILIDAD
# ==========================

# --- Seleccionar columnas relevantes ---
tabla_sensibilidad = df[[
    'CLASE',
    'Familia',
    'Especie_cientifica',
    'N. comun',
    'IUCN',
    'MADS (Resol 0126)',
    'CITES',
    'Dist_Geo',
    'Tipo_Migra'
]].copy()

# --- Renombrar columnas ---
tabla_sensibilidad = tabla_sensibilidad.rename(columns={
    'Especie_cientifica': 'Especie',
    'CLASE': 'Clase',
    'N. comun': 'N. común',
    'MADS (Resol 0126)': 'Res. 0126',
    'Dist_Geo': 'Distribución',
    'Tipo_Migra': 'Migración'
})

# --- Normalizar texto ---
for col in ['Clase','Familia', 'N. común', 'Distribución', 'Migración']:
    tabla_sensibilidad[col] = tabla_sensibilidad[col].astype(str).str.strip().str.title()

# --- Eliminar duplicados (una fila por especie) ---
tabla_sensibilidad = tabla_sensibilidad.drop_duplicates(subset=['Especie']).reset_index(drop=True)

# --- Mostrar tabla final ---
print(tabulate(tabla_sensibilidad.head(10), headers='keys', tablefmt='github', showindex=False))

# Mostrar nombres de las columnas
print("\n Columnas del DataFrame:")
print(tabla_sensibilidad.columns)

# --- Revisar valores únicos en las columnas de interés ---
print(" Valores únicos en la columna 'IUCN':")
print(df['IUCN'].dropna().unique())

print("\n Valores únicos en la columna 'MADS (Resol 0126)':")
print(df['MADS (Resol 0126)'].dropna().unique())

# --- Diccionario de abreviaciones IUCN ---
abreviaciones_iucn = {
    'Preocupación Menor (LC)': 'LC',
    'Preocupacin Menor (LC)': 'LC',  # error ortográfico corregido
    'Casi Amenazado (NT)': 'NT',
    'Casi Amenazada (NT)': 'NT',
    'Vulnerable (VU)': 'VU',
    'En Peligro (EN)': 'EN',
    'En Peligro Crítico (CR)': 'CR',
    'Extinto En Estado Silvestre (EW)': 'EW',
    'Extinto (EX)': 'EX',
    'Datos Insuficientes (DD)': 'DD',
    'No Evaluado (NE)': 'NE'
}

# --- Diccionario de abreviaciones Resolución 0126 (MADS) ---
abreviaciones_mads = {
    'Preocupación Menor (LC)': 'LC',
    'Casi Amenazada(NT)': 'NT',
    'Vulnerable (VU)': 'VU',
    'En Peligro (EN)': 'EN',
    'En Peligro Crítico (CR)': 'CR',
    'Extinto En Estado Silvestre (EW)': 'EW',
    'Extinto (EX)': 'EX',
    'No Listada': 'NL',
    'NL': 'NL',
    'No aplica': 'NA'
}
# --- Aplicar abreviaciones ---


tabla_sensibilidad['IUCN'] = tabla_sensibilidad['IUCN'].replace(abreviaciones_iucn)
tabla_sensibilidad['Res. 0126'] = tabla_sensibilidad['Res. 0126'].replace(abreviaciones_mads)

# Mostrar nombres de las columnas
print("\n Columnas del DataFrame:")
print(tabla_sensibilidad.columns)

# ====================================================
#  Filtrar especies sensibles sin modificar formato original
# ====================================================
# ---------- Filtrado robusto sin alterar formato original ----------
import re

# trabajar sobre copia
temp = tabla_sensibilidad.copy()

# columnas de interés
cols = ['IUCN', 'Res. 0126', 'CITES', 'Distribución', 'Migración']

# preparar columna normalizada para cada campo (solo para comparar)
def norm_iucn(v):
    if pd.isna(v): return ''
    s = str(v).strip().lower()
    # arreglar errores comunes
    s = s.replace('preocupacin', 'preocupación')
    s = re.sub(r'[^\w\(\) ]', '', s)  # quitar puntuación inusual salvo paréntesis
    # mapear variantes a códigos
    if s in ('lc', 'preocupación menor lc', 'preocupación menor (lc)', 'preocupación menor'):
        return 'lc'
    if s in ('nt', 'casi amenazado (nt)', 'casi amenazado'):
        return 'nt'
    if s in ('vu', 'vulnerable (vu)', 'vulnerable'):
        return 'vu'
    if s in ('en', 'en peligro (en)', 'en peligro'):
        return 'en'
    if s in ('cr', 'en peligro crítico (cr)', 'en peligro crítico'):
        return 'cr'
    if s in ('ew', 'extinto en estado silvestre (ew)', 'extinto en estado silvestre'):
        return 'ew'
    if s in ('ex', 'extinto (ex)', 'extinto'):
        return 'ex'
    if s in ('dd', 'datos insuficientes (dd)', 'datos insuficientes'):
        return 'dd'
    if s in ('ne', 'no evaluado (ne)', 'no evaluado', 'no evaluada (ne)', 'no evaluada'):
        return 'ne'
    # Si ya es la frase larga
    return s

def norm_res126(v):
    if pd.isna(v): return ''
    s = str(v).strip().lower()
    s = s.replace('preocupacin', 'preocupación')
    s = re.sub(r'[^\w\(\) ]', '', s)
    if s in ('nl', 'no listada', 'no listada (nl)'):
        return 'nl'
    if s in ('no aplica', 'na', ''):
        return ''
    # mapear igual que iucn si vienen clasificaciones iguales
    mapped = norm_iucn(s)
    return mapped if mapped in ('lc','nt','vu','en','cr','ew','ex','dd','ne') else s

def norm_cites(v):
    if pd.isna(v): return ''
    s = str(v).strip().lower()
    s = s.replace('á','a')
    s = s.replace('apendice', 'apendice')  # mantener ortografía
    # mapear apéndices
    if 'apendice i' in s or 'apendice i' == s or 'apendicei'==s:
        return 'apendice i'
    if 'apendice ii' in s or 'apendice ii' == s or 'apendiceii'==s:
        return 'apendice ii'
    if 'apendice iii' in s or 'apendice iii' == s or 'apendiceiii'==s:
        return 'apendice iii'
    if 'no aplica' in s or s in ('', 'na'):
        return ''
    return s

def norm_dist(v):
    if pd.isna(v): return ''
    s = str(v).strip().lower()
    # considerar listas/comas: si contiene 'neotropical' marcar como 'neotropical'
    if 'neotropical' in s:
        return 'neotropical'
    if 'cosmopolita' in s:
        return 'cosmopolita'
    if 'neartica' in s or 'nearctic' in s:
        return 'neartica'
    return s

def norm_mig(v):
    if pd.isna(v): return ''
    s = str(v).strip().lower()
    s = s.replace('.', '')
    if s in ('res', 'residente', 'resident'):
        return 'res'
    if 'lat' in s or 'lat-trans' in s or 'lat trans' in s:
        return 'lat-trans'
    if 'nomad' in s:
        return 'nomad'
    if 'residente' in s:
        return 'res'
    return s

# crear columnas normalizadas
temp['_iucn_norm'] = temp['IUCN'].apply(norm_iucn)
temp['_res126_norm'] = temp['Res. 0126'].apply(norm_res126)
temp['_cites_norm'] = temp['CITES'].apply(norm_cites)
temp['_dist_norm'] = temp['Distribución'].apply(norm_dist)
temp['_mig_norm'] = temp['Migración'].apply(norm_mig)

# ahora definir neutros usando los códigos canónicos (todo en minúscula)
iucn_neutro = {'lc', 'ne', ''}            # LC y NO EVALUADO consideramos neutro
res126_neutro = {'lc', '', 'nl'}                # NL o vacío = neutro
cites_neutro = {'', 'no aplica', "sin registro en apéndice cites"}                       # vacío = sin CITES
dist_neutro = {'neotropical', 'cosmopolita','nomadismo', '','neartica, neotropical'}         # neotropical se considera neutro
mig_neutro = {'res', ''}                  # residente se considera neutro

condicion_neutra = (
    temp['_iucn_norm'].isin(iucn_neutro) &
    temp['_res126_norm'].isin(res126_neutro) &
    temp['_cites_norm'].isin(cites_neutro) &
    temp['_dist_norm'].isin(dist_neutro) &
    temp['_mig_norm'].isin(mig_neutro)
)

# Filtrar sobre la tabla original (sin modificar su formato)
tabla_sensible = temp.loc[~condicion_neutra, tabla_sensibilidad.columns].copy()

print(f" Especies sensibles detectadas: {len(tabla_sensible)} de {len(tabla_sensibilidad)}")
print(tabla_sensible.head(20))


# ---  Ordenar clases: primero Aves, luego Mammalia, luego el resto ---
orden_clase = ['Aves', 'Mammalia']
tabla_sensible['Clase'] = pd.Categorical(
    tabla_sensible['Clase'],
    categories=orden_clase + sorted(set(tabla_sensible['Clase']) - set(orden_clase)),
    ordered=True
)

# ---  Ordenar la tabla por Clase, Orden, Familia y Especie ---
tabla_sensible = tabla_sensible.sort_values(['Clase', 'Familia', 'Especie']).reset_index(drop=True)

# ---  Agregar numeración reiniciada por Clase ---
tabla_sensible['N°'] = tabla_sensible.groupby('Clase').cumcount() + 1

# Mover la columna "N°" al inicio
cols = ['N°'] + [col for col in tabla_sensible.columns if col != 'N°']
tabla_sensible = tabla_sensible[cols]

# ---  Insertar fila con nombres de columnas justo antes de 'Mammalia' ---
idx_mam = tabla_sensible.index[tabla_sensible['Clase'] == 'Mammalia']
if len(idx_mam) > 0:
    insert_pos = idx_mam[0]
    fila_header = pd.DataFrame([{col: str(col) for col in tabla_sensible.columns}])  # mantiene los nombres de columnas como texto
    tabla_sensible = pd.concat(
        [tabla_sensible.iloc[:insert_pos], fila_header, tabla_sensible.iloc[insert_pos:]],
        ignore_index=True
    )

# ---  Mostrar resultado ---
tabla_sensible.head(20)



# Guardar tabla en Excel
output_path = os.path.join(output_folder, "12_tabla_sensibilidad.xlsx")
tabla_sensible.to_excel(output_path, index=False)
print(f"\n Archivo guardado en: {output_path}")


#---------------------------------- Reparar y formatear archivo de tabla_sensibilidad -----------------------
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import os

# --- Rutas ---
ruta_original = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\12_tabla_sensibilidad.xlsx"
ruta_limpia = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\12_tabla_sensibilidad.xlsx"

# --- Verificar existencia ---
if not os.path.exists(ruta_original):
    raise FileNotFoundError(f"⚠️ No se encontró el archivo: {ruta_original}")

# --- Leer archivo dañado con pandas ---
try:
    df = pd.read_excel(ruta_original)
    print(" Archivo leído correctamente con pandas.")
except Exception as e:
    raise RuntimeError(f" No se pudo leer el archivo: {e}")

# --- Reescribir el archivo limpio ---
df.to_excel(ruta_limpia, index=False)
print(f" Archivo reparado y guardado como:\n{ruta_limpia}")

# --- Aplicar formato con openpyxl ---
from openpyxl import load_workbook

wb = load_workbook(ruta_limpia)
ws = wb.active

# --- Estilos base ---
header_fill = PatternFill(start_color='BFD8B8', end_color='BFD8B8', fill_type='solid')
header_font = Font(bold=True, color='000000', name='Calibri')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# --- Aplicar formato y reemplazar vacíos ---
for row in ws.iter_rows():
    for cell in row:
        if cell.value is None or str(cell.value).strip() == '':
            cell.value = '-'
        cell.alignment = center_align
        cell.border = thin_border

# --- Encabezado ---
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align

# --- Ajustar ancho de columnas ---
for col in ws.columns:
    max_length = 0
    column = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            length = len(str(cell.value))
            if length > max_length:
                max_length = length
    ws.column_dimensions[column].width = max_length + 3

# --- Ajustar altura de filas ---
for row in ws.iter_rows():
    ws.row_dimensions[row[0].row].height = 18

# --- Guardar cambios ---
wb.save(ruta_limpia)
print(f' Archivo formateado y reparado correctamente:\n{ruta_limpia}')





#-------------------------------------------------


import os
import pandas as pd

# Carpeta de salida 
ruta_salida = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados"
os.makedirs(ruta_salida, exist_ok=True)

# Copias
reg = Registros.copy()
sens = tabla_sensible.copy()

# Quitar columna N° si existe en la tabla sensible
if "N°" in sens.columns:
    sens = sens.drop(columns=["N°"])

# Obtener lista de municipios
municipios = sorted(reg["MUNICIPIO"].dropna().unique())

for muni in municipios:
    print(f"\n=== Procesando municipio: {muni} ===")

    # 1. Filtrar registros del municipio
    df_muni = reg[reg["MUNICIPIO"] == muni]

    # 2. Obtener especies únicas del municipio
    especies_muni = sorted(df_muni["Especie"].dropna().unique())

    # 3. Filtrar solo las especies sensibles presentes en ese municipio
    tabla_muni = sens[sens["Especie"].isin(especies_muni)].copy()

    # 4. Ordenar por nombre de especie
    tabla_muni = tabla_muni.sort_values("Especie").reset_index(drop=True)

    # 5. Reiniciar numeración N°
    tabla_muni.insert(0, "N°", range(1, len(tabla_muni) + 1))

    # 6. Guardar archivo
    archivo = os.path.join(ruta_salida, f"Especies_Sensibles_{muni}.xlsx")
    tabla_muni.to_excel(archivo, index=False)

    print(f"✔ Archivo generado: {archivo}")

print("\n✔ PROCESO COMPLETADO.")

print(temp['_iucn_norm'].value_counts())
print(temp['_res126_norm'].value_counts())
####-----------------------------------------------------------


import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

# Carpeta de salida
ruta_salida = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados"
os.makedirs(ruta_salida, exist_ok=True)

# Copias
reg = Registros.copy()
sens = tabla_sensible.copy()

# Si existe la columna N°, eliminarla
if "N°" in sens.columns:
    sens = sens.drop(columns=["N°"])

# ===========================
#   FUNCIÓN PARA FORMATEAR
# ===========================
def formatear_excel(ruta_archivo):

    wb = load_workbook(ruta_archivo)
    ws = wb.active

    # Estilos
    header_fill = PatternFill(start_color='BFD8B8', end_color='BFD8B8', fill_type='solid')
    header_font = Font(bold=True, color='000000', name='Calibri')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Formato general
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or str(cell.value).strip() == "":
                cell.value = "-"
            cell.alignment = center_align
            cell.border = thin_border

    # Formato encabezado
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Ajustar ancho columnas
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                value = str(cell.value)
                if len(value) > max_length:
                    max_length = len(value)
        ws.column_dimensions[column].width = max_length + 3

    # Ajustar altura filas
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 18

    wb.save(ruta_archivo)


# ===========================
#   GENERAR TABLAS POR MUNICIPIO
# ===========================

municipios = sorted(reg["MUNICIPIO"].dropna().unique())

for muni in municipios:

    print(f"\n=== Procesando municipio: {muni} ===")

    df_muni = reg[reg["MUNICIPIO"] == muni]

    especies_muni = sorted(df_muni["Especie"].dropna().unique())

    tabla_muni = sens[sens["Especie"].isin(especies_muni)].copy()

    tabla_muni = tabla_muni.sort_values("Especie").reset_index(drop=True)

    tabla_muni.insert(0, "N°", range(1, len(tabla_muni) + 1))

    archivo = os.path.join(ruta_salida, f"Especies_Sensibles_{muni}.xlsx")
    tabla_muni.to_excel(archivo, index=False)

    # Aplicar formato
    formatear_excel(archivo)

    print(f"✔ Archivo generado y formateado: {archivo}")

print("\n✔ PROCESO COMPLETADO.")




















































































#---------------------------------------------








output_folder = (r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados")

import pandas as pd
import os

# ==========================================================
# FUNCIÓN PRINCIPAL
# ==========================================================
def interpretar_sensibilidad(tabla_sensible, output_folder):

    # Convertir columnas clave a string y normalizar
    tabla_sensible['IUCN'] = tabla_sensible['IUCN'].astype(str)
    tabla_sensible['Res. 0126'] = tabla_sensible['Res. 0126'].astype(str)
    tabla_sensible['CITES'] = tabla_sensible['CITES'].astype(str)

    # -------------------------------------------------------------
    # 1. Métricas generales automáticas
    # -------------------------------------------------------------
    total_especies = len(tabla_sensible)
    clases = tabla_sensible['Clase'].value_counts()
    familias = tabla_sensible['Familia'].value_counts()

    # Amenaza IUCN (solo categorías reales)
    categorias_amenaza = tabla_sensible[
        tabla_sensible['IUCN'].isin(['VU', 'NT', 'EN', 'CR'])
    ]['IUCN'].value_counts()

    # CITES
    cites = tabla_sensible[
        tabla_sensible['CITES'].str.contains("Apendice", case=False)
    ]['CITES'].value_counts()

    # Resolución 0126 (MADS)
    r0126 = tabla_sensible[
        tabla_sensible['Res. 0126'].isin(['VU', 'EN', 'CR'])
    ]['Res. 0126'].value_counts()

    # Distribución geográfica
    distrib = tabla_sensible['Distribución'].value_counts()

    # Migración
    migracion = tabla_sensible['Migración'].value_counts()

    # -------------------------------------------------------------
    # 2. Interpretación general automática
    # -------------------------------------------------------------
    texto_general = f"""
==========================================================
 INTERPRETACIÓN AUTOMÁTICA DE SENSIBILIDAD FAUNÍSTICA
==========================================================

Se registró un total de **{total_especies} especies sensibles** pertenecientes a {len(clases)} clases 
y {len(familias)} familias taxonómicas. La mayor representación corresponde a:

- **Clases más abundantes:** 
  {', '.join([f"{c}: {n}" for c, n in clases.items()])}

- **Familias con mayor número de especies:** 
  {', '.join([f"{f}: {n}" for f, n in familias.head(5).items()])}

- **Especies incluidas en el análisis:**
  {", ".join(sorted(tabla_sensible['Especie'].unique()))}

"""

    # -------------------------------------------------------------
    # 3. Estado de amenaza
    # -------------------------------------------------------------
    texto_amenaza = "2. --- Estado de amenaza (IUCN, CITES, Res.0126) ---\n\n"

    # IUCN
    if categorias_amenaza.empty:
        texto_amenaza += "No se registraron especies con categoría de amenaza IUCN.\n\n"
    else:
        texto_amenaza += "La evaluación IUCN indica presencia de especies sensibles:\n"
        for cat, n in categorias_amenaza.items():
            especies_cat = tabla_sensible[tabla_sensible['IUCN'] == cat]['Especie'].unique()
            texto_amenaza += (
                f" - {cat}: {n} especies\n"
                f"   Especies: {', '.join(especies_cat)}\n"
            )
        texto_amenaza += "\n"

    # CITES
    if cites.empty:
        texto_amenaza += "No se registraron especies incluidas en CITES.\n\n"
    else:
        texto_amenaza += "Regulaciones CITES detectadas:\n"
        for c, n in cites.items():
            especies_c = tabla_sensible[tabla_sensible['CITES'] == c]['Especie'].unique()
            texto_amenaza += (
                f" - {c}: {n} especies\n"
                f"   Especies: {', '.join(especies_c)}\n"
            )
        texto_amenaza += "\n"

    # Res. 0126
    if r0126.empty:
        texto_amenaza += "Ninguna especie está bajo categoría de amenaza según la Resolución 0126.\n\n"
    else:
        texto_amenaza += "Especies protegidas bajo Resolución 0126:\n"
        for cat, n in r0126.items():
            especies_r = tabla_sensible[tabla_sensible['Res. 0126'] == cat]['Especie'].unique()
            texto_amenaza += (
                f" - {cat}: {n} especies\n"
                f"   Especies: {', '.join(especies_r)}\n"
            )
        texto_amenaza += "\n"


    # -------------------------------------------------------------
    # 4. Distribución biogeográfica
    # -------------------------------------------------------------
    texto_distrib = "3. --- Patrones de distribución ---\n\n"

    texto_distrib += "Las especies registradas abarcan los siguientes patrones biogeográficos:\n"
    for d, n in distrib.items():
        especies_d = tabla_sensible[tabla_sensible['Distribución'] == d]['Especie'].unique()
        texto_distrib += (
            f" - {d}: {n} especies\n"
            f"   Especies: {', '.join(especies_d)}\n"
        )

    texto_distrib += "\n"


    # -------------------------------------------------------------
    # 5. Migración
    # -------------------------------------------------------------
    texto_migracion = "4. --- Patrones de migración ---\n\n"

    texto_migracion += "La comunidad presenta los siguientes patrones migratorios:\n"
    for m, n in migracion.items():
        especies_m = tabla_sensible[tabla_sensible['Migración'] == m]['Especie'].unique()
        texto_migracion += (
            f" - {m}: {n} especies\n"
            f"   Especies: {', '.join(especies_m)}\n"
        )

    texto_migracion += "\n"


    # -------------------------------------------------------------
    # 6. Síntesis ecológica integrada
    # -------------------------------------------------------------
    texto_sintesis = f"""
5. --- Síntesis ecológica integrada ---

La estructura faunística registrada combina especies residentes, migratorias y de amplia
distribución, junto con especies reguladas por CITES y algunas incluidas en categorías de 
amenaza. Esto indica que el área funciona como un nodo ecológico relevante para múltiples 
grupos funcionales.

Las especies evaluadas incluyen:
{", ".join(sorted(tabla_sensible['Especie'].unique()))}

La presencia de especies en CITES o con categoría IUCN eleva la importancia de conservación,
mientras que los patrones migratorios sugieren que el área cumple funciones críticas dentro
de corredores biológicos regionales y continentales.
"""

    # Unir texto
    texto_final = texto_general + texto_amenaza + texto_distrib + texto_migracion + texto_sintesis

    # -------------------------------------------------------------
    # GUARDAR TXT
    # -------------------------------------------------------------
    os.makedirs(output_folder, exist_ok=True)
    output_path = os.path.join(output_folder, "12.1_Interpretacion_Sensibilidad.txt")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(texto_final)

    print(f"\n Interpretación generada y guardada en:\n {output_path}\n")
    return texto_final


interpretar_sensibilidad(tabla_sensible, output_folder)

























































































































































import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.cm import get_cmap

df = tabla_sensible.copy()

# =============================
# CONTAR ESPECIES POR CATEGORÍA
# =============================
conteo_iucn = df["IUCN"].value_counts().sort_index()
conteo_migra = df["Migración"].value_counts().sort_index()
conteo_distri = df["Distribución"].value_counts().sort_index()
conteo_cites = df["CITES"].value_counts().sort_index()

categorias = [
    ("IUCN", conteo_iucn),
    ("Migración", conteo_migra),
    ("Distribución", conteo_distri),
    ("CITES", conteo_cites)
]

# =============================
# CONFIGURAR FIGURA GENERAL (2x2)
# =============================
fig, axes = plt.subplots(2, 2, figsize=(11, 11))
axes = axes.flatten()

# Paleta de colores tab20
cmap = get_cmap("Paired")

# =============================
# GRAFICAR
# =============================
for ax, (titulo, serie) in zip(axes, categorias):

    # Seleccionar colores únicos por barra
    colores = [cmap(i) for i in range(len(serie))]

    # Gráfico de barras
    barras = ax.bar(serie.index, serie.values, color=colores, edgecolor="black", linewidth=1)

    # Título
    ax.set_title(titulo, fontsize=16, fontweight="bold")

    # Eje Y
    ax.set_ylabel("Número de especies", fontsize=12)

    # Evitar que los datos toquen la línea superior
    max_val = max(serie.values)
    ax.set_ylim(0, max_val * 1.20)

    # Ticks X
    ax.set_xticks(range(len(serie.index)))
    ax.set_xticklabels(serie.index, rotation=45, ha="right", fontsize=10)

    # NÚMEROS SOBRE BARRAS (separados del borde)
    for b in barras:
        h = b.get_height()
        ax.text(
            b.get_x() + b.get_width()/2,
            h + max_val * 0.03,  # separación extra
            f"{int(h)}",
            ha="center",
            va="bottom",
            fontsize=11,
            fontweight="bold"
        )

# =============================
# AJUSTE GENERAL DE ESPACIADOS
# =============================
plt.tight_layout(pad=3)
plt.subplots_adjust(wspace=0.25, hspace=0.35)

# =============================
# GUARDAR ARCHIVO
# =============================

from pathlib import Path

# Carpeta de salida
ruta_salida = Path(r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados")

# Crear carpeta si no existe
ruta_salida.mkdir(parents=True, exist_ok=True)

nombre_archivo = "12.1_Grilla_IUCN_Migracion_Distribucion_CITES_Color.png"

# Ruta completa del archivo
ruta_completa = ruta_salida / nombre_archivo

# Guardar
plt.savefig(ruta_completa, dpi=300, bbox_inches="tight")
plt.show()

print("✔ Gráfica guardada en:", ruta_completa)





























#-------------------Graficar--------------------------

import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.cm import tab20

df = tabla_sensible.copy()

# =============================
# CONTAR ESPECIES POR CATEGORÍA
# =============================
conteo_iucn = df["IUCN"].value_counts().sort_index()
conteo_migra = df["Migración"].value_counts().sort_index()
conteo_distri = df["Distribución"].value_counts().sort_index()
conteo_cites = df["CITES"].value_counts().sort_index()

# =============================
# CONFIGURAR 4 SUBPLOTS
# =============================
fig, axes = plt.subplots(1, 4, figsize=(32, 8))

categorias = [
    ("IUCN", conteo_iucn),
    ("Migración", conteo_migra),
    ("Distribución", conteo_distri),
    ("CITES", conteo_cites)
]

# =============================
# GRAFICAR CADA PANEL
# =============================
for ax, (titulo, serie) in zip(axes, categorias):
    
    # Colores tab20 (tantos como categorías existan)
    colores = [tab20(i / 20) for i in range(len(serie.index))]

    # Crear barras con color
    barras = ax.bar(serie.index, serie.values, color=colores)

    # Título
    ax.set_title(titulo, fontsize=16, fontweight="bold")
    ax.set_ylabel("Número de especies", fontsize=12)

    # Configurar ticks
    ax.set_xticks(range(len(serie.index)))
    ax.set_xticklabels(serie.index, rotation=45, ha="right", fontsize=10)

    # Evitar que los datos toquen el borde superior (+10% de límite)
    ax.set_ylim(0, max(serie.values) * 1.15)

    # Números encima de columnas
    for barra in barras:
        altura = barra.get_height()
        ax.text(
            barra.get_x() + barra.get_width() / 2,
            altura + max(serie.values) * 0.03,
            f"{int(altura)}",
            ha="center",
            va="bottom",
            fontsize=11,
            fontweight="bold"
        )

# =================================
# AJUSTES DE DISEÑO
# =================================
plt.tight_layout(pad=2)
plt.subplots_adjust(wspace=0.30)

# =============================
# =============================
# GUARDAR ARCHIVO
# =============================
from pathlib import Path

# Ruta a la carpeta de resultados
ruta_salida = Path(r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados")

# Crear carpeta si no existe
ruta_salida.mkdir(parents=True, exist_ok=True)

# Nombre del archivo
nombre_archivo = "12.1_Grafica_IUCN_Migracion_Distribucion_CITES.png"

# Ruta completa
ruta_completa = ruta_salida / nombre_archivo

# Guardar el archivo
plt.savefig(ruta_completa, dpi=300, bbox_inches="tight")

print("✔ Gráfica guardada como:", ruta_completa)

plt.show()






