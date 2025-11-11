#-------------------------Dar formato al archivo Estimadores_Abundancia.xlsx-------------------------#
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import os

# --- Rutas ---
ruta_original = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\Estimadores_Abundancia.xlsx"
ruta_limpia = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\Estimadores_Abundancia.xlsx"

# --- Verificar existencia ---
if not os.path.exists(ruta_original):
    raise FileNotFoundError(f"⚠️ No se encontró el archivo: {ruta_original}")

# --- Leer archivo dañado con pandas ---
try:
    df = pd.read_excel(ruta_original)
    print("✅ Archivo leído correctamente con pandas.")
except Exception as e:
    raise RuntimeError(f"❌ No se pudo leer el archivo: {e}")

# --- Reescribir el archivo limpio ---
df.to_excel(ruta_limpia, index=False)
print(f"🧹 Archivo reparado y guardado como:\n{ruta_limpia}")

# --- Aplicar formato con openpyxl ---
from openpyxl import load_workbook

wb = load_workbook(ruta_limpia)
ws = wb.active

# --- Estilos base ---
header_fill = PatternFill(start_color='BFD8B8', end_color='BFD8B8', fill_type='solid')
header_font = Font(bold=True, color='000000', name='Calibri')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# --- Aplicar formato y reemplazar vacíos ---
for row in ws.iter_rows():
    for cell in row:
        if cell.value is None or str(cell.value).strip() == '':
            cell.value = '-'
        cell.alignment = center_align
        cell.border = thin_border

# --- Encabezado ---
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align

# --- Ajustar ancho de columnas ---
for col in ws.columns:
    max_length = 0
    column = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            length = len(str(cell.value))
            if length > max_length:
                max_length = length
    ws.column_dimensions[column].width = max_length + 3

# --- Ajustar altura de filas ---
for row in ws.iter_rows():
    ws.row_dimensions[row[0].row].height = 18

# --- Guardar cambios ---
wb.save(ruta_limpia)
print(f'📘 Archivo formateado y reparado correctamente:\n{ruta_limpia}')

#-------------------------Fin Dar formato al archivo Estimadores_Abundancia.xlsx-------------------------#

#-----------------------leer el archivo formateado-----------------------#
import pandas as pd
tabla_Abund = pd.read_excel(r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\Estimadores_Abundancia.xlsx")

tabla_Abund

names = tabla_Abund.columns.tolist()
names


#-----------------------Calcular efectividad de los estimadores de abundancia-----------------------#
import pandas as pd

# --- Cargar datos ---
ruta = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\Estimadores_Abundancia.xlsx"
tabla_Abund = pd.read_excel(ruta)

# --- Calcular efectividad para cada estimador ---
estimadores = ['Chao1_Chao_1984_Mean', 'Chao1_bc_Mean', 'iChao1_Chiu_et_al_2014_Mean',
               'ACE_Chao_Lee_1992_Mean', 'ACE_1_Chao_Lee_1992_Mean', '1st_order_jackknife_Mean', '2nd_order_jackknife_Mean', 
               'Bootstrap_Mean',]  # ajusta según tus columnas reales

efectividad = pd.DataFrame()
efectividad['Unidad'] = tabla_Abund['Unidad']
efectividad['Observadas_Mean'] = tabla_Abund['Observadas_Mean']

for est in estimadores:
    if est in tabla_Abund.columns:
        efectividad[est.replace('_Mean', '_Efectividad_%')] = (
            (tabla_Abund['Observadas_Mean'] / tabla_Abund[est]) * 100
        )

# --- Calcular  total de efectividad por estimador ---
# (último valor registrado en cada columna para cada estimador)
resumen = (
    efectividad
    .drop(columns=['Unidad', 'Observadas_Mean'])
    .tail(1)  # ✅ toma la última fila (la efectividad final)
    .melt(var_name='Estimador', value_name='Efectividad_Promedio_%')
    .reset_index(drop=True)
)

# --- Seleccionar los dos mejores estimadores según efectividad ---
resumen = resumen.sort_values(by='Efectividad_Promedio_%', ascending=False)

import pandas as pd

# --- Agrupar por tipo de estimador y obtener el de mayor efectividad ---
resumen['Grupo'] = resumen['Estimador'].apply(
    lambda x: (
        'Chao' if 'Chao' in x else
        'ACE' if 'ACE' in x else
        'Jackknife' if 'jackknife' in x.lower() else
        'Bootstrap' if 'Bootstrap' in x else
        'Otro'
    )
)

# --- Seleccionar el mejor (mayor efectividad) por grupo ---
mejores_por_grupo = (
    resumen.sort_values(by='Efectividad_Promedio_%', ascending=False)
           .groupby('Grupo')
           .head(1)   # uno por grupo
           .reset_index(drop=True)
)

# --- Ajustar nombres al formato de las columnas de la tabla de datos ---
top_estimadores = mejores_por_grupo['Estimador'].str.replace('_Efectividad_%', '_Mean').tolist()


resumen = pd.DataFrame(mejores_por_grupo)

print("🏆 Mejores estimadores por grupo:")
print(mejores_por_grupo[['Grupo', 'Estimador', 'Efectividad_Promedio_%']])
print("\n📊 Nombres finales para graficar:", top_estimadores)


#-----------------------Fin Calcular efectividad de los estimadores de abundancia-----------------------#
#-----------------------Guardar tabla de efectividad-----------------------#

ruta_salida = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\Efectividad_Estimadores.xlsx"
with pd.ExcelWriter(ruta_salida, engine='openpyxl') as writer:
    efectividad.to_excel(writer, sheet_name='Por_Unidad', index=False)
    resumen.to_excel(writer, sheet_name='Resumen_Efectividad', index=False)

print("✅ Tabla de efectividad exportada correctamente.")

#-----------------------Fin Guardar tabla de efectividad-----------------------#
#-----------------------Graficar curvas de acumulacion de especies-----------------------#
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np

# --- Escoger los dos mejores estimadores ---
top2 = resumen['Estimador'].head(3).str.replace('_Efectividad_%', '_Mean').tolist()
print("📊 Mejores estimadores:", top2)

# --- Crear figura ---
fig, ax = plt.subplots(figsize=(10, 6))

# Eje X dinámico según número de unidades
x = np.arange(1, len(tabla_Abund) + 1)

# --- Función automática de etiquetado sin solapamientos ---
etiquetas_previas = []

def colocar_etiqueta_automatica(x_val, y_val, texto, ax):
    ymin, ymax = ax.get_ylim()
    offset = (ymax - ymin) * 0.05
    for y_prev in etiquetas_previas:
        if abs(y_prev - y_val) < offset:
            y_val += offset
    etiquetas_previas.append(y_val)
    y_val = np.clip(y_val, ymin + offset, ymax - offset)
    x_val = min(x_val, ax.get_xlim()[1] - 0.5)
    ax.text(x_val + 0.2, y_val, f"{float(texto):.1f}",
            fontsize=9, ha='left', va='center', color='black')


# 🔹 1️⃣ Agrega este bloque justo antes de graficar “Observadas”
# Calcula el número total observado
n_obs = int(tabla_Abund['Observadas_Mean'].iloc[-1])

# --- Dibujar observadas ---
# 🔹 2️⃣ Cambia solo la etiqueta del label:
ax.plot(x, tabla_Abund['Observadas_Mean'], 'o-', color='black',
        label=f"Observadas ({n_obs} spp)")
colocar_etiqueta_automatica(x[-1], tabla_Abund['Observadas_Mean'].iloc[-1],
                            tabla_Abund['Observadas_Mean'].iloc[-1], ax)


# --- Dibujar los dos mejores estimadores ---
# 🔹 3️⃣ Aquí agregas la búsqueda de efectividad para cada estimador:
for est in top2:
    est_base = est.replace('_Mean', '_Efectividad_%')
    efectividad = resumen.loc[resumen['Estimador'] == est_base, 'Efectividad_Promedio_%'].values
    ef_txt = f" ({efectividad[0]:.1f}%)" if len(efectividad) > 0 else ""

    # 🔹 4️⃣ Modifica el label para que incluya la efectividad
    ax.plot(x, tabla_Abund[est], 'o--', label=est.replace('_Mean', '') + ef_txt)
    colocar_etiqueta_automatica(x[-1], tabla_Abund[est].iloc[-1],
                                tabla_Abund[est].iloc[-1], ax)


# --- Dibujar desviación estándar de Singletons ---
if 'Singletons_SD' in tabla_Abund.columns:
    ax.plot(x, tabla_Abund['Singletons_SD'], 'o--', color='gray', linewidth=2,
            label='Singletons_SD')
    colocar_etiqueta_automatica(x[-1], tabla_Abund['Singletons_SD'].iloc[-1],
                                tabla_Abund['Singletons_SD'].iloc[-1], ax)


# --- Ajustes automáticos del gráfico ---
ax.set_xlim(0.5, len(x) + 0.8)
ax.margins(y=0.1)
plt.title("Curva de acumulación de especies - Basada en Abundancias", fontsize=14)
plt.xlabel("Unidades de muestreo")
plt.ylabel("Riqueza estimada")
plt.legend()
plt.grid(True, linestyle='--', alpha=0.5)
plt.tight_layout()
plt.show()
#-----------------------Fin ajustes automáticos del gráfico ---#

# --- Guardar la gráfica en formato PNG ---
fig.savefig("D:/CORPONOR 2025/Backet/python_Proyect/Resultados/estimadores_riqueza.png",
            dpi=300, bbox_inches='tight', transparent=False)
print("✅ Gráfica guardada correctamente.")

#-----------------------Fin Graficar curvas de acumulacion de especies-----------------------#

#-----------------------Crear tabla resumen de estimadores-----------------------#
import pandas as pd

# --- Crear tabla resumen de efectividad ---
datos_tabla = []

# Valor observado final
obs_final = tabla_Abund['Observadas_Mean'].iloc[-1]
datos_tabla.append({
    "Estimador": "Observadas",
    "Individuos_estimados": obs_final,
    "Efectividad_%": None  # sin porcentaje
})

# Los estimadores del gráfico (automático según top2)
efectividades = []
for est in top2:
    valor_final = tabla_Abund[est].iloc[-1]
    # 🔹 Porcentaje de representatividad del observado respecto al estimado
    efectividad = (obs_final / valor_final) * 100
    efectividades.append(efectividad)
    datos_tabla.append({
        "Estimador": est.replace('_Mean', ''),
        "Individuos_estimados": valor_final,
        "Efectividad_%": efectividad
    })

# Agregar fila de promedio de efectividad
promedio_efectividad = sum(efectividades) / len(efectividades)
datos_tabla.append({
    "Estimador": "Promedio efectividad",
    "Individuos_estimados": None,
    "Efectividad_%": promedio_efectividad
})

# Convertir a DataFrame
tabla_resumen = pd.DataFrame(datos_tabla)

# --- Mostrar con formato redondeado ---
print("\n📋 Resumen de estimadores (valores finales):")
print(tabla_resumen.round(2).to_string(index=False))

# --- Guardar en Excel ---
ruta_salida = "D:/CORPONOR 2025/Backet/python_Proyect/Resultados/Resumen_estimadores.xlsx"
tabla_resumen.to_excel(ruta_salida, index=False)

print(f"\n✅ Archivo Excel guardado en:\n{ruta_salida}")
