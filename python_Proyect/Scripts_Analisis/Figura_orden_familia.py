#------------------Figura de Ordenes familias---------------

import matplotlib
matplotlib.use('Agg')
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# --- Copiar dataframe base ---
df = Registros.copy()

# --- Si tu DataFrame ya está cargado en Python como df, úsalo directamente ---
tabla = df.copy()

# --- Limpiar nombres de columnas por seguridad ---
tabla.columns = tabla.columns.str.strip()

# Limpiar nombres de familia
tabla['Familia'] = tabla['Familia'].astype(str).str.strip()

# Eliminar filas con nombres de familia vacíos o 'nan'
tabla = tabla[tabla['Familia'].notna()]
tabla = tabla[tabla['Familia'] != '']


# --- Asegurar que las columnas requeridas existen ---
tabla = tabla.dropna(subset=['Orden', 'Familia', 'Especie'])

# --- Crear tabla dinámica: número de especies únicas por Orden y Familia ---
pivot_df = (
    tabla.groupby(['Orden', 'Familia'])['Especie']
    .nunique()
    .reset_index()
    .pivot(index='Orden', columns='Familia', values='Especie')
    .fillna(0)
    .astype(int)
)
# --- Ruta de salida ---
output_folder = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados"
os.makedirs(output_folder, exist_ok=True)  # Crea la carpeta si no existe

# --- Exportar tabla a Excel ---
excel_path = os.path.join(output_folder, 'Riqueza_Orden_Familia.xlsx')
pivot_df.to_excel(excel_path, sheet_name='Tabla_dinamica')


# --- Crear gráfico de barras apiladas horizontal ---
sns.set(style='whitegrid')
fig, ax = plt.subplots(figsize=(12, 8))

pivot_df.plot(
    kind='barh',
    stacked=True,
    colormap='tab20',
    edgecolor='black',
    ax=ax
)

# --- Añadir etiquetas dentro de las barras ---
for container in ax.containers:
    # etiquetas solo si el valor del segmento > 0
    labels = [f'{w.get_width():.0f}' if w.get_width() > 0 else '' for w in container]
    ax.bar_label(
        container,
        labels=labels,
        label_type='center',     # posición centrada dentro del bloque
        fontsize=7,
        color='black',
        weight='bold'
    )

# --- Etiquetas y formato ---
ax.set_title('Riqueza de especies por Orden y Familia', fontsize=14, fontweight='bold')
ax.set_xlabel('Número de especies')
ax.set_ylabel('Orden')
# --- Ajuste de la leyenda para ocupar todo el alto ---
ax.legend(
    title='Familia',
    bbox_to_anchor=(1.02, 0, 0.25, 1),  # [x0, y0, ancho, alto] → ocupa toda la altura
    loc='upper left',
    ncol=2,                             # número de columnas
    fontsize=8,
    title_fontsize=9,
    frameon=False,
    mode='expand',                      # distribuye las entradas verticalmente en todo el alto
    borderaxespad=0.0,
    columnspacing=1.2,
    labelspacing=0.8
)

plt.tight_layout()


# --- Ruta de salida ---
output_folder = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados"
os.makedirs(output_folder, exist_ok=True)  # crea la carpeta si no existe

# --- Guardar el gráfico como imagen en la carpeta Resultados ---
img_path = os.path.join(output_folder, 'Grafico_Riqueza_Orden_Familia.png')
plt.savefig(img_path, dpi=300, bbox_inches='tight')
plt.close()

# --- Insertar el gráfico en el Excel ---
excel_path = os.path.join(output_folder, 'Riqueza_Orden_Familia.xlsx')
wb = load_workbook(excel_path)
ws = wb.create_sheet('Grafico')

# Insertar la imagen
img = Image(img_path)
ws.add_image(img, 'A1')

# Guardar el Excel final
wb.save(excel_path)

print('✅ Tabla dinámica y gráfico exportados en:', excel_path)


