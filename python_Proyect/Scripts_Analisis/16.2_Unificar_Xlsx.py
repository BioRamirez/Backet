#!/usr/bin/env python3
import os
import math
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# --- CONFIG: ajusta si tu carpeta Results está en otra ruta
RESULTADOS_DIR = Path(r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados")
RESULTADOS_FIN = Path(r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados_Final")
SALIDA_XLSX = RESULTADOS_FIN / "POF_Zulia_Resultado_Final_.xlsx"

# Extensiones a incluir / ignorar
EXT_DATOS = {".xlsx", ".xls", ".csv"}
EXT_IMAGEN = {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".gif"}
IGNORAR_EXT = {".pdf", ".doc", ".docx", ".txt"}

if not RESULTADOS_DIR.exists():
    raise SystemExit(f"Carpeta no encontrada: {RESULTADOS_DIR}")

# Recolectar archivos en 'Resultados' (no recursivo), ordenar por ctime (creación)
archivos = [p for p in RESULTADOS_DIR.iterdir() if p.is_file()]
archivos.sort(key=lambda p: p.stat().st_ctime)  # orden por creación

# Filtrar: mantener solo datos e imágenes, ignorar PDFs/DOCs
archivos = [p for p in archivos if p.suffix.lower() not in IGNORAR_EXT]

# Crear workbook
wb = Workbook()
default_sheet = wb.active
default_sheet.title = "Resumen"  # la mantendremos y la borraremos si agregamos otras hojas

# Helper: generar nombre de hoja válido y único (<=31 chars)
def sheet_name_unique(base, existing):
    base_clean = base[:31]
    if base_clean not in existing:
        return base_clean
    # si ya existe, agregar sufijo numérico corto
    i = 1
    while True:
        candidate = f"{base_clean[:28]}_{i}"[:31]
        if candidate not in existing:
            return candidate
        i += 1

# Escribir DataFrame a hoja (openpyxl)
def write_df_to_sheet(ws, df):
    # Escribir encabezados
    ws.append(list(df.columns))
    # Escribir filas (convertir NaN a empty)
    for row in df.itertuples(index=False, name=None):
        ws.append([("" if (pd.isna(x)) else x) for x in row])
    # Ajustar ancho de columnas (simple heurística)
    for i, col in enumerate(df.columns, start=1):
        maxlen = max(
            (len(str(x)) for x in df[col].astype(str).head(200)), default=len(str(col))
        )
        width = min(50, max(8, int(maxlen) + 2))
        ws.column_dimensions[get_column_letter(i)].width = width

# Llevar control de nombres de hojas creadas
hojas_existentes = {default_sheet.title}

# Contador para saber si hemos puesto al menos una hoja real
hojas_agregadas = 0

for p in archivos:
    ext = p.suffix.lower()
    basename = p.stem  # nombre sin extensión
    created = datetime.fromtimestamp(p.stat().st_ctime).isoformat(sep=" ", timespec="seconds")
    print(f"Procesando: {p.name}  (creado: {created})")

    if ext in EXT_DATOS:
        # Leer Excel/CSV/TXT
        try:
            if ext in {".xlsx", ".xls"}:
                # leer todas las hojas
                sheets = pd.read_excel(p, sheet_name=None)
                # si el archivo tiene solo 1 hoja, la usaremos con el nombre del archivo
                if len(sheets) == 1:
                    key, df = next(iter(sheets.items()))
                    sheet_base = basename
                    sheet_name = sheet_name_unique(sheet_base, hojas_existentes)
                    ws = wb.create_sheet(title=sheet_name)
                    write_df_to_sheet(ws, df)
                    hojas_existentes.add(sheet_name)
                    hojas_agregadas += 1
                else:
                    # varias hojas -> prefijar con nombre de archivo
                    for key, df in sheets.items():
                        sheet_base = f"{basename}_{key}"
                        sheet_name = sheet_name_unique(sheet_base, hojas_existentes)
                        ws = wb.create_sheet(title=sheet_name)
                        write_df_to_sheet(ws, df)
                        hojas_existentes.add(sheet_name)
                        hojas_agregadas += 1

            else:
                # csv o txt (autodetect delimiter)
                df = pd.read_csv(p, encoding="utf-8", sep=None, engine="python")
                sheet_name = sheet_name_unique(basename, hojas_existentes)
                ws = wb.create_sheet(title=sheet_name)
                write_df_to_sheet(ws, df)
                hojas_existentes.add(sheet_name)
                hojas_agregadas += 1

        except Exception as e:
            # si falla la lectura con pandas, crear hoja con mensaje de error
            sheet_name = sheet_name_unique(f"{basename}_ERROR", hojas_existentes)
            ws = wb.create_sheet(title=sheet_name)
            ws["A1"] = f"ERROR leyendo {p.name}"
            ws["A2"] = str(e)
            hojas_existentes.add(sheet_name)
            hojas_agregadas += 1

    elif ext in EXT_IMAGEN:
        # Insertar imagen en hoja nueva
        try:
            sheet_name = sheet_name_unique(basename, hojas_existentes)
            ws = wb.create_sheet(title=sheet_name)
            img = XLImage(str(p))
            # ajustar tamaño máximos razonables (mantener relación)
            max_w, max_h = 1000, 700  # en píxeles aproximados
            # si la imagen es demasiado grande, reducirla (openpyxl Image soporta width/height)
            try:
                if img.width > max_w:
                    scale = max_w / img.width
                    img.width = int(img.width * scale)
                    img.height = int(img.height * scale)
                if img.height > max_h:
                    scale = max_h / img.height
                    img.width = int(img.width * scale)
                    img.height = int(img.height * scale)
            except Exception:
                pass

            # colocar en celda A1
            ws.add_image(img, "A1")
            hojas_existentes.add(sheet_name)
            hojas_agregadas += 1

        except Exception as e:
            sheet_name = sheet_name_unique(f"{basename}_IMG_ERROR", hojas_existentes)
            ws = wb.create_sheet(title=sheet_name)
            ws["A1"] = f"ERROR insertando imagen {p.name}"
            ws["A2"] = str(e)
            hojas_existentes.add(sheet_name)
            hojas_agregadas += 1

    else:
        # Si llegamos aquí es por una extensión desconocida; la saltamos.
        print(f"  → Ignorado (ext no procesada): {p.suffix}")

# Borrar hoja resumen por defecto si ya agregamos otras hojas
if hojas_agregadas > 0:
    try:
        # si la hoja "Resumen" está vacía la borramos
        wb.remove(wb["Resumen"])
    except Exception:
        # intentar eliminar la hoja activa si no hay necesidad
        pass

# Guardar
if hojas_agregadas == 0:
    print("No se encontraron archivos de datos ni imágenes en la carpeta Resultados.")
else:
    wb.save(SALIDA_XLSX)
    print(f"\n✅ Archivo final guardado en:\n{SALIDA_XLSX}\n")


#---------------------------Agregar datos de entrada---------------------------------

import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

# ------------------------------
# RUTAS
# ------------------------------
DIR_DATA = Path(r"D:\CORPONOR 2025\Backet\python_Proyect\data")
SALIDA_XLSX = Path(r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados_Final\POF_Zulia_Resultado_Final_.xlsx")

# ------------------------------
# 1. Obtener archivos de /data
# ------------------------------
archivos_data = sorted(
    [f for f in DIR_DATA.iterdir() if f.suffix.lower() in {".xlsx", ".xls", ".csv"}],
    key=lambda x: x.name
)

# ------------------------------
# 2. Cargar el Excel final existente
# ------------------------------
wb = load_workbook(SALIDA_XLSX)

# ------------------------------
# 3. Insertar las hojas de /data como primeras
# ------------------------------
for idx, file in enumerate(archivos_data):

    # Cargar el archivo como DataFrame
    if file.suffix.lower() in {".xlsx", ".xls"}:
        df = pd.read_excel(file)
    else:
        df = pd.read_csv(file)

    # Crear una hoja nueva al inicio
    nombre_hoja = f"DATA_{file.stem}"[:31]
    ws = wb.create_sheet(title=nombre_hoja, index=idx)

    # ---- 🔹 Agregar encabezados ----
    ws.append(list(df.columns))

    # ---- 🔹 Agregar filas de datos ----
    for fila in df.itertuples(index=False, name=None):
        ws.append(list(fila))

# ------------------------------
# 4. Guardar Excel final modificado
# ------------------------------
wb.save(SALIDA_XLSX)

print("\n✅ Se agregaron correctamente las hojas con ENCABEZADOS al inicio en:")
print(SALIDA_XLSX)























#---------------------------------- Reparar y formatear archivo final ----------------------------------
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

# --- Rutas ---
ruta_original = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados_Final\POF_Zulia_Resultado_Final_.xlsx"
ruta_limpia = ruta_original  # se sobrescribe

# --- Verificar existencia ---
if not os.path.exists(ruta_original):
    raise FileNotFoundError(f"⚠️ No se encontró el archivo: {ruta_original}")

print("\n🔍 Reparando archivo...")

# --- Cargar el archivo completo con pandas (solo la primera hoja) para verificar que no esté corrupto ---
try:
    pd.read_excel(ruta_original, sheet_name=0)
    print("✓ Archivo leído sin errores.")
except Exception as e:
    raise RuntimeError(f"❌ No se pudo leer el archivo: {e}")

print("✓ Reparación inicial terminada.")

# --- Abrir para formateo con openpyxl ---
wb = load_workbook(ruta_limpia)

# --- Estilos base ---
header_fill = PatternFill(start_color='BFD8B8', end_color='BFD8B8', fill_type='solid')
header_font = Font(bold=True, color='000000', name='Calibri')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

print("\n🎨 Aplicando formato a todas las hojas...\n")

# -------------------------------------------------------------------------
# 🚨 NOTA IMPORTANTE:
#   Hojas con IMÁGENES se deben saltar (no tienen celdas con datos)
# -------------------------------------------------------------------------

for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    print(f" → Formateando hoja: {sheetname}")

    # Detectar si la hoja contiene solo una imagen
    solo_imagen = False
    if ws.max_row == 1 and ws.max_column == 1:
        if ws["A1"].value is None and ws._images:
            solo_imagen = True

    if solo_imagen:
        print("    (Hoja con imagen — sin formateo de celdas)\n")
        continue

    # --- Aplicar formato a todas las celdas ---
    for row in ws.iter_rows():
        for cell in row:
            # Completar vacíos
            if cell.value is None or str(cell.value).strip() == "":
                cell.value = "-"
            cell.alignment = center_align
            cell.border = thin_border

    # --- Encabezados ---
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # --- Ajustar ancho de columnas ---
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 3, 50)

    # --- Ajustar altura de filas ---
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 18

print("\n💾 Guardando archivo final...")
wb.save(ruta_limpia)

print(f"\n✅ Archivo formateado correctamente:\n{ruta_limpia}\n")









#---------------------------PDF------------------------













import os
from pathlib import Path
from datetime import datetime
from docx2pdf import convert as docx_to_pdf
from reportlab.pdfgen import canvas
from PyPDF2 import PdfMerger

# ---------------------------------------
# CONFIGURACIÓN
# ---------------------------------------
RESULTADOS_DIR = Path(r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados")
RESULTADOS_FIN = Path(r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados_Final")

RESULTADOS_FIN.mkdir(exist_ok=True)

IGNORAR_EXT = {".pdf", ".doc", ".docx", ".txt"}

PDF_SALIDA = RESULTADOS_FIN / "POF_Zulia_Fauna_Interpretacion_Unificado.pdf"
CARPETA_TEMP = RESULTADOS_FIN / "_TEMP_PDF"

CARPETA_TEMP.mkdir(exist_ok=True)

# ---------------------------------------
# FUNCIONES AUXILIARES
# ---------------------------------------

def convertir_txt_a_pdf(txt_path, pdf_path):
    """Convierte un archivo TXT a PDF usando reportlab."""
    c = canvas.Canvas(str(pdf_path))
    c.setFont("Helvetica", 11)

    y = 800
    with open(txt_path, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            c.drawString(40, y, line.strip())
            y -= 15
            if y < 40:  # Nueva página
                c.showPage()
                c.setFont("Helvetica", 11)
                y = 800

    c.save()


def convertir_doc_docx_a_pdf(doc_path, pdf_path):
    """Convierte DOC o DOCX a PDF usando Word vía docx2pdf."""
    docx_to_pdf(str(doc_path), str(pdf_path))


# ---------------------------------------
# 1. Detectar archivos ignorados
# ---------------------------------------
print("\n📁 Buscando archivos ignorados para conversion...\n")

archivos_ignorar = []

for file in RESULTADOS_DIR.iterdir():
    if file.is_file() and file.suffix.lower() in IGNORAR_EXT:
        archivos_ignorar.append(file)

if not archivos_ignorar:
    print("⚠️ No hay archivos ignorados para procesar.")
    exit()

# Ordenar por fecha de creación
archivos_ignorar = sorted(archivos_ignorar, key=lambda f: f.stat().st_ctime)

print("📌 Archivos encontrados:")
for f in archivos_ignorar:
    print(" →", f.name)

# ---------------------------------------
# 2. Convertir cada archivo a PDF
# ---------------------------------------
pdf_generados = []

print("\n🧾 Convirtiendo a PDF...\n")

for archivo in archivos_ignorar:
    ext = archivo.suffix.lower()
    nombre_sin_ext = archivo.stem

    pdf_destino = CARPETA_TEMP / f"{nombre_sin_ext}.pdf"

    try:
        if ext == ".txt":
            convertir_txt_a_pdf(archivo, pdf_destino)

        elif ext in {".doc", ".docx"}:
            convertir_doc_docx_a_pdf(archivo, pdf_destino)

        elif ext == ".pdf":
            # Copiar directamente
            pdf_destino.write_bytes(archivo.read_bytes())

        pdf_generados.append(pdf_destino)
        print(f"✔ Convertido: {archivo.name} → {pdf_destino.name}")

    except Exception as e:
        print(f"❌ ERROR con {archivo.name}: {e}")

# ---------------------------------------
# 3. Unir todos los PDF en orden
# ---------------------------------------
print("\n📚 Uniendo todos los PDF en uno solo...\n")

merger = PdfMerger()

for pdf in pdf_generados:
    merger.append(str(pdf))

merger.write(str(PDF_SALIDA))
merger.close()

print(f"\n✅ PDF FINAL GENERADO:\n{PDF_SALIDA}")

# ---------------------------------------
# (Opcional) Limpiar PDF temporales
# ---------------------------------------
# import shutil
# shutil.rmtree(CARPETA_TEMP)



































from reportlab.platypus import SimpleDocTemplate, Paragraph, PageBreak, Spacer
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from pathlib import Path
from datetime import datetime
import re


# -------------------------------------------------------------
# CONFIGURACIÓN DE DIRECTORIOS
# -------------------------------------------------------------
RESULTADOS_DIR = Path(r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados")
RESULTADOS_FIN = Path(r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados_Final")

RESULTADOS_FIN.mkdir(exist_ok=True)

PDF_SALIDA = RESULTADOS_FIN / "Informe_Biodiversidad_Fauna.pdf"


# -------------------------------------------------------------
# ESTILOS PROFESIONALES DEL INFORME
# -------------------------------------------------------------
styles = getSampleStyleSheet()

styles.add(ParagraphStyle(
    name="Titulo",
    parent=styles["Title"],
    fontSize=28,
    leading=30,
    alignment=1,
    spaceAfter=40
))

styles.add(ParagraphStyle(
    name="Subtitulo",
    parent=styles["Heading2"],
    fontSize=17,
    leading=20,
    spaceBefore=20,
    spaceAfter=12
))

styles.add(ParagraphStyle(
    name="Texto",
    parent=styles["BodyText"],
    fontSize=11,
    leading=15,
    alignment=4,
    spaceAfter=12
))


# -------------------------------------------------------------
# LIMPIEZA DE TEXTO
# -------------------------------------------------------------
def limpiar_texto(texto):
    texto = re.sub(r"\n{3,}", "\n\n", texto)
    texto = texto.replace("•", "• ")          # arregla viñetas
    texto = texto.replace(" -", " — ")         # convierte guiones en en-dash
    return texto.strip()


# -------------------------------------------------------------
# GENERAR TITULO A PARTIR DEL NOMBRE DEL TXT
# -------------------------------------------------------------
def generar_titulo(nombre):
    nombre = nombre.lower()

    # Ejemplos:
    if "rich" in nombre or "riqueza" in nombre:
        return "Análisis de Riqueza y Diversidad de Especies"

    if "abun" in nombre:
        return "Análisis de Abundancia de la Comunidad"

    if "espec" in nombre:
        return "Listado de Especies Registradas"

    if "sensib" in nombre:
        return "Evaluación de Sensibilidad Ambiental"

    if "muest" in nombre:
        return "Metodología y Esfuerzo de Muestreo"

    # Si no se reconoce, se limpia y capitaliza
    nombre_limpio = re.sub(r"[_-]+", " ", nombre)
    return nombre_limpio.title()


# -------------------------------------------------------------
# OBTENER Y ORDENAR LOS TXT
# -------------------------------------------------------------
txt_files = sorted(
    [f for f in RESULTADOS_DIR.iterdir() if f.suffix.lower() == ".txt"],
    key=lambda f: f.stat().st_ctime
)

if not txt_files:
    print("⚠️ No se encontraron archivos TXT.")
    exit()


# -------------------------------------------------------------
# CREACIÓN DEL DOCUMENTO PDF
# -------------------------------------------------------------
pdf = SimpleDocTemplate(
    str(PDF_SALIDA),
    pagesize=letter,
    leftMargin=1 * inch,
    rightMargin=1 * inch,
    topMargin=1 * inch,
    bottomMargin=1 * inch
)

contenido = []


# -------------------------------------------------------------
# PORTADA
# -------------------------------------------------------------
fecha = datetime.now().strftime("%d/%m/%Y")

contenido.append(Paragraph("Informe de Biodiversidad – Fauna", styles["Titulo"]))
contenido.append(Paragraph(f"Fecha de generación del informe: {fecha}", styles["Subtitulo"]))
contenido.append(PageBreak())


# -------------------------------------------------------------
# AÑADIR CONTENIDOS DE CADA TXT
# -------------------------------------------------------------
for txt in txt_files:

    # (✓) Genera un título profesional basado en el nombre
    titulo = generar_titulo(txt.stem)

    contenido.append(Paragraph(titulo, styles["Subtitulo"]))

    with open(txt, encoding="utf-8", errors="ignore") as f:
        texto = limpiar_texto(f.read())

    parrafos = texto.split("\n")

    for p in parrafos:
        p = p.strip()
        if p:
            contenido.append(Paragraph(p, styles["Texto"]))
        else:
            contenido.append(Spacer(1, 0.2 * inch))

    contenido.append(PageBreak())


# -------------------------------------------------------------
# GENERAR EL PDF
# -------------------------------------------------------------
pdf.build(contenido)

print("\n=====================================")
print(" PDF PROFESIONAL GENERADO CORRECTAMENTE")
print(" Archivo final:", PDF_SALIDA)
print("=====================================\n")




































import os

ruta = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados"

# Listar todos los archivos con su ruta completa
archivos = [os.path.join(ruta, f) for f in os.listdir(ruta)]

print("ARCHIVOS EN Resultados:\n")
for a in archivos:
    print(a)


#------------------------------------

import os
import re

ruta = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados"
archivos = os.listdir(ruta)

estructura = {}

for archivo in archivos:
    # Obtener número inicial (1, 2, 3, 10, 11...)
    match = re.match(r"(\d+)", archivo)
    if not match:
        continue

    num = int(match.group(1))  # número de análisis

    if num not in estructura:
        estructura[num] = {"graficos": [], "tablas": [], "interpretaciones": []}

    # Clasificar por extensión
    if archivo.lower().endswith((".png", ".jpg", ".jpeg")):
        estructura[num]["graficos"].append(archivo)

    elif archivo.lower().endswith((".xlsx", ".csv")):
        estructura[num]["tablas"].append(archivo)

    elif archivo.lower().endswith(".txt"):
        estructura[num]["interpretaciones"].append(archivo)

# Mostrar estructura organizada
for num, contenido in sorted(estructura.items()):
    print(f"\n=== ANÁLISIS {num} ===")
    print("Gráficos:", contenido["graficos"])
    print("Tablas:", contenido["tablas"])
    print("Interpretaciones:", contenido["interpretaciones"])





































