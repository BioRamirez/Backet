#--------------## Cargar librerias necesarias------------------------------

# Si no las tienes instaladas, ejecuta esta celda una vez:
# Salir del interprete con: exit() exit() python   pip install tabulate pandas numpy scipy scikit-bio openpyxl
#
# !pip install pandas numpy matplotlib tabulate openpyxl

import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from tabulate import tabulate
import openpyxl


# Carpeta donde guardarás los gráficos (solo una vez)
output_folder = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados"
os.makedirs(output_folder, exist_ok=True)
#--------------## Leer archivo y revisar columnas------------------------------
# Ruta del archivo
ruta = r"D:\Forestal Consultores\2026\FAUNA\BD\BD_SANROQUE.xlsx"
# Leer el archivo Excel
Registros = pd.read_excel(ruta)

# Mostrar las primeras filas
print(" Primeras filas del archivo:")
print(Registros.head())

# Mostrar nombres de las columnas
print("\n Columnas del DataFrame:")
print(Registros.columns)

# =========================================================
# FILTRAR UNA SOLA CLASE
# =========================================================

grupo = "MAMIFEROS"   # AVES, MAMIFEROS, REPTILES,ANFIBIOS etc.

Registros = Registros[
    Registros["CLASE"].astype(str).str.upper() == grupo.upper()
].copy()

# Reiniciar índice
Registros.reset_index(drop=True, inplace=True)





# Copia de trabajo
df = Registros.copy()

# Normalizar texto
df['Uso'] = df['Uso'].astype(str).str.strip().str.lower()

# Unificación de categorías
df['Uso'] = df['Uso'].replace({
    'sin uso conocido': 'sin_uso',
    'uso cultural': 'cultural',
    'cultural': 'cultural',
    'medicinal': 'medicinal',
    'mascota': 'mascotas',
    'mascotas': 'mascotas',
    'subsistencia': 'subsistencia',
    'otro': 'otro'
})


# Quitar espacios y separar por coma
df['Uso'] = df['Uso'].str.replace(' ', '').str.split(',')

# Expandir usos múltiples a filas individuales
df_uso = df.explode('Uso')


tabla_final = (
    df_uso[['CLASE', 'Orden', 'Familia', 'Especie', 'N. comun', 'Uso']]
    .dropna()
    .drop_duplicates()
    .sort_values(['Orden', 'Familia', 'Especie'])
    .reset_index(drop=True)
)

tabla_final.rename(columns={'Uso': 'Uso'}, inplace=True)

print(tabla_final)


# 1. Copia del dataframe final
df2 = tabla_final.copy()

# 2. Renombrar la columna a "Clase"
df2.rename(columns={'CLASE': 'Clase'}, inplace=True)

# 3. Eliminar especies SIN USO
#df2 = df2[df2['Uso'] != 'sin_uso']

# 4. Ordenar primero Aves luego Mamíferos
df2['Clase'] = df2['Clase'].str.upper()

orden_clases = {'AVES': 1, 'MAMMALIA': 2}
df2['Clase_orden'] = df2['Clase'].map(orden_clases)

df2 = df2.sort_values(['Clase_orden', 'Orden', 'Familia', 'N. comun', 'Especie'])

# 5. Reiniciar índices por clase y agregar columna "N°"
df2['N°'] = df2.groupby('Clase').cumcount() + 1

# 6. Ordenar columnas como solicitaste
tabla_ordenada = df2[['N°', 'Clase', 'Orden', 'Familia', 'Especie', 'N. comun', 'Uso']]

# 7. Eliminar la columna auxiliar
tabla_ordenada = tabla_ordenada.reset_index(drop=True)

print(tabla_ordenada)

# --- 1. Definir el orden deseado para 'Uso' ---
orden_uso = ['cultural', 'medicinal', 'mascotas', 'subsistencia', 'otro', 'sin_uso']

tabla_ordenada['Uso'] = pd.Categorical(tabla_ordenada['Uso'], categories=orden_uso, ordered=True)

# --- 2. Ordenar por Clase y luego por Uso ---
tabla_ordenada = tabla_ordenada.sort_values(
    by=['Clase', 'Uso', 'Orden', 'Familia', 'Especie', 'N. comun']
).reset_index(drop=True)

# --- 3. Volver a crear N° independiente por Clase ---
tabla_ordenada['N°'] = tabla_ordenada.groupby('Clase').cumcount() + 1

# --- 4. Reordenar columnas para que N° quede primero ---
tabla_ordenada = tabla_ordenada[['N°', 'Clase', 'Orden', 'Familia', 'Especie', 'N. comun', 'Uso']]

tabla_ordenada['Clase'] = tabla_ordenada['Clase'].str.capitalize()
tabla_ordenada['Uso'] = tabla_ordenada['Uso'].str.capitalize()


print(tabla_ordenada)

# Guardar el resultado en un nuevo archivo Excel
# Carpeta donde quieres guardar
#output_folder = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados"

# Asegurar que la carpeta existe
#os.makedirs(output_folder, exist_ok=True)

# Nombre final del archivo
output_file = os.path.join(output_folder, "13_tabla_Uso_Cultural.xlsx")

# Guardar archivo
tabla_ordenada.to_excel(output_file, index=False)

print(f"\nArchivo guardado en: {output_file}")

#--------------## Fin del análisis------------------------------
#----------------------------------------------------------------------
#---------------------------------- Reparar y formatear archivo de tabla_Uso_Cultural -----------------------
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import os

# --- Rutas ---
ruta_original = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\13_tabla_Uso_Cultural.xlsx"
ruta_limpia = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\13_tabla_Uso_Cultural.xlsx"

# --- Verificar existencia ---
if not os.path.exists(ruta_original):
    raise FileNotFoundError(f"⚠️ No se encontró el archivo: {ruta_original}")

# --- Leer archivo dañado con pandas ---
try:
    df = pd.read_excel(ruta_original)
    print(" Archivo leído correctamente con pandas.")
except Exception as e:
    raise RuntimeError(f" No se pudo leer el archivo: {e}")

# --- Reescribir el archivo limpio ---
df.to_excel(ruta_limpia, index=False)
print(f" Archivo reparado y guardado como:\n{ruta_limpia}")

# --- Aplicar formato con openpyxl ---
from openpyxl import load_workbook

wb = load_workbook(ruta_limpia)
ws = wb.active

# --- Estilos base ---
header_fill = PatternFill(start_color='BFD8B8', end_color='BFD8B8', fill_type='solid')
header_font = Font(bold=True, color='000000', name='Calibri')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# --- Aplicar formato y reemplazar vacíos ---
for row in ws.iter_rows():
    for cell in row:
        if cell.value is None or str(cell.value).strip() == '':
            cell.value = '-'
        cell.alignment = center_align
        cell.border = thin_border

# --- Encabezado ---
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align

# --- Ajustar ancho de columnas ---
for col in ws.columns:
    max_length = 0
    column = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            length = len(str(cell.value))
            if length > max_length:
                max_length = length
    ws.column_dimensions[column].width = max_length + 3

# --- Ajustar altura de filas ---
for row in ws.iter_rows():
    ws.row_dimensions[row[0].row].height = 18

# --- Guardar cambios ---
wb.save(ruta_limpia)
print(f' Archivo formateado y reparado correctamente:\n{ruta_limpia}')

#-------------------------------------------------------------------------------






# ================================================================
# ========== CREAR TABLA DE USO POR MUNICIPIO ====================
# ================================================================
#--
import os
import pandas as pd

# Carpeta de salida
ruta_salida_uso_muni = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\USO_MUNICIPIOS"
os.makedirs(ruta_salida_uso_muni, exist_ok=True)

# Copias de seguridad
reg_muni = Registros.copy()
tabla_uso = tabla_ordenada.copy()

# Asegurar que no traiga columna N° previa
if "N°" in tabla_uso.columns:
    tabla_uso = tabla_uso.drop(columns=["N°"])

# 🔥 FILTRAR PARA ELIMINAR ESPECIES SIN USO
tabla_uso = tabla_uso[tabla_uso["Uso"].str.upper() != "SIN_USO"]
tabla_uso = tabla_uso[tabla_uso["Uso"].str.upper() != "Sin Uso"]
tabla_uso = tabla_uso[tabla_uso["Uso"].notna()]   # Eliminar vacíos
tabla_uso = tabla_uso[tabla_uso["Uso"] != ""]     # Eliminar cadenas vacías

# Lista de municipios
municipios = sorted(reg_muni["MUNICIPIO"].dropna().unique())

print("\n=== INICIANDO GENERACIÓN DE TABLAS DE USO POR MUNICIPIO ===")

for muni in municipios:

    print(f"\n--- Procesando municipio: {muni} ---")

    # 1. Filtrar registros del municipio
    df_muni = reg_muni[reg_muni["MUNICIPIO"] == muni]

    # 2. Especies presentes en ese municipio
    especies_muni = sorted(df_muni["Especie"].dropna().unique())

    # 3. Filtrar solo especies con uso (ya filtrado arriba)
    tabla_muni = tabla_uso[tabla_uso["Especie"].isin(especies_muni)].copy()

    # Si no hay especies con uso en ese municipio → omitir archivo
    if tabla_muni.empty:
        print(f"⚠ No hay especies con uso para el municipio {muni}. Se omite.")
        continue

    # 4. Ordenar igual que la tabla original
    tabla_muni = tabla_muni.sort_values(
        by=['Clase', 'Uso', 'Orden', 'Familia', 'Especie', 'N. comun']
    ).reset_index(drop=True)

    # 5. Añadir numeración
    tabla_muni.insert(0, "N°", range(1, len(tabla_muni) + 1))

    # 6. Guardar archivo
    archivo = os.path.join(
        ruta_salida_uso_muni,
        f"Tabla_Uso_{muni}.xlsx"
    )

    tabla_muni.to_excel(archivo, index=False)
    print(f"✔ Archivo generado: {archivo}")

print("\n✔ PROCESO COMPLETADO PARA TODOS LOS MUNICIPIOS.")











#---------------------Dar formato a cada municipio-----------------

#---------------------------------- FORMATEAR TODOS LOS ARCHIVOS DE USO POR MUNICIPIO -----------------------
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

# Carpeta donde ya están los archivos creados
ruta_salida_uso_muni = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\USO_MUNICIPIOS"

# --- Estilos base ---
header_fill = PatternFill(start_color='BFD8B8', end_color='BFD8B8', fill_type='solid')
header_font = Font(bold=True, color='000000', name='Calibri')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

print("\n=== APLICANDO FORMATO A TODOS LOS MUNICIPIOS ===")

# --- Recorrer todos los archivos .xlsx en el directorio ---
for archivo in os.listdir(ruta_salida_uso_muni):

    if archivo.lower().endswith(".xlsx"):
        ruta_archivo = os.path.join(ruta_salida_uso_muni, archivo)

        print(f"\n✔ Formateando: {archivo}")

        # --- Leer archivo con pandas (limpieza) ---
        try:
            df = pd.read_excel(ruta_archivo)
        except:
            print(f"❌ No se pudo leer {archivo}. Se omite.")
            continue

        # Reescribir para reparar si es necesario
        df.to_excel(ruta_archivo, index=False)

        # --- Formatear con openpyxl ---
        wb = load_workbook(ruta_archivo)
        ws = wb.active

        # Aplicar formato general y reemplazar vacíos
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None or str(cell.value).strip() == "":
                    cell.value = "-"
                cell.alignment = center_align
                cell.border = thin_border

        # Encabezado
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)

            for cell in col:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length

            ws.column_dimensions[column].width = max_length + 3

        # Ajustar altura de filas
        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 18

        # Guardar cambios
        wb.save(ruta_archivo)

print("\n✔ FORMATO APLICADO A TODOS LOS ARCHIVOS.")
print("===========================================")





#----------------------------------------------------------------------

import matplotlib.pyplot as plt
import pandas as pd

# --- 1. Filtrar usos válidos ---
usos_validos = ['Cultural', 'Medicinal', 'Mascotas', 'Subsistencia', 'Otro', 'Sin_uso']
df_plot = tabla_ordenada[tabla_ordenada['Uso'].isin(usos_validos)]

# --- 2. Contar especies por Clase y Uso ---
conteos = df_plot.groupby(['Clase', 'Uso'])['Especie'].nunique().reset_index()

# --- 3. Convertir en tabla pivote para graficar ---
pivot = conteos.pivot(index='Clase', columns='Uso', values='Especie').fillna(0)

# --- 4. Crear gráfico ---
ax = pivot.plot(kind='bar', figsize=(12, 7))

plt.title("Tipos de Uso por Especies", fontsize=14)
plt.xlabel("Clase", fontsize=12)
plt.ylabel("Número de especies", fontsize=12)
plt.xticks(rotation=0)
plt.legend(title="Uso", fontsize=10)

# --- 5. Agregar etiquetas de datos ---
for container in ax.containers:
    ax.bar_label(container, fmt='%d', padding=3)

plt.tight_layout()

# --- 6. Guardar el gráfico en PNG ---
output_path = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\13.1_Grafico_Uso_Cultural.png"
plt.savefig(output_path, dpi=300, bbox_inches='tight')


plt.show()

print(f"Gráfico guardado en: {output_path}")



















# ------------------------Segundo grafico---------------


import matplotlib.pyplot as plt
import pandas as pd

# --- 1. Normalizar valores de "Uso" ---
tabla_ordenada['Uso'] = tabla_ordenada['Uso'].str.strip()

usos_validos = ['Cultural', 'Medicinal', 'Mascotas', 'Subsistencia', 'Otro', 'Sin_uso']

df_plot = tabla_ordenada[tabla_ordenada['Uso'].isin(usos_validos)]

# --- 2. Contar especies por Clase y Uso ---
conteos = df_plot.groupby(['Clase', 'Uso'])['Especie'].nunique().reset_index()

# --- 3. Tabla pivote ---
pivot = conteos.pivot(index='Clase', columns='Uso', values='Especie').fillna(0)

# --- 4. Crear gráfico ---
fig, ax = plt.subplots(figsize=(12, 7))
pivot.plot(kind='bar', ax=ax)

plt.title("Tipos de Uso por Especies", fontsize=14)
plt.xlabel("Clase", fontsize=12)
plt.ylabel("Número de especies (escala log)", fontsize=12)
plt.xticks(rotation=0)
plt.legend(title="Uso", fontsize=10)

# --- 🔥 Aplicar escala logarítmica ---
ax.set_yscale('log')

# --- 5. Agregar etiquetas de datos ---
for container in ax.containers:
    ax.bar_label(container, fmt='%d', padding=3)

plt.tight_layout()

# --- 6. Guardar gráfico ---
output_path = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\13.2_Grafico_Uso_Cultural2.png"
plt.savefig(output_path, dpi=300, bbox_inches='tight')

plt.show()

print(f"Gráfico guardado en: {output_path}")







# ------------------------ Segundo gráfico (ESTILO PROFESIONAL) ------------------------

import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd
import numpy as np

# --- 1. Normalizar valores de "Uso" ---
tabla_ordenada['Uso'] = tabla_ordenada['Uso'].str.strip()

usos_validos = ['Cultural', 'Medicinal', 'Mascotas', 'Subsistencia', 'Otro', 'Sin_uso']

df_plot = tabla_ordenada[tabla_ordenada['Uso'].isin(usos_validos)]

# --- 2. Contar especies por Clase y Uso ---
conteos = df_plot.groupby(['Clase', 'Uso'])['Especie'].nunique().reset_index()

# --- 3. Tabla pivote ---
pivot = conteos.pivot(index='Clase', columns='Uso', values='Especie').fillna(0)

# -------------------------------
# 🎨 PALETA PAIRED CONSISTENTE
# -------------------------------
num_usos = len(pivot.columns)
palette = sns.color_palette("Paired", num_usos)

# --- 4. Crear gráfico ---
fig, ax = plt.subplots(figsize=(12, 7))

pivot.plot(
    kind='bar',
    ax=ax,
    color=palette,
    edgecolor="black",
    linewidth=0.6
)

# -------------------------------
# 📝 TÍTULOS Y ETIQUETAS
# -------------------------------
plt.title("Tipos de Uso por Especies", fontsize=16, fontweight='bold')
plt.xlabel("Clase", fontsize=13)
plt.ylabel("Número de especies (escala log)", fontsize=13)
plt.xticks(rotation=0, fontsize=12)
plt.yticks(fontsize=11)

# --- 🔥 Escala logarítmica ---
ax.set_yscale('log')

# -------------------------------
# 🔢 Etiquetas de datos
# -------------------------------
for container in ax.containers:
    ax.bar_label(container, fmt='%d', padding=3, fontsize=10, fontweight="bold")

# -------------------------------
# 🎛 Leyenda elegante
# -------------------------------
ax.legend(
    title="Uso",
    fontsize=11,
    title_fontsize=12,
    frameon=True,
    fancybox=True,
    edgecolor="gray"
)

# -------------------------------
# 🎨 Estética del gráfico
# -------------------------------
ax.grid(axis='y', linestyle='--', alpha=0.35)

# Borde inferior más grueso
ax.spines['bottom'].set_linewidth(1.4)

# Quitar bordes no necesarios
ax.spines['right'].set_visible(True)
ax.spines['top'].set_visible(True)

plt.tight_layout()

# --- 6. Guardar gráfico ---
output_path = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\13.2_Grafico_Uso_Cultural2.png"
plt.savefig(output_path, dpi=350, bbox_inches='tight')

plt.show()

print(f"✔ Gráfico guardado en: {output_path}")













#------------------------INTERPRETACION------------------

#------------------------INTERPRETACION------------------

def interpretar_tabla_funcional(tabla):
    """
    Genera una interpretación funcional basada únicamente en la columna:
    Uso
    """

    texto = []

    # ===============================
    # 1. ANÁLISIS GENERAL
    # ===============================
    total_registros = len(tabla)
    total_usos = tabla["Uso"].nunique()
    total_especies = tabla["Especie"].nunique()

    texto.append(
        f"1. Análisis general del ensamblaje avifaunístico\n\n"
        f"La tabla analizada reúne un total de {total_registros} registros correspondientes "
        f"a {total_especies} especies que presentan {total_usos} categorías de uso. "
        f"Este patrón refleja la diversidad de interacciones entre la avifauna y los servicios "
        f"ecosistémicos o valores socioculturales a los cuales se encuentran asociadas."
    )

    # ===============================
    # 2. DISTRIBUCIÓN POR CATEGORÍA DE USO
    # ===============================
    texto.append(
        "\n2. Distribución de especies según categorías de uso\n\n"
        "Las especies fueron clasificadas según su relación con servicios ecosistémicos y "
        "funciones culturales. La distribución general es la siguiente:\n"
    )

    usos = tabla.groupby("Uso").size()

    for uso, n in usos.items():
        especies_uso = tabla[tabla["Uso"] == uso]["Especie"].nunique()
        texto.append(
            f"- {uso}: {n} registros que corresponden a {especies_uso} especies."
        )

    texto.append(
        "\nLas categorías de uso permiten comprender la relación entre la biodiversidad y los "
        "servicios ecosistémicos, destacando el papel cultural, ecológico o funcional que "
        "desempeñan las aves dentro de los paisajes estudiados."
    )

    # ===============================
    # 3. CONSIDERACIONES FINALES
    # ===============================
    texto.append(
        "\n3. Consideraciones finales\n\n"
        "El análisis funcional evidencia que la avifauna registrada cumple múltiples roles "
        "dentro de los sistemas socioecológicos evaluados. La identificación de categorías de uso "
        "constituye un insumo clave para la gestión de la biodiversidad, al permitir reconocer "
        "aquellos grupos con valor cultural, económico o ecológico, y orientar estrategias de "
        "conservación y manejo basadas en su contribución funcional."
    )

    return "\n".join(texto)

texto_interpretacion = interpretar_tabla_funcional(tabla_ordenada)
print(texto_interpretacion)


# ==========================================
# GUARDAR EL TEXTO EN UN ARCHIVO .TXT
# ==========================================

ruta_salida = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\13_Interpretacion_Uso.txt"

with open(ruta_salida, "w", encoding="utf-8") as file:
    file.write(texto_interpretacion)

print(f"Archivo guardado correctamente en:\n{ruta_salida}")
