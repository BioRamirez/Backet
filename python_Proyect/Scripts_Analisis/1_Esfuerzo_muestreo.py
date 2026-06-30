#--------------## Cargar librerias necesarias------------------------------

# Si no las tienes instaladas, ejecuta esta celda una vez:
# Salir del interprete con: exit() exit() python   pip install tabulate pandas numpy scipy scikit-bio openpyxl
#
# !pip install pandas numpy matplotlib tabulate openpyxl

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from tabulate import tabulate
import openpyxl

#--------------## Leer archivo y revisar columnas----------------D:\CORPONOR 2025\Backet\python_Proyect\data\SRF_LAM_5235_AVES_SAMORE_HERPETOS.xlsx--------------
# Ruta del archivo
#ruta = r"D:\CORPONOR 2025\Backet\python_Proyect\data\POF_ZULIA_2025_BD_AVES_MAMIFEROS.xlsx"
ruta = r"D:\Forestal Consultores\2026\FAUNA\BD\BD_SANROQUE.xlsx"

#
# Leer el archivo Excel 
Registros = pd.read_excel(ruta)

# =========================================================
# FILTRAR UNA SOLA CLASE
# =========================================================

grupo = "MAMIFEROS"   # AVES, MAMIFEROS, REPTILES, ANFIBIOS etc.

Registros = Registros[
    Registros["CLASE"].astype(str).str.upper() == grupo.upper()
].copy()

# Reiniciar índice
Registros.reset_index(drop=True, inplace=True)

# =========================================================
# VERIFICACIÓN
# =========================================================

print(f"\nGrupo seleccionado: {grupo}")
print(f"Número de registros: {len(Registros)}")

print("\nPrimeras filas:")
print(Registros.head())

# Mostrar las primeras filas
print(" Primeras filas del archivo:")
print(Registros.head())

# Mostrar nombres de las columnas
print("\n Columnas del DataFrame:")
print(Registros.columns)

# Mostrar valores únicos de algunas columnas
for col in ["METODOLOGIA", "METODO", "ID", "Gremio"]:
    if col in Registros.columns:
        print(f"\n Valores únicos en '{col}':")
        print(Registros[col].unique())
    else:
        print(f"\n⚠️ La columna '{col}' no existe en el DataFrame.")

#--------------## Esfuerzo de Muestreo------------------------------

import pandas as pd
from tabulate import tabulate

# --- Copiar dataframe base ---
df = Registros.copy()

# --- Aplicar reemplazos (corrección lectura de coberturas) ---Bosque denso alto con vegetación secundaria

df['COBERTURA'] = (
    df['COBERTURA']
    .astype(str)
    .str.strip()
    .str.lower()
)

abreviaciones_cobertura = {
    'bosque de galería y ripario': 'Bgr',
    'bosque de galeria y ripario': 'Bgr',
    'bosque denso alto de tierra firme': 'Bda',
    'bosque denso alto con vegetacion scundaria': 'Bdavs',
    'bosque denso bajo con vegetacion secundaria': 'Bda',
    'bosque denso bajo de tierra firme': 'Bdb',
    'bosque fragmentado con vegetación secundaria': 'Bfvs',
    'bosque fragmentado': 'BF',
    'pastos limpios': 'PL',
    'pastos enmalezados': 'PE',
    'pastos arbolados': 'PA',
    'cultivos permanentes arboreos': 'CPA',
    'mosaico de cultivos, pastos y espacios naturales': 'MEN',
    'zonas de extracción minera': 'ZEM',
    'sin dato': 'NA'
}

df['COBERTURA'] = df['COBERTURA'].map(abreviaciones_cobertura).fillna(df['COBERTURA'])

print(df['COBERTURA'].value_counts())


# --- Aplicar reemplazos (con control de mayúsculas y tildes) ---
df['COBERTURA'] = df['COBERTURA'].apply(
    lambda x: abreviaciones_cobertura.get(x.strip().title(), x)
)

# --- Reemplazar vacíos y nulos por 'Sin dato' ---
df['METODO'] = df['METODO'].replace('', 'Sin dato').fillna('Sin dato')
df['COBERTURA'] = df['COBERTURA'].replace('', 'Sin dato').fillna('Sin dato')


# --- Validar que exista la nueva columna de horas ---
if 'Hora_Hombre' not in df.columns:
    raise ValueError(' La columna Hora_Hombre no existe en el dataframe Registros.')

# --- Calcular totales de individuos (sin perder registros) ---
individuos = (
    df.groupby(['METODO', 'COBERTURA'], dropna=False, as_index=False)['INDIVIDUOS']
      .sum(min_count=1)
)

# --- Calcular esfuerzo total único por ID ---
# (Evitamos duplicar horas si un ID aparece varias veces)
esfuerzo_unico = df[['ID', 'METODO', 'COBERTURA', 'Hora_Hombre']].drop_duplicates()

# --- Calcular esfuerzo total (solo una vez por ID) ---
esfuerzo = (
    esfuerzo_unico.groupby(['METODO', 'COBERTURA'], dropna=False, as_index=False)['Hora_Hombre']
    .sum(min_count=1)
    .rename(columns={'Hora_Hombre': 'Esfuerzo_horas'})
)


# --- Unir tablas ---
tabla = individuos.merge(esfuerzo, on=['METODO', 'COBERTURA'], how='outer')
tabla['Exito_captura'] = tabla['INDIVIDUOS'] / tabla['Esfuerzo_horas']

# --- Calcular totales por método ---
totales = tabla.groupby('METODO', as_index=False).agg({
    'INDIVIDUOS': 'sum',
    'Esfuerzo_horas': 'sum'
})
totales['Exito_captura'] = totales['INDIVIDUOS'] / totales['Esfuerzo_horas']
totales['COBERTURA'] = 'Total'

# --- Unir con la tabla principal ---
tabla_final = pd.concat([tabla, totales], ignore_index=True)

# --- Reestructurar para salida ---
tabla_melt = pd.melt(
    tabla_final,
    id_vars=['METODO', 'COBERTURA'],
    value_vars=['INDIVIDUOS', 'Esfuerzo_horas', 'Exito_captura'],
    var_name='Indice',
    value_name='Valor'
)

# --- Cambiar nombres de los índices ---
tabla_melt['Indice'] = tabla_melt['Indice'].replace({
    'INDIVIDUOS': 'Número de individuos',
    'Esfuerzo_horas': 'Esfuerzo captura (horas-hombre)',
    'Exito_captura': 'Éxito de captura (individuos/horas-hombre)'
})


# --- Orden lógico de los índices ---
orden_indices = [
    'Número de individuos',
    'Esfuerzo captura (horas-hombre)',
    'Éxito de captura (individuos/horas-hombre)'
]
tabla_melt['Indice'] = pd.Categorical(tabla_melt['Indice'], categories=orden_indices, ordered=True)

# --- Renombrar columna ---
tabla_melt = tabla_melt.rename(columns={'METODO': 'Metodologia'})

# --- Orden personalizado de metodologias ---
orden_metodologia = [
    'Transecto',
    'Punto de observacion',
    'Red de niebla',
    'Camara trampa',
    'Informacion Secundaria'
]

tabla_melt['Metodologia'] = pd.Categorical(tabla_melt['Metodologia'], categories=orden_metodologia, ordered=True)

# --- Orden personalizado de Metodologias ---
orden_COBERTURA = [
    'Bgr',
    'Bfvs',
    'Bdavs',
    'Bda',
    'Bdb',
    'BF',
    'ZEM',
    'MEN',
    'CPA',
    'PA',
    'PL',
    'PE',
    'Total'
]

tabla_melt['COBERTURA'] = pd.Categorical(tabla_melt['COBERTURA'], categories=orden_COBERTURA, ordered=True)

# --- Pivotar ---
tabla_pivot = tabla_melt.pivot_table(
    index=['Metodologia', 'Indice'],
    columns='COBERTURA',
    values='Valor',
    aggfunc='first'
).reset_index()

# --- Redondear ---
tabla_pivot = tabla_pivot.round(3)

# --- Mostrar resumen en consola ---
print(tabulate(tabla_pivot, headers='keys', tablefmt='fancy_grid', floatfmt='.3f'))


# --- Exportar a Excel a una ruta específica ---
import os

# Definir la ruta exacta donde guardar el archivo
output_path = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados"
output_file = os.path.join(output_path, "1_Esfuerzo_Muestreo.xlsx")

# Exportar el DataFrame a Excel
tabla_pivot.to_excel(output_file, index=False)

# Confirmar la ubicación del archivo guardado
print(f" Archivo exportado correctamente en:\n{output_file}")



#---------------Dar formato a archivo generato o tabla---------------


from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import os

# --- Nombre del archivo a formatear ---
output_file = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\1_Esfuerzo_Muestreo.xlsx"

# --- Verificar que el archivo existe ---
if not os.path.exists(output_file):
    raise FileNotFoundError(f"⚠️ No se encontró el archivo: {output_file}")

# --- Cargar el archivo ---
wb = load_workbook(output_file)
ws = wb.active

# --- Estilos base ---
header_fill = PatternFill(start_color='BFD8B8', end_color='BFD8B8', fill_type='solid')
header_font = Font(bold=True, color='000000', name='Calibri')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

# --- Bordes finos para toda la tabla ---
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# --- Aplicar formato y reemplazar vacíos ---
for row in ws.iter_rows():
    for cell in row:
        # Reemplazar vacíos o None por guion
        if cell.value is None or str(cell.value).strip() == '':
            cell.value = '-'
        # Aplicar formato general
        cell.border = thin_border
        cell.alignment = center_align

# --- Aplicar formato al encabezado ---
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align

# --- Ajustar ancho de columnas automáticamente ---
for col in ws.columns:
    max_length = 0
    column = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            length = len(str(cell.value))
            if length > max_length:
                max_length = length
    ws.column_dimensions[column].width = max_length + 3

# --- Ajustar altura de filas automáticamente ---
for row in ws.iter_rows():
    max_height = 15
    for cell in row:
        if cell.value and "\n" in str(cell.value):
            lines = str(cell.value).count('\n') + 1
            if lines > 1:
                max_height = 15 * lines
    ws.row_dimensions[cell.row].height = max_height

# --- Guardar cambios ---
wb.save(output_file)
print(f' Archivo formateado con éxito:\n{output_file}')


#-------------Mostrar archivo creado-----------------

import pandas as pd
from IPython.display import display, HTML

# --- Leer el archivo Excel ---
#  Usa una cadena RAW (r"...") para evitar errores con las barras invertidas
tabla = pd.read_excel(r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\1_Esfuerzo_Muestreo.xlsx")

# --- Mostrar tabla con desplazamiento vertical ---
display(HTML(f"""
<h3>Vista del archivo <code>Esfuerzo_Muestreo.xlsx</code></h3>
<div style="
    height: 400px;
    overflow-y: scroll;
    border: 1px solid #ccc;
    padding: 8px;
    font-size: 14px;
">
{tabla.to_html(index=False)}
</div>
"""))




#---------------------------INTERPRETACION-------------------------
#-----------------------------------------------------------------
import pandas as pd
import os

def interpretar_metodos(tabla, output_folder):

    # Crear carpeta de resultados si no existe
    os.makedirs(output_folder, exist_ok=True)

    ruta_salida = os.path.join(output_folder, "1.1_Interpretacion_Esfuerzo.txt")

    with open(ruta_salida, "w", encoding="utf-8") as f:

        f.write("INTERPRETACIÓN AUTOMÁTICA DE RESULTADOS POR METODOLOGÍA\n")
        f.write("--------------------------------------------------------\n\n")

        metodologias = tabla["Metodologia"].unique()

        # Lista para generar el texto global sintetizado
        resumen_global = []

        for metodo in metodologias:
            sub = tabla[tabla["Metodologia"] == metodo]

            f.write(f"\n### {metodo}\n")

            # --------------------------------------
            # Número de individuos
            # --------------------------------------
            fila_ind = sub[sub["Indice"] == "Número de individuos"].iloc[0]
            total_ind = fila_ind["Total"]

            coberturas = [c for c in fila_ind.index if c not in ["Metodologia", "Indice", "Total"]]
            datos_ind = {c: fila_ind[c] for c in coberturas if isinstance(fila_ind[c], (int, float))}

            f.write("\n1. Número de individuos registrados:\n")
            f.write(f"   - Total registrado: {total_ind:.3f}\n")

            if datos_ind:
                mayor_ind = max(datos_ind, key=datos_ind.get)
                f.write(f"   - Mayor aporte: {mayor_ind} ({datos_ind[mayor_ind]:.3f} individuos).\n")

                otras = {c: v for c, v in datos_ind.items() if c != mayor_ind}
                if otras:
                    f.write("   - Otros aportes relevantes:\n")
                    for c, v in otras.items():
                        f.write(f"        * {c}: {v:.3f} individuos.\n")
            else:
                f.write("   - Este método no registró individuos.\n")

            # --------------------------------------
            # Esfuerzo
            # --------------------------------------
            fila_esf = sub[sub["Indice"] == "Esfuerzo captura (horas-hombre)"].iloc[0]
            total_esf = fila_esf["Total"]
            datos_esf = {c: fila_esf[c] for c in coberturas if isinstance(fila_esf[c], (int, float))}

            f.write("\n2. Esfuerzo aplicado:\n")
            f.write(f"   - Total: {total_esf:.3f} horas-hombre.\n")

            if datos_esf:
                mayor_esf = max(datos_esf, key=datos_esf.get)
                f.write(f"   - Mayor esfuerzo en: {mayor_esf} ({datos_esf[mayor_esf]:.3f}).\n")

            # --------------------------------------
            # Éxito
            # --------------------------------------
            fila_exc = sub[sub["Indice"] == "Éxito de captura (individuos/horas-hombre)"].iloc[0]
            total_exc = fila_exc["Total"]
            datos_exc = {c: fila_exc[c] for c in coberturas if isinstance(fila_exc[c], (int, float))}

            f.write("\n3. Éxito de captura:\n")
            f.write(f"   - Éxito general: {total_exc:.3f} individuos por hora-hombre.\n")

            if datos_exc:
                mayor_exc = max(datos_exc, key=datos_exc.get)
                f.write(f"   - Mayor éxito en: {mayor_exc} ({datos_exc[mayor_exc]:.3f}).\n")

                otras_exc = {c: v for c, v in datos_exc.items() if c != mayor_exc}
                if otras_exc:
                    f.write("   - Éxitos observados en otras coberturas:\n")
                    for c, v in otras_exc.items():
                        f.write(f"        * {c}: {v:.3f}.\n")

            f.write("\n--------------------------------------------------------\n")

            # Agregar al resumen global sintetizado
            resumen_global.append(
                f"El método {metodo} registró un total de {total_ind:.3f} individuos, "
                f"con un esfuerzo acumulado de {total_esf:.3f} horas-hombre y un "
                f"éxito promedio de {total_exc:.3f} individuos/hora-hombre."
            )

        # -------------------------------------------------------------
        # PÁRRAFO GLOBAL SINTETIZADO
        # -------------------------------------------------------------
        f.write("\n\nRESUMEN INTERPRETATIVO GLOBAL\n")
        f.write("--------------------------------------------------------\n\n")

        texto_global = (
            "En términos generales, los distintos métodos de muestreo aplicados "
            "mostraron variaciones importantes en abundancia registrada, esfuerzo "
            "empleado y eficiencia (éxito de captura). De manera global, se observa que "
            "los métodos difieren en su capacidad de detección según la cobertura y el "
            "tipo de técnica aplicada. A continuación, se sintetizan los patrones "
            "generales observados:\n\n"
        )

        f.write(texto_global)

        for linea in resumen_global:
            f.write(f"- {linea}\n")

        f.write("\nEste resumen global permite comprender la eficiencia relativa de cada metodología "
                "y su aporte al inventario general, facilitando la interpretación ecológica del esfuerzo "
                "de muestreo aplicado en el área de estudio.\n")

    print(f"\nArchivo generado en: {ruta_salida}\n")


# ------------------------------------------------
# EJEMPLO DE USO
# ------------------------------------------------
# tabla = pd.read_excel("tu_archivo.xlsx")
# interpretar_metodos(tabla, output_folder="Resultados")

#-----------------------------------------

interpretar_metodos(tabla, output_folder= r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados")