

#--------------## Cargar librerias necesarias------------------------------

# Si no las tienes instaladas, ejecuta esta celda una vez:
# Salir del interprete con: exit() exit() python   pip install tabulate pandas numpy scipy scikit-bio openpyxl
#
# !pip install pandas numpy matplotlib tabulate openpyxl

import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from tabulate import tabulate
import openpyxl

output_folder = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados"
os.makedirs(output_folder, exist_ok=True)

#--------------## Leer archivo y revisar columnas------------------------------
# Ruta del archivo "D:\Forestal Consultores\2026\FAUNA\BD\BASE_DATOS.xlsx"

ruta = r"D:\Forestal Consultores\2026\FAUNA\BD\BD_SANROQUE.xlsx"

# Leer el archivo Excel
Registros = pd.read_excel(ruta)

# Mostrar las primeras filas
print(" Primeras filas del archivo:")
print(Registros.head())

# Mostrar nombres de las columnas
print("\n Columnas del DataFrame:")
print(Registros.columns)

#-----------------------------RUTA SIG______________

# Ruta del archivo
ruta = r"D:\Forestal Consultores\2026\FAUNA\BD\Puntos_Transectos_SIG_SanRoque.xlsx"

# Leer el archivo Excel
datosSIG = pd.read_excel(ruta)
# Mostrar las primeras filas
print(" Primeras filas del archivo:")
print(datosSIG.head())

# Mostrar nombres de las columnas
print("\n Columnas del DataFrame:")
print(datosSIG.columns)


#-----------------------------RUTA OPERADOR______________

# Ruta del archivo
ruta = r"D:\Forestal Consultores\2026\FAUNA\BD\Operador_Info_SanRoque.xlsx"

# Leer el archivo Excel
OPER_INFO = pd.read_excel(ruta)
# Mostrar las primeras filas
print(" Primeras filas del archivo:")
print(OPER_INFO.head())

# Mostrar nombres de las columnas
print("\n Columnas del DataFrame:")
print(OPER_INFO.columns)

#------------------Crear taabla punto muestreo fauna----------------------------

import pandas as pd

estructura_punto_muestreo = {
    "EXPEDIENTE": str,
    "OPERADOR": str,
    "PROYECTO": str,
    "NUM_ACT_AD": str,
    "FEC_ACT_AD": "datetime64[ns]",
    "ART_ACT_AD": str,
    "VEREDA": str,
    "MUNICIPIO": str,
    "DEPTO": str,
    "NOMBRE": str,
    "ID_MUES_PT": str,
    "N_COBERT": str,
    "NOMENCLAT": "Int64",
    "T_MUEST": float,
    "HABITAT": str,
    "DESCRIP": str,
    "FEC_MUEST": "datetime64[ns]",
    "ESTACIONAL": float,
    "CUERPO_AGU": str,
    "COTA": float,
    "COOR_ESTE": float,
    "COOR_NORTE": float
}

df_punto_muestreo = pd.DataFrame(columns=estructura_punto_muestreo.keys())
df_punto_muestreo = df_punto_muestreo.astype(estructura_punto_muestreo)

df_punto_muestreo

# Mostrar nombres de las columnas
print("\n Columnas del DataFrame:")
print(df_punto_muestreo.columns)

#----------------Rellenar ID_MUES_PT----------------------
import pandas as pd
import re

# --- 1. Expresión regular: busca C, P o R en los últimos 3 caracteres ---
patron = r"[CPRcp r]{1}[A-Za-z0-9]{0,2}$"

# --- 2. Filtrar IDs válidos ---
ids_validos = Registros[
    Registros["ID"].astype(str).str[-3:].str.contains(r"[CPRcpr]", regex=True)
]["ID"]

# --- 3. Eliminar duplicados ---
ids_unicos = ids_validos.drop_duplicates().reset_index(drop=True)

# --- 4. Crear df_punto_muestreo base con las columnas correctas ---
columnas_punto = [
    "EXPEDIENTE", "OPERADOR", "PROYECTO", "NUM_ACT_AD", "FEC_ACT_AD", "ART_ACT_AD",
    "VEREDA", "MUNICIPIO", "DEPTO", "NOMBRE", "ID_MUES_PT", "N_COBERT", "NOMENCLAT",
    "T_MUEST", "HABITAT", "DESCRIP", "FEC_MUEST", "ESTACIONAL", "CUERPO_AGU",
    "COTA", "COOR_ESTE", "COOR_NORTE"
]

df_punto_muestreo = pd.DataFrame(columns=columnas_punto)

# --- 5. Llenar únicamente ID_MUES_PT y NOMBRE (que será igual al ID) ---
df_punto_muestreo["ID_MUES_PT"] = ids_unicos
df_punto_muestreo["NOMBRE"] = ids_unicos

print("\n df_punto_muestreo creado con IDs filtrados:")
print(df_punto_muestreo.head())
print(f"\nTotal de ID_MUES_PT generados: {len(df_punto_muestreo)}")

#---------------------Rellenar iformacion del proyecto y operador-----------------------

# Tomamos la primera fila de OPER_INFO como diccionario
info = OPER_INFO.iloc[0].to_dict()

# Para cada columna que exista en df_punto_muestreo y en OPER_INFO
for col in ["EXPEDIENTE", "OPERADOR", "PROYECTO", "NUM_ACT_AD", "FEC_ACT_AD", "ART_ACT_AD"]:
    if col in df_punto_muestreo.columns:
        df_punto_muestreo[col] = info[col]



#----------------------Rellenar coordenadas y otros datos desde datosSIG-----------------------

# --- 1. Seleccionar columnas relevantes de datosSIG ---
sig_cols = [
    "ID",
    "VEREDA",
    "MUNICIPIO",
    "DEPARTAMENTO",
    "COTA",
    "POINT_X_MAGNA_F",
    "POINT_Y_MAGNA_F"
]

sig_info = datosSIG[sig_cols].drop_duplicates(subset=["ID"])


# --- 2. Preparamos SOLO la columna de llave del df principal ---
clave = df_punto_muestreo[["ID_MUES_PT"]].copy()


# --- 3. Merge limpio SIN traer columnas duplicadas ---
df_merge = clave.merge(
    sig_info,
    left_on="ID_MUES_PT",
    right_on="ID",
    how="left"
)

# (opcional) eliminar la columna ID que solo se usó para unir
df_merge = df_merge.drop(columns=["ID"])


# --- 4. Sobrescribir columnas oficiales ---
df_punto_muestreo["VEREDA"]     = df_merge["VEREDA"]
df_punto_muestreo["MUNICIPIO"]  = df_merge["MUNICIPIO"]
df_punto_muestreo["DEPTO"]      = df_merge["DEPARTAMENTO"]
df_punto_muestreo["COTA"]       = df_merge["COTA"]
df_punto_muestreo["COOR_ESTE"]  = df_merge["POINT_X_MAGNA_F"]
df_punto_muestreo["COOR_NORTE"] = df_merge["POINT_Y_MAGNA_F"]

#---------------------------Definir coberturas y codigos----------------------------

codigo_a_cobertura = {
    "1": "TERRITORIOS ARTIFICIALIZADOS",
    "11": "Zonas urbanizadas",
    "111": "Tejido urbano continuo",
    "112": "Tejido urbano discontinuo",
    "12": "Zonas industriales o comerciales y redes de comunicación",
    "121": "Zonas industriales o comerciales",
    "1211": "Zonas industriales",
    "1212": "Zonas comerciales",
    "122": "Red vial, ferroviaria y terrenos asociados",
    "1221": "Red vial y terrenos asociados",
    "1222": "Red ferroviaria y terrenos asociados",
    "123": "Zonas portuarias",
    "1231": "Zonas portuarias fluviales",
    "1232": "Zonas portuarias marítimas",
    "124": "Aeropuertos",
    "1241": "Aeropuerto con infraestructura asociada",
    "1242": "Aeropuerto sin infraestructura asociada",
    "125": "Obras hidráulicas",
    "13": "Zonas de extracción minera y escombreras",
    "131": "Zonas de extracción minera",
    "1311": "Otras explotaciones mineras",
    "1312": "Explotación de hidrocarburos",
    "1313": "Explotación de carbón",
    "1314": "Explotación de oro",
    "1315": "Explotación de materiales de construcción",
    "1316": "Explotación de sal",
    "1317": "Explotación de esmeraldas",
    "132": "Zonas de disposición de residuos",
    "1321": "Otros sitios de disposición de residuos a cielo abierto",
    "1322": "Escombreras",
    "1323": "Vertederos",
    "1324": "Relleno sanitario",
    "14": "Zonas verdes artificializadas, no agrícolas",
    "141": "Zonas verdes urbanas",
    "1411": "Otras zonas verdes urbanas",
    "1412": "Parques cementerios",
    "1413": "Jardines botánicos",
    "1414": "Zoológicos",
    "1415": "Parques urbanos",
    "1416": "Rondas de cuerpos de agua de zonas urbanas",
    "142": "Instalaciones recreativas",
    "1421": "Áreas culturales",
    "1422": "Áreas deportivas",
    "1423": "Áreas turísticas",

    "2": "TERRITORIOS AGRÍCOLAS",
    "21": "Cultivos transitorios",
    "211": "Otros cultivos transitorios",
    "212": "Cereales",
    "2121": "Arroz",
    "2122": "Maíz",
    "2123": "Sorgo",
    "2124": "Cebada",
    "2125": "Trigo",
    "213": "Oleaginosas y leguminosas",
    "2131": "Algodón",
    "2132": "Ajonjolí",
    "2133": "Fríjol",
    "2134": "Soya",
    "2135": "Maní",
    "214": "Hortalizas",
    "2141": "Cebolla",
    "2142": "Zanahoria",
    "2143": "Remolacha",
    "215": "Tubérculos",
    "2151": "Papa",
    "2152": "Yuca",

    "22": "Cultivos permanentes",
    "221": "Cultivos permanentes herbáceos",
    "2211": "Otros cultivos permanentes herbáceos",
    "2212": "Caña",
    "22121": "Caña de azúcar",
    "22122": "Caña panelera",
    "2213": "Plátano y banano",
    "2214": "Tabaco",
    "2215": "Papaya",
    "2216": "Amapola",
    "222": "Cultivos permanentes arbustivos",
    "2221": "Otros cultivos permanentes arbustivos",
    "2222": "Café",
    "2223": "Cacao",
    "2224": "Viñedos",
    "2225": "Coca",
    "223": "Cultivos permanentes arbóreos",
    "2231": "Otros cultivos permanentes arbóreos",
    "2232": "Palma de aceite",
    "2233": "Cítricos",
    "2234": "Mango",
    "224": "Cultivos agroforestales",
    "2241": "Pastos y árboles plantados",
    "2242": "Cultivos y árboles plantados",
    "225": "Cultivos confinados",

    "23": "Pastos",
    "231": "Pastos limpios",
    "232": "Pastos arbolados",
    "233": "Pastos enmalezados",

    "24": "Áreas agrícolas heterogéneas",
    "241": "Mosaico de cultivos",
    "242": "Mosaico de pastos y cultivos",
    "243": "Mosaico de cultivos, pastos y espacios naturales",
    "244": "Mosaico de pastos con espacios naturales",
    "245": "Mosaico de cultivos y espacios naturales",

    "3": "BOSQUES Y ÁREAS SEMINATURALES",
    "31": "Bosques",
    "311": "Bosque denso",
    "3111": "Bosque denso alto",
    "31111": "Bosque denso alto de tierra firme",
    "31112": "Bosque denso alto inundable",
    "311121": "Bosque denso alto inundable heterogéneo",
    "311122": "Manglar denso alto",
    "311123": "Palmares",
    "3112": "Bosque denso bajo",
    "31121": "Bosque denso bajo de tierra firme",
    "311211": "Caatingas",
    "311212": "Bosque enano del Caribe",
    "311213": "Bosque denso altoandino",
    "31122": "Bosque denso bajo inundable",

    "312": "Bosque abierto",
    "3121": "Bosque abierto alto",
    "31211": "Bosque abierto alto de tierra firme",
    "31212": "Bosque abierto alto inundable",
    "3122": "Bosque abierto bajo",
    "31221": "Bosque abierto bajo de tierra firme",
    "31222": "Bosque abierto bajo inundable",

    "313": "Bosque fragmentado",
    "3131": "Bosque fragmentado con pastos y cultivos",
    "3132": "Bosque fragmentado con vegetación secundaria",

    "314": "Bosque de galería y ripario",
    "315": "Plantación forestal",
    "3151": "Plantación de coníferas",
    "3152": "Plantación de latifoliadas",

    "32": "Áreas con vegetación herbácea y/o arbustiva",
    "3211": "Herbazal denso",
    "32111": "Herbazal denso de tierra firme",
    "321111": "Herbazal denso de tierra firme no arbolado",
    "3211111": "Herbazal denso alto de tierra firme no arbolado",
    "3211112": "Herbazal denso bajo de tierra firme no arbolado",
    "321112": "Herbazal denso de tierra firme arbolado",
    "321113": "Herbazal denso de tierra firme con arbustos",
    "32112": "Herbazal denso inundable",
    "321121": "Herbazal denso inundable no arbolado",
    "321122": "Herbazal denso inundable arbolado",
    "321123": "Arracachal",
    "321124": "Helechal",

    "3212": "Herbazal abierto",
    "32121": "Herbazal abierto arenoso",
    "32122": "Herbazal abierto rocoso",

    "3221": "Arbustal denso",
    "3222": "Arbustal abierto",
    "32221": "Arbustal abierto esclerófilo",
    "32222": "Arbustal abierto mesófilo",

    "323": "Vegetación secundaria o en transición",
    "3231": "Vegetación secundaria alta",
    "3232": "Vegetación secundaria baja",

    "33": "Áreas abiertas sin o con poca vegetación",
    "331": "Zonas arenosas naturales",
    "3311": "Playas",
    "3312": "Arenales",
    "3313": "Campos de dunas",
    "332": "Afloramientos rocosos",
    "333": "Tierras desnudas y degradadas",
    "334": "Zonas quemadas",
    "335": "Zonas glaciares y nivales",
    "3351": "Zonas glaciares",
    "3352": "Zonas nivales",

    "4": "ÁREAS HÚMEDAS",
    "41": "Áreas húmedas continentales",
    "411": "Zonas pantanosas",
    "412": "Turberas",
    "413": "Vegetación acuática sobre cuerpos de agua",

    "42": "Áreas húmedas costeras",
    "421": "Pantanos costeros",
    "422": "Salitral",
    "423": "Sedimentos expuestos en bajamar",

    "5": "SUPERFICIES DE AGUA",
    "51": "Aguas continentales",
    "511": "Ríos (50 m)",
    "512": "Lagunas, lagos y ciénagas naturales",
    "513": "Canales",
    "514": "Cuerpos de agua artificiales",
    "5141": "Embalses",
    "5142": "Lagunas de oxidación",
    "5143": "Estanques para acuicultura continental",

    "52": "Aguas marítimas",
    "521": "Lagunas costeras",
    "522": "Mares y océanos",
    "5221": "Otros fondos",
    "5222": "Fondos coralinos someros",
    "5223": "Praderas de pastos marinos someras",
    "5224": "Fondos someros de arenas y cascajo",
    "523": "Estanques para acuicultura marina"
}


# --- 1. Crear diccionario invertido para obtener códigos desde nombres ---
codigo_a_cobertura = {v: k for k, v in codigo_a_cobertura.items()}

# --- 2. Merge para traer N_COBERT desde datosSIG ---
df_merge = df_punto_muestreo[["ID_MUES_PT"]].merge(
    datosSIG[["ID", "COBERTURA"]],
    left_on="ID_MUES_PT",
    right_on="ID",
    how="left"
)

# --- 3. Sobrescribir el nombre de la cobertura ---
df_punto_muestreo["N_COBERT"] = df_merge["COBERTURA"]

# --- 4. Asignar NOMENCLAT usando el diccionario oficial ---
df_punto_muestreo["NOMENCLAT"] = df_punto_muestreo["N_COBERT"].map(codigo_a_cobertura)

#-----------------Rellenar Habitat, Descripcion y Metodo de Muestreo----------------------

# Diccionario del tipo de muestreo según METODO
dicc_tipo_muestreo = {
    "Camara trampa": "Puntual",
    "Punto de observacion": "Puntual",
    "Red de niebla": "Puntual",
    "Nido": "Puntual",
    "Parcela": "Parcela"
}

# Diccionario grupos
grupos_dicc = {
    "A": "aves",
    "M": "mamíferos",
    "F": "fauna (Aves, mamiferos, anfibios y reptiles)",
    "R": "reptiles",
    "AN": "Anfibios",
    "H": "herpetos"
}


# Función para interpretar los grupos del ID
def interpretar_grupos(id_value):
    import re
    prefijo = re.match(r"[A-Z]+", id_value.upper()).group(0)
    grupos = [grupos_dicc[l] for l in prefijo if l in grupos_dicc]

    if len(grupos) == 1:
        return grupos[0]
    elif len(grupos) == 2:
        return f"{grupos[0]} y {grupos[1]}"
    else:
        return ", ".join(grupos[:-1]) + f" y {grupos[-1]}"

# Métodos que generan descripción
metodos_puntuales = ["Camara Trampa", "Punto de observacion", "Red de niebla"]

# Construir columnas
df_punto_muestreo["T_MUEST"] = df_punto_muestreo["ID_MUES_PT"].map(
    lambda x: dicc_tipo_muestreo.get(Registros.loc[Registros["ID"] == x, "METODO"].iloc[0], None)
)

# HABITAT igual a N_COBERT
df_punto_muestreo["HABITAT"] = df_punto_muestreo["N_COBERT"]

# DESCRIP
def construir_descripcion(id_value, metodo):
    grupos = interpretar_grupos(id_value)
    if metodo in metodos_puntuales:
        return f"Muestreo de {grupos} mediante el uso de {metodo.lower()}"
    else:
        return f"Muestreo de {grupos} mediante el método {metodo.lower()}"

df_punto_muestreo["DESCRIP"] = df_punto_muestreo.apply(
    lambda row: construir_descripcion(row["ID_MUES_PT"], Registros.loc[Registros["ID"] == row["ID_MUES_PT"], "METODO"].iloc[0]),
    axis=1
)
import pandas as pd

#-----------Asignar fecha de muestreo y temporada estacional----------------

import pandas as pd

# ------------------------------------------------------------
# 1. FECHA → FEC_MUEST (rellenando df_punto_muestreo)
# ------------------------------------------------------------

import pandas as pd

# Asegurar formato fecha
Registros["FECHA"] = pd.to_datetime(Registros["FECHA"], errors="coerce")

# Crear tabla única por ID
fecha_unica = (
    Registros.groupby("ID")["FECHA"]
    .first()                # o .min() / .max() si lo prefieres
    .reset_index()
)

# Rellenar directamente FEC_MUEST en df_punto_muestreo
df_punto_muestreo["FEC_MUEST"] = df_punto_muestreo["ID_MUES_PT"].map(
    fecha_unica.set_index("ID")["FECHA"]
)

# ------------------------------------------------------------
# 2. Asignación estacional IDEAM (rellenando df_punto_muestreo)
# ------------------------------------------------------------
def obtener_estacional(fecha):
    # Si es NaT o None
    if pd.isna(fecha):
        return None
    
    mes = fecha.month
    
    # Temporada húmeda (lluvias)
    if mes in [3, 4, 5, 9, 10, 11]:
        return "Húmedo"
    
    # Temporada seca
    if mes in [6, 7, 8, 12, 1, 2]:
        return "Seco"
    
    return None

df_punto_muestreo["ESTACIONAL"] = df_punto_muestreo["FEC_MUEST"].apply(obtener_estacional)

print(df_punto_muestreo.columns)



# ------------------------------------------------------------
# 3. Verificación
# ------------------------------------------------------------
print(df_punto_muestreo[["ID_MUES_PT", "FEC_MUEST", "ESTACIONAL"]].head())



print(type(df_punto_muestreo["FEC_MUEST"]))
print(df_punto_muestreo["FEC_MUEST"].head())

#-------------------Guardar df_punto_muestreo para inspección----------------------

df_punto_muestreo["FEC_MUEST"] = df_punto_muestreo["FEC_MUEST"].dt.strftime("%Y-%m-%d")


# Guardar el resultado en un nuevo archivo Excel

# Archivo final
output_file = os.path.join(output_folder, "14_Punto_muestreo_fauna.xlsx")

# Guardar resultado
df_punto_muestreo.to_excel(output_file, index=False)
print(f"\n Archivo guardado en: {output_file}")


#---------------------------------- Reparar y formatear archivo de Punto_muestreo_fauna -----------------------
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import os

# --- Rutas ---
ruta_original = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\14_Punto_muestreo_fauna.xlsx"
ruta_limpia = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\14_Punto_muestreo_fauna.xlsx"

# --- Verificar existencia ---
if not os.path.exists(ruta_original):
    raise FileNotFoundError(f"⚠️ No se encontró el archivo: {ruta_original}")

# --- Leer archivo dañado con pandas ---
try:
    df = pd.read_excel(ruta_original)
    print(" Archivo leído correctamente con pandas.")
except Exception as e:
    raise RuntimeError(f" No se pudo leer el archivo: {e}")

# --- Reescribir el archivo limpio ---
df.to_excel(ruta_limpia, index=False)
print(f" Archivo reparado y guardado como:\n{ruta_limpia}")

# --- Aplicar formato con openpyxl ---
from openpyxl import load_workbook

wb = load_workbook(ruta_limpia)
ws = wb.active

# --- Estilos base ---
header_fill = PatternFill(start_color='BFD8B8', end_color='BFD8B8', fill_type='solid')
header_font = Font(bold=True, color='000000', name='Calibri')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# --- Aplicar formato y reemplazar vacíos ---
for row in ws.iter_rows():
    for cell in row:
        if cell.value is None or str(cell.value).strip() == '':
            cell.value = '-'
        cell.alignment = center_align
        cell.border = thin_border

# --- Encabezado ---
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align

# --- Ajustar ancho de columnas ---
for col in ws.columns:
    max_length = 0
    column = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            length = len(str(cell.value))
            if length > max_length:
                max_length = length
    ws.column_dimensions[column].width = max_length + 3

# --- Ajustar altura de filas ---
for row in ws.iter_rows():
    ws.row_dimensions[row[0].row].height = 18

# --- Guardar cambios ---
wb.save(ruta_limpia)
print(f' Archivo formateado y reparado correctamente:\n{ruta_limpia}')






