#--------------## Cargar librerias necesarias------------------------------

# Si no las tienes instaladas, ejecuta esta celda una vez:
# Salir del interprete con: exit() exit() python   pip install tabulate pandas numpy scipy scikit-bio openpyxl
#
# !pip install pandas numpy matplotlib tabulate openpyxl

import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from tabulate import tabulate
import openpyxl

output_folder = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados"
os.makedirs(output_folder, exist_ok=True)

#--------------## Leer archivo y revisar columnas------------------------------
# Ruta del archivo
ruta = r"D:\Forestal Consultores\2026\FAUNA\BD\BD_SANROQUE.xlsx"

# Leer el archivo Excel
Registros = pd.read_excel(ruta)

# Mostrar las primeras filas
print(" Primeras filas del archivo:")
print(Registros.head())

# Mostrar nombres de las columnas
print("\n Columnas del DataFrame:")
print(Registros.columns)

#-----------------------------RUTA SIG______________

# Ruta del archivo
ruta = r"D:\Forestal Consultores\2026\FAUNA\BD\Puntos_Transectos_SIG_SanRoque.xlsx"

# Leer el archivo Excel
datosSIG = pd.read_excel(ruta)
# Mostrar las primeras filas
print(" Primeras filas del archivo:")
print(datosSIG.head())

# Mostrar nombres de las columnas
print("\n Columnas del DataFrame:")
print(datosSIG.columns)


#-----------------------------RUTA OPERADOR______________

# Ruta del archivo
ruta = r"D:\Forestal Consultores\2026\FAUNA\BD\Operador_Info_SanRoque.xlsx"

# Leer el archivo Excel
OPER_INFO = pd.read_excel(ruta)
# Mostrar las primeras filas
print(" Primeras filas del archivo:")
print(OPER_INFO.head())

# Mostrar nombres de las columnas
print("\n Columnas del DataFrame:")
print(OPER_INFO.columns)

#-----------------------------------------------
# Crear tabla MuestreoFaunaTB (estructura ANLA)
#-----------------------------------------------



#-----------------------------------------------
# Crear tabla MuestreoFaunaTB (estructura ANLA)
#-----------------------------------------------

import pandas as pd
import re

estructura_muestreo_fauna = {
    "EXPEDIENTE": str,
    "ID_MUES_PT": str,
    "ID_MUES_TR": str,
    "DETERM": "float64",
    "OT_DETERM": str,
    "DIVISION": str,
    "CLASE": str,
    "ORDEN": str,
    "FAMILIA": str,
    "GENERO": str,
    "ESPECIE": str,
    "N_COMUN": str,
    "ABUND_ABS": "float64",
    "OBSERV": str
}

# Crear DataFrame vacío
df_muestreo_fauna = pd.DataFrame(columns=estructura_muestreo_fauna.keys())
df_muestreo_fauna = df_muestreo_fauna.astype(estructura_muestreo_fauna)

print(df_muestreo_fauna.head())

#------------------------
# ============================================================
# 1. Identificar IDs de transectos y puntos
# ============================================================

Registros["ID"] = Registros["ID"].astype(str)

ids_tr = Registros[Registros["ID"].str.match(r".*T\d+$")]["ID"].unique()
ids_pt = Registros[Registros["ID"].str.match(r".*[CPR]\d+$")]["ID"].unique()


# ============================================================
# 2. Aplicar equivalencias Registros → ANLA
# ============================================================

map_equiv = {
    "DIVISION": "DIVISION",
    "Orden": "ORDEN",
    "Familia": "FAMILIA",
    "Genero": "GENERO",
    "ESPECIE": "ESPECIE",
    "N. comun": "N_COMUN",
    "INDIVIDUOS": "ABUND_ABS",
    "Conglomerado": "OBSERV",
    "CLASE": "CLASE",
}

map_equiv = {k: v for k, v in map_equiv.items() if k in Registros.columns}

Registros_eq = Registros.rename(columns=map_equiv)

# Asegurar que nunca existan columnas ID_MUES en la base antes de crearlas
for col in ["ID_MUES_PT", "ID_MUES_TR"]:
    if col in Registros_eq.columns:
        Registros_eq.drop(columns=[col], inplace=True)


# ============================================================
# 3. Separar PUNTOS y TRANSECTOS
# ============================================================

df_PT = Registros_eq[Registros_eq["ID"].isin(ids_pt)].copy()
df_TR = Registros_eq[Registros_eq["ID"].isin(ids_tr)].copy()

# Asignar cada ID en su columna correspondiente
df_PT["ID_MUES_PT"] = df_PT["ID"]
df_PT["ID_MUES_TR"] = ""

df_TR["ID_MUES_TR"] = df_TR["ID"]
df_TR["ID_MUES_PT"] = ""


# ============================================================
# 4. Eliminar la columna ID original (ya no la necesitamos)
# ============================================================

df_PT.drop(columns=["ID"], inplace=True)
df_TR.drop(columns=["ID"], inplace=True)


# ============================================================
# 5. Asegurar todas las columnas ANLA (rellenar faltantes)
# ============================================================

columnas_ANLA = list(estructura_muestreo_fauna.keys())


def completar_columnas(df):
    for col, tipo in estructura_muestreo_fauna.items():
        if col not in df.columns:
            df[col] = "" if tipo == str else np.nan
    return df[columnas_ANLA]  # ordenar columnas finales ANLA


df_PT = completar_columnas(df_PT)
df_TR = completar_columnas(df_TR)


# ============================================================
# 6. Unir PUNTOS y TRANSECTOS uno debajo del otro
# ============================================================

df_muestreo_fauna = pd.concat([df_PT, df_TR], ignore_index=True)


# ============================================================
# 7. Agregar EXPEDIENTE y convertir tipos finales
# ============================================================

df_muestreo_fauna["EXPEDIENTE"] = OPER_INFO.loc[0, "EXPEDIENTE"]
df_muestreo_fauna = df_muestreo_fauna.astype(estructura_muestreo_fauna, errors="ignore")


print("\n Tabla final ANLA generada correctamente:")
print(df_muestreo_fauna.head())
print(f"\nTotal de registros: {len(df_muestreo_fauna)}")









# ============================================================
# 4. Completar METODOLOGIA, DETERM y OT_DETERM
# ============================================================

# -----------------------------------------
# 4.1 Crear diccionario único ID → METODOLOGIA
# -----------------------------------------
metodologia_dict = (
    Registros.groupby("ID")["METODOLOGIA"]
    .first()              # garantiza un valor único por ID
    .to_dict()
)

# -----------------------------------------
# 4.2 Mapear METODOLOGIA a TR y PT
# -----------------------------------------
df_muestreo_fauna["METODOLOGIA"] = (
    df_muestreo_fauna["ID_MUES_TR"].map(metodologia_dict)
).fillna(
    df_muestreo_fauna["ID_MUES_PT"].map(metodologia_dict)
)

# -----------------------------------------
# 4.3 Diccionario de determinación ANLA
# -----------------------------------------
determinacion_dict = {
    "AUDITIVO": "Detección auditiva",
    "AVISTAMIENTO": "Observación",
    "CAPTURA": "Captura de individuos",
    "CUEVA": "Huellas",
    "ENTREVISTA": "Otro",
    "FOTOGRAFIA": "Observación",
    "HECES": "Heces",
    "Huellas": "Huellas",
    "INFORMACION MCNUP": "Otro",
    "MARCAS DE PRESENCIA": "Huellas",
    "Rastros": "Rastros",
    "VIDEO": "Observación"
}

def map_deter(m):
    if pd.isna(m):
        return "Otro"
    x = str(m).strip().upper()
    return determinacion_dict.get(x, "Otro")

# -----------------------------------------
# 4.4 Completar DETERM
# -----------------------------------------
df_muestreo_fauna["DETERM"] = df_muestreo_fauna["METODOLOGIA"].apply(map_deter)

# -----------------------------------------
# 4.5 Completar OT_DETERM
# Solo se llena cuando DETERM = "Otro"
# -----------------------------------------
df_muestreo_fauna["OT_DETERM"] = df_muestreo_fauna.apply(
    lambda row: row["METODOLOGIA"] if row["DETERM"] == "Otro" else None,
    axis=1
)

print(df_muestreo_fauna)


# 4.6 Eliminar columna METODOLOGIA para exportar
df_muestreo_fauna.drop(columns=["METODOLOGIA"], inplace=True)



print(df_muestreo_fauna)




















#-------------------Guardar df_punto_muestreo para inspección----------------------


# Guardar el resultado en un nuevo archivo Excel

# Archivo final
output_file = os.path.join(output_folder, "14.2_Muestreo_Fauna_TB.xlsx")

# Guardar resultado
df_muestreo_fauna.to_excel(output_file, index=False)
print(f"\n Archivo guardado en: {output_file}")


#---------------------------------- Reparar y formatear archivo de Transecto_muestreo_fauna -----------------------
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import os

# --- Rutas ---
ruta_original = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\14.2_Muestreo_Fauna_TB.xlsx"
ruta_limpia = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\14.2_Muestreo_Fauna_TB.xlsx"

# --- Verificar existencia ---
if not os.path.exists(ruta_original):
    raise FileNotFoundError(f"⚠️ No se encontró el archivo: {ruta_original}")

# --- Leer archivo dañado con pandas ---
try:
    df = pd.read_excel(ruta_original)
    print(" Archivo leído correctamente con pandas.")
except Exception as e:
    raise RuntimeError(f" No se pudo leer el archivo: {e}")

# --- Reescribir el archivo limpio ---
df.to_excel(ruta_limpia, index=False)
print(f" Archivo reparado y guardado como:\n{ruta_limpia}")

# --- Aplicar formato con openpyxl ---
from openpyxl import load_workbook

wb = load_workbook(ruta_limpia)
ws = wb.active

# --- Estilos base ---
header_fill = PatternFill(start_color='BFD8B8', end_color='BFD8B8', fill_type='solid')
header_font = Font(bold=True, color='000000', name='Calibri')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# --- Aplicar formato y reemplazar vacíos ---
for row in ws.iter_rows():
    for cell in row:
        if cell.value is None or str(cell.value).strip() == '':
            cell.value = '-'
        cell.alignment = center_align
        cell.border = thin_border

# --- Encabezado ---
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align

# --- Ajustar ancho de columnas ---
for col in ws.columns:
    max_length = 0
    column = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            length = len(str(cell.value))
            if length > max_length:
                max_length = length
    ws.column_dimensions[column].width = max_length + 3

# --- Ajustar altura de filas ---
for row in ws.iter_rows():
    ws.row_dimensions[row[0].row].height = 18

# --- Guardar cambios ---
wb.save(ruta_limpia)
print(f' Archivo formateado y reparado correctamente:\n{ruta_limpia}')











