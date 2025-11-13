
#--------------## Cargar librerias necesarias------------------------------
import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from tabulate import tabulate
import openpyxl

# Carpeta donde guardarás los gráficos (solo una vez)
output_folder = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados"
os.makedirs(output_folder, exist_ok=True)
#--------------## Leer archivo y revisar columnas------------------------------
# Ruta del archivo
ruta = r"D:\CORPONOR 2025\Backet\python_Proyect\data\POF_ZULIA_2025_BD_AVES_MAMIFEROS.xlsx"

# Leer el archivo Excel
Registros = pd.read_excel(ruta)

# Mostrar las primeras filas
print("📄 Primeras filas del archivo:")
print(Registros.head())

# Mostrar nombres de las columnas
print("\n📋 Columnas del DataFrame:")
print(Registros.columns)

# --------------------------------------------------
# 2️⃣ Funciones para abreviar nombres de coberturas
# --------------------------------------------------

def generar_abreviacion(nombre):
    """
    Genera abreviaciones automáticas a partir de nombres de coberturas.
    Ejemplo: 'Bosque de galería y ripario' → 'Bgr'
    """
    # Convertir a minúsculas y dividir en palabras
    palabras = nombre.lower().split()

    # Eliminar conectores comunes
    palabras = [p for p in palabras if p not in ['de', 'del', 'la', 'el', 'y', 'con', 'en', 'los', 'las']]

    # Tomar la primera letra de cada palabra
    abreviacion = ''.join([p[0] for p in palabras])

    # Asegurar que tenga al menos 3 caracteres
    if len(abreviacion) < 3:
        abreviacion = abreviacion.ljust(3, '_')

    return abreviacion.capitalize()


def abreviar_coberturas(df, columna='COBERTURA'):
    """
    Crea un diccionario de abreviaciones y reemplaza los nombres en el DataFrame.
    """
    coberturas_unicas = df[columna].dropna().unique()
    abreviaciones = {c: generar_abreviacion(c) for c in coberturas_unicas}

    print("\n🔤 Abreviaciones generadas automáticamente:")
    for original, abrev in abreviaciones.items():
        print(f"  {original} → {abrev}")

    # Reemplazar en el DataFrame
    df[columna] = df[columna].replace(abreviaciones)

    return df, abreviaciones


# --- Aplicar las abreviaciones ---
Registros, abreviaciones_cobertura = abreviar_coberturas(Registros, columna='COBERTURA')

import pandas as pd
import matplotlib.pyplot as plt
import numpy as np


# --- Verificar columnas clave ---
print(Registros.columns)

# --- Convertir 'INDIVIDUOS' a numérico ---
Registros["INDIVIDUOS"] = pd.to_numeric(Registros["INDIVIDUOS"], errors="coerce")

# --- Eliminar filas sin especie o sin valor de individuos ---
Registros = Registros.dropna(subset=["ESPECIE", "INDIVIDUOS"])

# --- 3. Asegurar que los nombres están bien ---
# (Por si hay diferencias en mayúsculas/minúsculas)
#Registros.columns = Registros.columns.str.strip().str.upper()

# --- 4. Calcular abundancia total por especie (todas las coberturas) ---
abund_total = (
    Registros.groupby("ESPECIE")["INDIVIDUOS"]
    .sum()
    .sort_values(ascending=False)
    .reset_index()
)
abund_total["LOG10_ABUND"] = np.log10(abund_total["INDIVIDUOS"] + 1)
abund_total["RANGO"] = range(1, len(abund_total) + 1)
abund_total["COBERTURA"] = "TOTAL"

# --- 5. Calcular abundancia por cobertura ---
abund_cob = (
    Registros.groupby(["COBERTURA", "ESPECIE"])["INDIVIDUOS"]
    .sum()
    .reset_index()
)

# Crear una lista para guardar todos los DataFrames (total + coberturas)
curvas = [abund_total]

# --- 6. Procesar cada cobertura individual ---
for cobertura, df in abund_cob.groupby("COBERTURA"):
    df_sorted = df.sort_values(by="INDIVIDUOS", ascending=False).reset_index(drop=True)
    df_sorted["LOG10_ABUND"] = np.log10(df_sorted["INDIVIDUOS"] + 1)
    df_sorted["RANGO"] = range(1, len(df_sorted) + 1)
    df_sorted["COBERTURA"] = cobertura
    curvas.append(df_sorted)

# --- 7. Unir todas las curvas ---
curvas_df = pd.concat(curvas, ignore_index=True)

# --- 8. Exportar a Excel ---
ruta_salida_xlsx = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\Curvas_Rango_Abundancia.xlsx"

# Exportar el DataFrame a Excel
curvas_df.to_excel(ruta_salida_xlsx, index=False)

print(f"✅ Archivo Excel exportado exitosamente en: {ruta_salida_xlsx}")

#---------------------------Dar formato a tabla curvas rango-abundancia por cobertura----------------------------
#---------------------------------- Reparar y formatear archivo de Curvas_Rango_Abundancia -----------------------
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import os

# --- Rutas ---
ruta_original = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\Curvas_Rango_Abundancia.xlsx"
ruta_limpia = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\Curvas_Rango_Abundancia.xlsx"

# --- Verificar existencia ---
if not os.path.exists(ruta_original):
    raise FileNotFoundError(f"⚠️ No se encontró el archivo: {ruta_original}")

# --- Leer archivo dañado con pandas ---
try:
    df = pd.read_excel(ruta_original)
    print("✅ Archivo leído correctamente con pandas.")
except Exception as e:
    raise RuntimeError(f"❌ No se pudo leer el archivo: {e}")

# --- Reescribir el archivo limpio ---
df.to_excel(ruta_limpia, index=False)
print(f"🧹 Archivo reparado y guardado como:\n{ruta_limpia}")

# --- Aplicar formato con openpyxl ---
from openpyxl import load_workbook

wb = load_workbook(ruta_limpia)
ws = wb.active

# --- Estilos base ---
header_fill = PatternFill(start_color='BFD8B8', end_color='BFD8B8', fill_type='solid')
header_font = Font(bold=True, color='000000', name='Calibri')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# --- Aplicar formato y reemplazar vacíos ---
for row in ws.iter_rows():
    for cell in row:
        if cell.value is None or str(cell.value).strip() == '':
            cell.value = '-'
        cell.alignment = center_align
        cell.border = thin_border

# --- Encabezado ---
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align

# --- Ajustar ancho de columnas ---
for col in ws.columns:
    max_length = 0
    column = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            length = len(str(cell.value))
            if length > max_length:
                max_length = length
    ws.column_dimensions[column].width = max_length + 3

# --- Ajustar altura de filas ---
for row in ws.iter_rows():
    ws.row_dimensions[row[0].row].height = 18

# --- Guardar cambios ---
wb.save(ruta_limpia)
print(f'📘 Archivo formateado y reparado correctamente:\n{ruta_limpia}')




# --- 8. Graficar ---


import matplotlib.pyplot as plt

# --- 1. Calcular la abundancia total por cobertura ---
abundancia_por_cobertura = (
    curvas_df.groupby("COBERTURA")["INDIVIDUOS"].sum().sort_values(ascending=False)
)

# --- 2. Ordenar coberturas (TOTAL primero si existe) ---
orden_coberturas = ["TOTAL"] + [c for c in abundancia_por_cobertura.index if c != "TOTAL"]

# --- Calcular límites globales del eje Y ---
ymin = curvas_df["LOG10_ABUND"].min() - 0.2
ymax = curvas_df["LOG10_ABUND"].max() + 0.2

# --- 3. Crear figura (una columna por cobertura) ---
fig, axes = plt.subplots(1, len(orden_coberturas), figsize=(4 * len(orden_coberturas), 5), sharey=True)
if len(orden_coberturas) == 1:
    axes = [axes]

# --- 4. Función para abreviar nombres de especies ---
def abreviar_nombre(nombre):
    partes = nombre.split()
    if len(partes) >= 2:
        return f"{partes[0][0]}. {partes[1]}"
    else:
        return nombre

# --- 5. Dibujar las curvas de rango-abundancia ---
for i, (ax, cobertura) in enumerate(zip(axes, orden_coberturas)):
    sub = curvas_df[curvas_df["COBERTURA"] == cobertura].sort_values("RANGO")
    
    if sub.empty:
        ax.text(0.5, 0.5, "Sin datos", ha="center", va="center", fontsize=12)
        continue

    # Curva de abundancia
    ax.plot(sub["RANGO"], sub["LOG10_ABUND"], marker="o", linestyle="-", color=f"C{i}")

    # Etiquetas (5 especies salteadas por curva)
    etiquetas_idx = sub.index[::max(1, len(sub)//5)]
    for j in etiquetas_idx:
        especie = abreviar_nombre(sub.loc[j, "ESPECIE"])
        ax.text(sub.loc[j, "RANGO"], sub.loc[j, "LOG10_ABUND"] + 0.1, especie,
                rotation=45, fontsize=8)

    ax.set_title(cobertura, fontsize=11, fontweight="bold")
    ax.set_xlabel("Rango de especies")
    if i == 0:
        ax.set_ylabel("Log₁₀ (Abundancia + 1)")
    ax.grid(True, linestyle="--", alpha=0.5)

# --- 6. Ajustes finales y visualización ---
plt.suptitle("Curvas rango-abundancia por cobertura", fontsize=14, fontweight="bold")
plt.tight_layout(rect=[0, 0, 1, 0.95])

plt.show()




#-------------------grafico curvas rango-abundancia estilo publicación--------------------

import matplotlib.pyplot as plt
import numpy as np
from adjustText import adjust_text

#----------------------------------- FUNCIÓN PARA ABREVIAR NOMBRES -----------------------------------
def abreviar_nombre(nombre):
    """Convierte 'Chlorophanes atratus' → 'C. atratus'."""
    partes = nombre.split()
    if len(partes) >= 2:
        return f"{partes[0][0]}. {partes[1]}"
    else:
        return nombre

#----------------------------------- FUNCIÓN PRINCIPAL -----------------------------------
def plot_curvas_comparativas_estilo(curvas_df, figsize=(12, 7)):
    """
    Dibuja curvas rango-abundancia desplazadas horizontalmente,
    con etiquetas ajustadas automáticamente (sin solapamientos)
    y estilo limpio tipo publicación.
    """
    plt.style.use('seaborn-v0_8-whitegrid')

    colores = plt.cm.tab10.colors  # paleta
    abundancia_por_cobertura = (
        curvas_df.groupby("COBERTURA")["INDIVIDUOS"].sum().sort_values(ascending=False)
    )

    fig, ax = plt.subplots(figsize=figsize)
    textos = []
    offset = 0  # desplazamiento horizontal

    #----------------------------------- GRAFICAR CADA CURVA -----------------------------------
    for i, (cobertura, _) in enumerate(abundancia_por_cobertura.items()):
        sub = curvas_df[curvas_df["COBERTURA"] == cobertura].copy()
        if sub.empty:
            continue

        sub = sub.sort_values(by="INDIVIDUOS", ascending=False).reset_index(drop=True)
        sub["RANGO"] = np.arange(1, len(sub) + 1) + offset
        sub["LOG10_ABUND"] = np.log10(sub["INDIVIDUOS"] + 1)

        # --- Graficar curva tipo escalon con puntos ---
        ax.plot(
            sub["RANGO"], sub["LOG10_ABUND"],
            marker="o", markersize=4, linestyle="-",
            color=colores[i % len(colores)], linewidth=1.8, alpha=0.9,
            label=cobertura, zorder=3
        )

        #----------------------------------- ETIQUETAS (5 especies distribuidas) -----------------------------------
        n_especies = len(sub)
        indices_etiqueta = [0]  # siempre la más abundante

        if n_especies > 5:
            indices_otros = np.linspace(1, n_especies - 1, 4, dtype=int)
            indices_etiqueta.extend(indices_otros)
        else:
            indices_etiqueta = list(range(n_especies))

        for j_idx in indices_etiqueta:
            especie = abreviar_nombre(sub.loc[j_idx, "ESPECIE"])
            x = sub.loc[j_idx, "RANGO"]
            y = sub.loc[j_idx, "LOG10_ABUND"]

            txt = ax.text(
                x, y, especie, fontsize=10, ha='center', va='bottom',
                rotation=0, style='italic', color='black', zorder=5
            )
            textos.append(txt)
        # --- Control preciso del desplazamiento entre curvas ---
        factor_desplazamiento = 1  # 🔹 AJUSTA AQUÍ (por ejemplo 0.2, 

        offset += len(sub) * factor_desplazamiento  # desplazar siguiente curva

    #----------------------------------- AJUSTAR ETIQUETAS -----------------------------------
    adjust_text(
        textos, ax=ax,
        arrowprops=dict(arrowstyle="-", lw=0.7, color="gray", alpha=0.6),
        expand_points=(1.2, 1.2), force_text=0.7
    )

    #----------------------------------- ESTILO Y DETALLES -----------------------------------
    ax.set_title("Curvas rango–abundancia comparativas por cobertura", fontsize=14, fontweight='bold')
    ax.set_xlabel("Rango de especies", fontsize=12)
    ax.set_ylabel("Log10 (Abundancia + 1)", fontsize=12)

    # --- Grilla suave y ejes sutiles ---
    ax.grid(True, which='major', linestyle='--', alpha=0.4)
    ax.grid(True, which='minor', linestyle=':', alpha=0.2)
    ax.minorticks_on()

    for spine in ax.spines.values():
        spine.set_visible(True)
        spine.set_linewidth(0.8)
        spine.set_color("gray")

    # --- Leyenda arriba a la derecha ---
    ax.legend(title="Cobertura", loc="upper right", frameon=True,
              fontsize=12, title_fontsize=14)

    plt.tight_layout()
    plt.show()

#----------------------------------- GUARDAR GRÁFICO -----------------------------------
    import os  # Para manejar rutas y crear carpetas
    
       #----------------------------------- GUARDAR GRÁFICO -----------------------------------
    output_folder = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados"
    os.makedirs(output_folder, exist_ok=True)
    ruta_salida = os.path.join(output_folder, "Curvas_Rango_Abundancia_Combinado.png")

    fig.savefig(ruta_salida, dpi=300, bbox_inches="tight")
    print(f"✅ Gráfico guardado en:\n{ruta_salida}")


#----------------------------------- USO -----------------------------------
plot_curvas_comparativas_estilo(curvas_df, figsize=(12, 7))
#-----------------------------fin del Grafico unificado-----------------------------







#----------------------------Graficar un panel por cobertura----------------------------

#------------------- GRAFICAR UN PANEL POR COBERTURA --------------------

import matplotlib.pyplot as plt
import numpy as np
from adjustText import adjust_text

#----------------------------------- FUNCIÓN PARA ABREVIAR NOMBRES -----------------------------------
def abreviar_nombre(nombre):
    """Convierte 'Chlorophanes atratus' → 'C. atratus'."""
    partes = nombre.split()
    if len(partes) >= 2:
        return f"{partes[0][0]}. {partes[1]}"
    else:
        return nombre

#----------------------------------- FUNCIÓN POR COBERTURA -----------------------------------
def plot_curvas_por_cobertura_estilo(curvas_df, figsize=(8, 6)):
    """
    Genera un gráfico independiente por cobertura (una figura por cobertura).
    Mantiene el estilo tipo publicación, etiquetas ajustadas y escala log10.
    """

    plt.style.use('seaborn-v0_8-whitegrid')
    colores = plt.cm.tab10.colors  # paleta

    # --- Identificar coberturas en orden descendente de abundancia total ---
    abundancia_por_cobertura = (
        curvas_df.groupby("COBERTURA")["INDIVIDUOS"].sum().sort_values(ascending=False)
    )

    # --- Iterar por cobertura ---
    for i, (cobertura, _) in enumerate(abundancia_por_cobertura.items()):
        sub = curvas_df[curvas_df["COBERTURA"] == cobertura].copy()
        if sub.empty:
            continue

        # --- Preparar datos ---
        sub = sub.sort_values(by="INDIVIDUOS", ascending=False).reset_index(drop=True)
        sub["RANGO"] = np.arange(1, len(sub) + 1)
        sub["LOG10_ABUND"] = np.log10(sub["INDIVIDUOS"] + 1)

        # --- Crear figura ---
        fig, ax = plt.subplots(figsize=figsize)
        color = colores[i % len(colores)]

        # --- Graficar curva ---
        ax.plot(
            sub["RANGO"], sub["LOG10_ABUND"],
            marker="o", markersize=5, linestyle="-",
            color=color, linewidth=1.8, alpha=0.9,
            label=cobertura, zorder=3
        )

        #----------------------------------- ETIQUETAS -----------------------------------
        textos = []
        n_especies = len(sub)
        indices_etiqueta = [0]  # Siempre incluir la más abundante

        if n_especies > 5:
            indices_otros = np.linspace(1, n_especies - 1, 4, dtype=int)
            indices_etiqueta.extend(indices_otros)
        else:
            indices_etiqueta = list(range(n_especies))

        for j_idx in indices_etiqueta:
            especie = abreviar_nombre(sub.loc[j_idx, "ESPECIE"])
            x = sub.loc[j_idx, "RANGO"]
            y = sub.loc[j_idx, "LOG10_ABUND"]

            txt = ax.text(
                x, y, especie, fontsize=9, ha='center', va='bottom',
                rotation=0, style='italic', color='black', zorder=5
            )
            textos.append(txt)

        # --- Ajustar etiquetas automáticamente ---
        adjust_text(
            textos, ax=ax,
            arrowprops=dict(arrowstyle="-", lw=0.7, color="gray", alpha=0.6),
            expand_points=(1.3, 1.3), force_text=0.8
        )

        #----------------------------------- ESTILO -----------------------------------
        ax.set_title(f"Curva rango–abundancia: {cobertura}", fontsize=13, fontweight='bold')
        ax.set_xlabel("Rango de especies", fontsize=12)
        ax.set_ylabel("Log10 (Abundancia + 1)", fontsize=12)

        # --- Grilla suave y ejes visibles ---
        ax.grid(True, which='major', linestyle='--', alpha=0.4)
        ax.grid(True, which='minor', linestyle=':', alpha=0.2)
        ax.minorticks_on()

        for spine in ax.spines.values():
            spine.set_visible(True)
            spine.set_linewidth(0.8)
            spine.set_color("gray")

        # --- Leyenda discreta arriba a la derecha ---
        ax.legend(title="Cobertura", loc="upper right", frameon=True,
                  fontsize=10, title_fontsize=11)

        plt.tight_layout()
        plt.show()
        
        #----------------------------------- GUARDAR GRÁFICO -----------------------------------
        #import os  # ✅ dentro de la función (con la misma indentación que plt.show)

        output_folder = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados"
        os.makedirs(output_folder, exist_ok=True)

        nombre_archivo = f"Curva_Rango_Abundancia_{cobertura.replace(' ', '_')}.png"
        ruta_salida = os.path.join(output_folder, nombre_archivo)
        fig.savefig(ruta_salida, dpi=300, bbox_inches="tight")
        print(f"✅ Gráfico guardado: {ruta_salida}")

        plt.close(fig)  # ✅ también dentro del for
#----------------------------------- USO -----------------------------------

plot_curvas_por_cobertura_estilo(curvas_df, figsize=(8, 6))














#------------------------------------fin del Grafico por cobertura----------------------------

#------------------Graficar todas las coberturas en una grilla estilo publicación------------------

#------------------- Curvas rango–abundancia por cobertura (en grilla, estilo publicación) --------------------
import matplotlib.pyplot as plt
import numpy as np
from adjustText import adjust_text

# --- Función para abreviar nombres científicos (C. atratus, etc.) ---
def abreviar_nombre(nombre):
    """Convierte 'Chlorophanes atratus' → 'C. atratus'."""
    partes = nombre.split()
    if len(partes) >= 2:
        return f"{partes[0][0]}. {partes[1]}"
    else:
        return nombre

# --- Función principal ---
def plot_curvas_grid_estilo(curvas_df, ncols=3, figsize_per_col=(5, 6)):
    """
    Dibuja curvas rango-abundancia en una grilla estilo publicación.
    - ncols: número de columnas en la grilla.
    - figsize_per_col: tamaño base de cada columna.
    - Ajusta etiquetas automáticamente y mantiene escala uniforme entre paneles.
    """

    plt.style.use('seaborn-v0_8-whitegrid')  # 🎨 estilo limpio y moderno

    # --- Ordenar coberturas (TOTAL primero si existe) ---
    abund_por_cob = curvas_df.groupby("COBERTURA")["INDIVIDUOS"].sum().sort_values(ascending=False)
    orden_cob = ["TOTAL"] + [c for c in abund_por_cob.index if c != "TOTAL" and c in curvas_df["COBERTURA"].unique()]

    # --- Límites globales del eje Y (mismo para todas las coberturas) ---
    ymin = curvas_df["LOG10_ABUND"].min() - 0.2
    ymax = curvas_df["LOG10_ABUND"].max() + 0.2

    # --- Calcular número de filas/columnas y tamaño total ---
    n = len(orden_cob)
    ncols = max(1, int(ncols))
    nrows = int(np.ceil(n / ncols))
    fig_w = figsize_per_col[0] * ncols
    fig_h = figsize_per_col[1] * nrows + 1.5
    fig, axes = plt.subplots(nrows, ncols, figsize=(fig_w, fig_h), sharey=True)

    # --- Normalizar ejes (por si hay una sola fila o columna) ---
    axes = np.atleast_2d(axes).reshape(nrows, ncols)

    # --- Paleta de colores consistente ---
    colores = plt.cm.tab10.colors

    # --- Iterar coberturas y graficar ---
    idx = 0
    for r in range(nrows):
        for c in range(ncols):
            ax = axes[r, c]
            if idx >= n:
                ax.axis('off')  # Apagar panel vacío
                continue

            cov = orden_cob[idx]
            sub = curvas_df[curvas_df["COBERTURA"] == cov].copy()

            if sub.empty:
                ax.text(0.5, 0.5, "Sin datos", ha="center", va="center", fontsize=12)
            else:
                # --- Preparar datos ordenados por abundancia ---
                sub = sub.sort_values(by="INDIVIDUOS", ascending=False).reset_index(drop=True)
                sub["RANGO"] = np.arange(1, len(sub) + 1)
                sub["LOG10_ABUND"] = np.log10(sub["INDIVIDUOS"] + 1)

                # --- Graficar curva principal ---
                ax.plot(
                    sub["RANGO"], sub["LOG10_ABUND"],
                    marker="o", markersize=4,
                    linestyle="-", color=colores[idx % len(colores)],
                    linewidth=1.8, alpha=0.9, zorder=3
                )

                # --- Etiquetar hasta 5 especies (primera + 4 distribuidas) ---
                textos = []
                n_especies = len(sub)
                indices_etiqueta = [0]  # siempre la más abundante

                if n_especies > 5:
                    indices_otros = np.linspace(1, n_especies - 1, 4, dtype=int)
                    indices_etiqueta.extend(indices_otros)
                else:
                    indices_etiqueta = list(range(n_especies))

                # --- Crear etiquetas ---
                for j_idx in indices_etiqueta:
                    especie = abreviar_nombre(sub.loc[j_idx, "ESPECIE"])
                    x = sub.loc[j_idx, "RANGO"]
                    y = sub.loc[j_idx, "LOG10_ABUND"]
                    txt = ax.text(
                        x, y, especie, fontsize=8, ha='center', va='bottom',
                        rotation=45, style='italic', color='black', zorder=5
                    )
                    textos.append(txt)

                # --- Ajuste automático de etiquetas para evitar solapes ---
                if textos:
                    adjust_text(
                        textos, ax=ax,
                        arrowprops=dict(arrowstyle='-', lw=0.6, color='gray', alpha=0.6),
                        expand_points=(1.2, 1.2), force_text=0.7
                    )

            # --- Personalización del panel ---
            ax.set_title(cov, fontsize=12, fontweight="bold", pad=10)
            ax.set_ylim(ymin, ymax)
            ax.grid(True, linestyle="--", alpha=0.4)
            ax.minorticks_on()
            for spine in ax.spines.values():
                spine.set_visible(True)
                spine.set_linewidth(0.8)
                spine.set_color("gray")

            # --- Etiquetas de ejes (solo bordes exteriores) ---
            if r == nrows - 1:
                ax.set_xlabel("Rango de especies", fontsize=10)
            else:
                ax.set_xlabel("")
            if c == 0:
                ax.set_ylabel("Log₁₀ (Abundancia + 1)", fontsize=10)
            else:
                ax.set_ylabel("")

            idx += 1

    # --- Título general del gráfico ---
    plt.suptitle("Curvas rango_abundancia por cobertura", fontsize=15, fontweight="bold", y=0.99)
    plt.tight_layout(rect=[0, 0, 1, 0.97], h_pad=1.5, w_pad=0.7)
    # --- Guardar figura ---
    nombre_archivo = "Curvas_rango_abundancia_por_cobertura_grid.png"
    ruta_salida = os.path.join(output_folder, nombre_archivo)
    fig.savefig(ruta_salida, dpi=300, bbox_inches="tight")
    print(f"✅ Gráfico guardado en: {ruta_salida}") 
   
   
   
    plt.show()


# Ejemplo de uso:
plot_curvas_grid_estilo(curvas_df, ncols=3)

#------------------------------------fin del Grafico en grilla----------------------------
