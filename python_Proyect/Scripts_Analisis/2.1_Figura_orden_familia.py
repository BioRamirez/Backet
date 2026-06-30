
#------------------Figura de Ordenes familias---------------


import matplotlib
matplotlib.use('Agg')
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os

# Ruta del archivo
# Ruta del archivo
ruta = r"D:\Forestal Consultores\2026\FAUNA\BD\AVES\Aves_Secundario_San_Roque.xlsx"
# Leer el archivo Excel
Registros = pd.read_excel(ruta)



# =========================================================
# FILTRAR UNA SOLA CLASE
# =========================================================

grupo = "AVES"   # AVES, MAMIFEROS, REPTILES,ANFIBIOS etc.

Registros = Registros[
    Registros["CLASE"].astype(str).str.upper() == grupo.upper()
].copy()

# Reiniciar índice
Registros.reset_index(drop=True, inplace=True)

# =========================================================
# VERIFICACIÓN
# =========================================================

print(f"\nGrupo seleccionado: {grupo}")
print(f"Número de registros: {len(Registros)}")

print("\nPrimeras filas:")
print(Registros.head())

# Mostrar las primeras filas
print(" Primeras filas del archivo:")
print(Registros.head())

# Mostrar nombres de las columnas
print("\n Columnas del DataFrame:")
print(Registros.columns)

# --- Copiar dataframe base ---
df = Registros.copy()

# --- Si tu DataFrame ya está cargado en Python como df, úsalo directamente ---
tabla = df.copy()

# --- Limpiar nombres de columnas por seguridad ---
tabla.columns = tabla.columns.str.strip()

# Limpiar nombres de familia
tabla['Familia'] = tabla['Familia'].astype(str).str.strip()

# Eliminar filas con nombres de familia vacíos o 'nan'
tabla = tabla[tabla['Familia'].notna()]
tabla = tabla[tabla['Familia'] != '']


# --- Asegurar que las columnas requeridas existen ---
tabla = tabla.dropna(subset=['Orden', 'Familia', 'Especie'])

# --- Crear tabla dinámica: número de especies únicas por Orden y Familia ---
pivot_df = (
    tabla.groupby(['Orden', 'Familia'])['Especie']
    .nunique()
    .reset_index()
    .pivot(index='Orden', columns='Familia', values='Especie')
    .fillna(0)
    .astype(int)
)
# --- Ruta de salida ---
output_folder = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados"
os.makedirs(output_folder, exist_ok=True)  # Crea la carpeta si no existe

# --- Exportar tabla a Excel ---
excel_path = os.path.join(output_folder, '2.1_Riqueza_Orden_Familia.xlsx')
pivot_df.to_excel(excel_path, sheet_name='Tabla_dinamica')


# --- Crear gráfico de barras apiladas horizontal ---
sns.set(style='whitegrid')
fig, ax = plt.subplots(figsize=(12, 8))

pivot_df.plot(
    kind='barh',
    stacked=True,
    colormap='tab20',
    edgecolor='black',
    ax=ax
)

# --- Añadir etiquetas dentro de las barras ---
for container in ax.containers:
    # etiquetas solo si el valor del segmento > 0
    labels = [f'{w.get_width():.0f}' if w.get_width() > 0 else '' for w in container]
    ax.bar_label(
        container,
        labels=labels,
        label_type='center',     # posición centrada dentro del bloque
        fontsize=7,
        color='black',
        weight='bold'
    )

# --- Etiquetas y formato ---
ax.set_title('', fontsize=14, fontweight='bold')
ax.set_xlabel('Número de especies')
ax.set_ylabel('Orden')
# --- Ajuste de la leyenda para ocupar todo el alto ---
ax.legend(
    title='Familia',
    bbox_to_anchor=(1.02, 0, 0.25, 1),  # [x0, y0, ancho, alto] → ocupa toda la altura
    loc='upper left',
    ncol=1,                             # número de columnas
    fontsize=8,
    title_fontsize=9,
    frameon=False,
    mode='expand',                      # distribuye las entradas verticalmente en todo el alto
    borderaxespad=0.0,
    columnspacing=1.2,
    labelspacing=0.8
)

plt.tight_layout()


# --- Ruta de salida ---
output_folder = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados"
os.makedirs(output_folder, exist_ok=True)  # crea la carpeta si no existe

# --- Guardar el gráfico como imagen en la carpeta Resultados ---
img_path = os.path.join(output_folder, '2.1.1_Grafico_Riqueza_Orden_Familia.png')
plt.savefig(img_path, dpi=300, bbox_inches='tight')
plt.close()

# --- Insertar el gráfico en el Excel ---
excel_path = os.path.join(output_folder, '2.1_Riqueza_Orden_Familia.xlsx')
wb = load_workbook(excel_path)
ws = wb.create_sheet('Grafico')

# Insertar la imagen
img = Image(img_path)
ws.add_image(img, 'A1')

# Guardar el Excel final
wb.save(excel_path)

print(' Tabla dinámica y gráfico exportados en:', excel_path)


#---------------------------------- Reparar y formatear archivo de riqueza orden familia -----------------------
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import os

# --- Rutas ---
ruta_original = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\2.1_Riqueza_Orden_Familia.xlsx"
ruta_limpia = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\2.1_Riqueza_Orden_Familia.xlsx"

# --- Verificar existencia ---
if not os.path.exists(ruta_original):
    raise FileNotFoundError(f"⚠️ No se encontró el archivo: {ruta_original}")

# --- Leer archivo dañado con pandas ---
try:
    df = pd.read_excel(ruta_original)
    print(" Archivo leído correctamente con pandas.")
except Exception as e:
    raise RuntimeError(f" No se pudo leer el archivo: {e}")

# --- Reescribir el archivo limpio ---
df.to_excel(ruta_limpia, index=False)
print(f" Archivo reparado y guardado como:\n{ruta_limpia}")

# --- Aplicar formato con openpyxl ---
from openpyxl import load_workbook

wb = load_workbook(ruta_limpia)
ws = wb.active

# --- Estilos base ---
header_fill = PatternFill(start_color='BFD8B8', end_color='BFD8B8', fill_type='solid')
header_font = Font(bold=True, color='000000', name='Calibri')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# --- Aplicar formato y reemplazar vacíos ---
for row in ws.iter_rows():
    for cell in row:
        if cell.value is None or str(cell.value).strip() == '':
            cell.value = '-'
        cell.alignment = center_align
        cell.border = thin_border

# --- Encabezado ---
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align

# --- Ajustar ancho de columnas ---
for col in ws.columns:
    max_length = 0
    column = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            length = len(str(cell.value))
            if length > max_length:
                max_length = length
    ws.column_dimensions[column].width = max_length + 3

# --- Ajustar altura de filas ---
for row in ws.iter_rows():
    ws.row_dimensions[row[0].row].height = 18

# --- Guardar cambios ---
wb.save(ruta_limpia)
print(f' Archivo formateado y reparado correctamente:\n{ruta_limpia}')
#------------------Fin Formaterar tabla de Riqueza_Orden_Familia------------------#















#------------------Figura de Ordenes familias (VERSIÓN PROFESIONAL)---------------

import matplotlib
matplotlib.use('Agg')
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import os

# =============================
# Preparación de datos
# =============================

pivot_df = (
    tabla.groupby(['Orden', 'Familia'])['Especie']
    .nunique()
    .reset_index()
    .pivot(index='Orden', columns='Familia', values='Especie')
    .fillna(0)
    .astype(int)
)

# =============================
# ESTILO PROFESIONAL PREMIUM
# =============================
sns.set_theme(style="whitegrid")
palette = sns.color_palette("Paired", n_colors=len(pivot_df.columns))

plt.figure(figsize=(15, 10))

ax = pivot_df.plot(
    kind='barh',
    stacked=True,
    color=palette,
    edgecolor='black',
    linewidth=0.6,
    figsize=(15, 10)
)

# =============================
# Etiquetas internas mejoradas
# =============================
for container in ax.containers:
    for bar in container:
        width = bar.get_width()
        if width > 0:
            ax.text(
                bar.get_x() + width / 2,
                bar.get_y() + bar.get_height() / 2,
                f'{int(width)}',
                ha='center', va='center',
                fontsize=14, fontweight='bold'
            )

# =============================
# Ajustes estéticos
# =============================
#Riqueza de Familias por Orden
ax.set_title(
    "",
    fontsize=18, fontweight="bold", pad=14
)

ax.set_xlabel("Número de especies", fontsize=14)
ax.set_ylabel("Orden", fontsize=14)

ax.tick_params(axis='both', labelsize=12)

# Quitar bordes superiores y derecho
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)

# Margen superior mejorado
plt.ylim(-0.5, len(pivot_df) - 0.5)

# =============================
# Leyenda elegante
# =============================
ax.legend(
    title='Familia',
    title_fontsize=14,
    fontsize=14,
    frameon=True,
    edgecolor="lightgray",
    loc='upper left',
    bbox_to_anchor=(1.02, 1),
    ncol=2,
    columnspacing=1.2,
    handletextpad=0.6,
    labelspacing=0.6,
    borderaxespad=0.0
)


plt.tight_layout()

# =============================
# Guardar gráfico
#

# -------------------------------
# 📁 Guardar resultado
# -------------------------------
ruta_fig = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados"
os.makedirs(ruta_fig, exist_ok=True)

plt.savefig(
    os.path.join(ruta_fig, "2.1_Orden_Familia_PRO.png"),
    dpi=350,
    bbox_inches="tight"
)

plt.show()
plt.close()







