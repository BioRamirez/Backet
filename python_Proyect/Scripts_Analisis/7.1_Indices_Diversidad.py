#--------------## Cargar librerias necesarias------------------------------
# Si no las tienes instaladas, ejecuta esta celda una vez:
# Salir del interprete con: exit() exit() python   pip install tabulate pandas numpy scipy scikit-bio openpyxl
#
# !pip install pandas numpy matplotlib tabulate openpyxl

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from tabulate import tabulate
import openpyxl

#--------------## Leer archivo y revisar columnas------------------------------
# Ruta del archivo
ruta = r"D:\Forestal Consultores\2026\FAUNA\BD\BD_SANROQUE.xlsx"
# Leer el archivo Excel
Registros = pd.read_excel(ruta)

# Mostrar las primeras filas
print(" Primeras filas del archivo:")
print(Registros.head())

# Mostrar nombres de las columnas
print("\n Columnas del DataFrame:")
print(Registros.columns)

# =========================================================
# FILTRAR UNA SOLA CLASE
# =========================================================

grupo = "MAMIFEROS"   # AVES, MAMIFEROS, REPTILES, ANFIBIOS etc.

Registros = Registros[
    Registros["CLASE"].astype(str).str.upper() == grupo.upper()
].copy()

# Reiniciar índice
Registros.reset_index(drop=True, inplace=True)

# =========================================================
# VERIFICACIÓN
# =========================================================

print(f"\nGrupo seleccionado: {grupo}")
print(f"Número de registros: {len(Registros)}")

print("\nPrimeras filas:")
print(Registros.head())
#------------------------------## Crear matriz de abundancia------------------------------

#------------------------------Calcular indices de diversidad------------------------------
import re

def generar_abreviacion(nombre):
    """
    Genera abreviaciones automáticas a partir de nombres de coberturas.
    Ejemplo: 'Bosque de galería y ripario' → 'Bgr'
    """
    # Convertir a minúsculas y dividir en palabras
    palabras = nombre.lower().split()

    # Eliminar conectores comunes
    palabras = [p for p in palabras if p not in ['de', 'del', 'la', 'el', 'y', 'con', 'en', 'los', 'las']]

    # Tomar la primera letra de cada palabra
    abreviacion = ''.join([p[0] for p in palabras])

    # Asegurar que tenga al menos 3 caracteres (rellena si es corta)
    if len(abreviacion) < 3:
        abreviacion = abreviacion.ljust(3, ' ')

    return abreviacion.upper()

#return abreviacion.capitalize()

def abreviar_coberturas(df, columna='COBERTURA'):
    """
    Crea un diccionario de abreviaciones y reemplaza los nombres en el DataFrame.
    """
    coberturas_unicas = df[columna].unique()
    abreviaciones = {c: generar_abreviacion(c) for c in coberturas_unicas}

    print(" Abreviaciones generadas automáticamente:")
    for original, abrev in abreviaciones.items():
        print(f"  {original} → {abrev}")

    # Reemplazar en el DataFrame
    df[columna] = df[columna].replace(abreviaciones)

    return df, abreviaciones


# --- Aplicar las abreviaciones en el DataFrame ---
Registros, abreviaciones_cobertura = abreviar_coberturas(Registros, columna='COBERTURA')

print(" Abreviaciones aplicadas a la columna 'COBERTURA':")
print(Registros['COBERTURA'].unique())


import pandas as pd
import numpy as np
from scipy.stats import entropy
import matplotlib.pyplot as plt

# --- 1. Crear matriz de abundancia por cobertura ---
matriz_abundancia = Registros.pivot_table(
    index='COBERTURA',
    columns='ESPECIE',
    values='INDIVIDUOS',
    aggfunc='sum',
    fill_value=0
)

print(" Matriz de abundancia creada:")
print(matriz_abundancia.head())

import numpy as np
import pandas as pd
from skbio.diversity.alpha import shannon, simpson
from math import log, sqrt

# --- 1. Función para calcular índices ecológicos con scikit-bio ---
def calcular_indices_skbio(abundancias):
    abundancias = np.array(abundancias)
    N = abundancias.sum()
    S = np.count_nonzero(abundancias)
    
    if N == 0 or S == 0:
        return {
            'Riqueza (S)': 0,
            'Abundancia (N)': 0,
            "Shannon (H')": 0,
            'Simpson (1-D)': 0,
            'Dominancia (D)': 0,
            "Equidad (J')": 0,
            'Margalef (DMg)': 0,
            'Menhinick (DMn)': 0
        }

    # Índices calculados con scikit-bio
    H = shannon(abundancias, base=np.e)         # Diversidad de Shannon (H’)
    D_simpson = simpson(abundancias)            # Dominancia de Simpson (D)
    one_minus_D = 1 - D_simpson                 # Diversidad de Simpson (1 - D)

    # Índices clásicos adicionales
    J = H / log(S)                              # Equidad de Pielou (J’)
    DMg = (S - 1) / log(N)                      # Índice de Margalef
    DMn = S / sqrt(N)                           # Índice de Menhinick

    return {
        'Riqueza (S)': S,
        'Abundancia (N)': N,
        "Shannon (H')": H,
        'Simpson (1-D)': D_simpson,
        'Dominancia (D)': one_minus_D, 
        "Equidad (J')": J,
        'Margalef (DMg)': DMg,
        'Menhinick (DMn)': DMn
    }

# --- 2. Aplicar la función por cobertura ---
calcular_indices = calcular_indices_skbio

indices_diversidad = matriz_abundancia.apply(calcular_indices, axis=1, result_type='expand')

print("\n Índices de diversidad por cobertura (basados en scikit-bio):")
print(indices_diversidad)



# --- 4. Calcular índices totales (todas las coberturas combinadas) ---
abundancia_total = matriz_abundancia.sum(axis=0)
indices_totales = calcular_indices(abundancia_total)

print("\n Índices de diversidad total (todas las coberturas combinadas):")
for k, v in indices_totales.items():
    print(f"{k}: {v:.4f}")

# Convertir los índices totales a DataFrame
indices_totales_df = pd.DataFrame([indices_totales], index=['Total'])


# Combinar ambos DataFrames
indices_combinados = pd.concat([indices_diversidad, indices_totales_df])

print(" Tabla combinada de índices de diversidad:")
print(indices_combinados)

indices_combinados = indices_combinados.round(4)
print(" Tabla combinada de índices de diversidad (redondeada):")
print(indices_combinados)

indices_pivot = indices_combinados.T
indices_pivot.index.name = 'Índice'
indices_pivot.columns.name = 'Cobertura'

print(indices_pivot)



indices_pivot.to_excel("D:/CORPONOR 2025/Backet/python_Proyect/Resultados/7.1_Indices_Diversidad_Combinados.xlsx")
# Confirmar la ubicación del archivo guardado
print(f" Archivo exportado correctamente en:\nD:/CORPONOR 2025/Backet/python_Proyect/Resultados/7.1_Indices_Diversidad_Combinados.xlsx")

#--------------------Formatear grafica--------------------

#---------------------------------- Reparar y formatear archivo de Indices_Diversidad_Combinados -----------------------
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import os

# --- Rutas ---
ruta_original = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\7.1_Indices_Diversidad_Combinados.xlsx"
ruta_limpia = r"D:\CORPONOR 2025\Backet\python_Proyect\Resultados\7.1_Indices_Diversidad_Combinados.xlsx"

# --- Verificar existencia ---
if not os.path.exists(ruta_original):
    raise FileNotFoundError(f"⚠️ No se encontró el archivo: {ruta_original}")

# --- Leer archivo dañado con pandas ---
try:
    df = pd.read_excel(ruta_original)
    print(" Archivo leído correctamente con pandas.")
except Exception as e:
    raise RuntimeError(f" No se pudo leer el archivo: {e}")

# --- Reescribir el archivo limpio ---
df.to_excel(ruta_limpia, index=False)
print(f" Archivo reparado y guardado como:\n{ruta_limpia}")

# --- Aplicar formato con openpyxl ---
from openpyxl import load_workbook

wb = load_workbook(ruta_limpia)
ws = wb.active

# --- Estilos base ---
header_fill = PatternFill(start_color='BFD8B8', end_color='BFD8B8', fill_type='solid')
header_font = Font(bold=True, color='000000', name='Calibri')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# --- Aplicar formato y reemplazar vacíos ---
for row in ws.iter_rows():
    for cell in row:
        if cell.value is None or str(cell.value).strip() == '':
            cell.value = '-'
        cell.alignment = center_align
        cell.border = thin_border

# --- Encabezado ---
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align

# --- Ajustar ancho de columnas ---
for col in ws.columns:
    max_length = 0
    column = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            length = len(str(cell.value))
            if length > max_length:
                max_length = length
    ws.column_dimensions[column].width = max_length + 3

# --- Ajustar altura de filas ---
for row in ws.iter_rows():
    ws.row_dimensions[row[0].row].height = 18

# --- Guardar cambios ---
wb.save(ruta_limpia)
print(f' Archivo formateado y reparado correctamente:\n{ruta_limpia}')

#-------------Fin del formateo del archivo----------------------------
#---------------------------Interpreetar tabla de Indices de diversidad---------------------------
import pandas as pd
import numpy as np

def clasificar_indice(valor, rangos):
    """Clasifica un valor numérico según los rangos definidos."""
    for categoria, (minv, maxv) in rangos.items():
        if minv <= valor < maxv:
            return categoria
    return "Fuera de rango"

def interpretar_indices_completo(df):
    """
    Interpreta automáticamente una tabla de índices ecológicos.
    Incluye descripción de cada índice, su interpretación y nivel (bajo, medio, alto).
    Filas = índices, columnas = coberturas (la última puede ser 'Total').
    """
    interpretaciones = []
    
    # --- Rangos ecológicos generales ---
    rangos_dict = {
    "Shannon": {"Baja": (0, 2), "Media": (2, 3.5), "Alta": (3.5, 100)},
    "Simpson": {"Baja": (0, 0.5), "Media": (0.5, 0.75), "Alta": (0.75, 1.01)},
    "Dominancia": {"Alta": (0.6, 1), "Media": (0.05, 0.1), "Baja": (0, 0.3)},
    "Equidad": {"Baja": (0, 0.6), "Media": (0.6, 0.8), "Alta": (0.8, 1.01)},
    "Margalef": {"Baja": (0, 3), "Media": (3, 5), "Alta": (5, 100)},
    "Menhinick": {"Baja": (0, 2), "Media": (2, 4), "Alta": (4, 100)},
    "Riqueza (S)": {"Baja": (0, 40), "Media": (40, 90), "Alta": (90, 500)},
    "Abundancia (N)": {"Baja": (0, 500), "Media": (500, 1500), "Alta": (1500, 10000)}
}

    # --- Descripciones breves de cada índice ---
    descripciones = {
        "Riqueza": "El índice de **riqueza (S)** representa el número total de especies registradas en una cobertura. No considera la abundancia, solo cuántas especies hay.",
        "Abundancia": "La **abundancia (N)** refleja el número total de individuos registrados; una alta abundancia puede indicar hábitats más productivos o mejor muestreados.",
        "Shannon": "El índice de **Shannon (H’)** mide la diversidad teniendo en cuenta tanto la riqueza de especies como su equidad. Valores altos indican comunidades más diversas y equilibradas.",
        "Simpson": "El índice de **Simpson (1–D)** expresa la probabilidad de que dos individuos seleccionados al azar pertenezcan a especies diferentes. Valores cercanos a 1 reflejan alta diversidad.",
        "Dominancia": "El índice de **Dominancia (D)** mide el grado en que una o pocas especies dominan el ensamblaje. Valores altos indican dominancia de pocas especies.",
        "Equidad": "La **equidad (J’)** describe cuán uniformemente se distribuyen los individuos entre las especies. Valores altos indican distribución equitativa.",
        "Margalef": "El índice de **Margalef (DMg)** ajusta la riqueza de especies en función del número de individuos, útil para comparar entre coberturas con distinto esfuerzo de muestreo.",
        "Menhinick": "El índice de **Menhinick (DMn)** también ajusta la riqueza según la abundancia total, proporcionando una medida estandarizada de riqueza relativa."
    }

    # --- Copia de seguridad del DataFrame ---
    datos = df.copy()
    
    # Identificar la columna de Totales si existe
    col_total = None
    for col in datos.columns:
        if "total" in col.lower():
            col_total = col
            break
    
    coberturas = [c for c in datos.columns if c != col_total]

    # --- Interpretación por índice ---
    for indice in datos.index:
        nombre_limpio = indice.split(" ")[0].replace("(", "").replace(")", "")
        valores = datos.loc[indice, coberturas]

        if np.issubdtype(valores.dtype, np.number):
            cobertura_max = valores.idxmax()
            cobertura_min = valores.idxmin()
            val_max = valores.max()
            val_min = valores.min()

            # --- Descripción ---
            tipo = next((k for k in rangos_dict.keys() if k.lower() in nombre_limpio.lower()), None)
            desc = next((v for k, v in descripciones.items() if k.lower() in nombre_limpio.lower()), None)
            if not desc:
                desc = f"El índice **{indice}** evalúa un aspecto ecológico particular del ensamblaje de especies."

            interpretaciones.append(f"\n {desc}")

            # --- Clasificación ---
            if tipo:
                nivel_max = clasificar_indice(val_max, rangos_dict[tipo])
                nivel_min = clasificar_indice(val_min, rangos_dict[tipo])
                interpretaciones.append(
                    f" En este índice, la cobertura con valor más alto es **{cobertura_max}** "
                    f"({val_max:.3f}, categoría {nivel_max.lower()}) y la más baja es **{cobertura_min}** "
                    f"({val_min:.3f}, categoría {nivel_min.lower()})."
                )
            else:
                interpretaciones.append(
                    f" La cobertura con mayor valor de **{indice}** es **{cobertura_max}** "
                    f"({val_max:.3f}), mientras que la más baja es **{cobertura_min}** ({val_min:.3f})."
                )

    # --- Interpretación general ---
    interpretaciones.append("\n **Síntesis ecológica general:**")
    interpretaciones.append(
        "Altos valores en los índices de **Shannon** y **Simpson** reflejan comunidades con gran diversidad y "
        "una distribución equilibrada de individuos entre especies. "
        "En contraste, valores altos de **Dominancia** indican concentración de abundancia en pocas especies. "
        "Los índices de **Equidad** expresan el grado de uniformidad en la distribución de individuos, "
        "mientras que **Margalef** y **Menhinick** complementan la evaluación de la riqueza relativa ajustada por abundancia."
    )

    if col_total:
        interpretaciones.append(
            f"\n Finalmente, la columna **{col_total}** resume los valores combinados del conjunto total de coberturas, "
            "brindando una visión general de la diversidad del ecosistema muestreado."
        )

    return "\n".join(interpretaciones)


# --- Ejemplo de uso ---
interpretacion = interpretar_indices_completo(indices_pivot)
print(interpretacion)

# ---------------------------- Crear informe TXT ----------------------------
import pandas as pd

# Ruta de salida del TXT
ruta_salida = r"D:/CORPONOR 2025/Backet/python_Proyect/Resultados/7.1.1_Interpretacion_Indices_Diversidad.txt"

# Construcción del contenido del TXT
contenido = []

# --- Título principal ---
contenido.append("INFORME DE ÍNDICES DE DIVERSIDAD")
contenido.append("=" * 60)
contenido.append("\n")

# --- Subtítulo ---
contenido.append("TABLA DE ÍNDICES CALCULADOS")
contenido.append("-" * 60)

# --- Convertir tabla a texto ---
tabla_txt = indices_pivot.to_string(index=True)
contenido.append(tabla_txt)
contenido.append("\n")

# --- Interpretación automática ---
contenido.append("INTERPRETACIÓN AUTOMÁTICA DE LOS RESULTADOS")
contenido.append("-" * 60)
contenido.append(interpretacion)
contenido.append("\n")

# --- Guardar como archivo TXT ---
with open(ruta_salida, "w", encoding="utf-8") as f:
    f.write("\n".join(contenido))

print(f"Archivo TXT guardado en:\n{ruta_salida}")

#----------------------------Fin del codigo----------------------------

#------------------Validar calculos con scikit-bio------------------------------
 # --- 2. Función para calcular índices ecológicos ---
def calcular_indices(abundancias):
    abundancias = np.array(abundancias)
    N = abundancias.sum()
    S = np.count_nonzero(abundancias)
    
    if N == 0 or S == 0:
        return {
            'Riqueza (S)': 0,
            'Abundancia (N)': 0,
            'Shannon (H\')': 0,
            'Simpson (1-D)': 0,
            'Dominancia (D)': 0,
            'Equidad (J\')': 0,
            'Margalef (DMg)': 0,
            'Menhinick (DMn)': 0
        }

    # Proporciones
    p = abundancias / N

    # Índices de diversidad
    shannon = entropy(p, base=np.e)
    simpson = 1 - np.sum(p**2)
    dominancia = np.sum(p**2)
    pielou = shannon / np.log(S)
    margalef = (S - 1) / np.log(N)
    menhinick = S / np.sqrt(N)

    return {
        'Riqueza (S)': S,
        'Abundancia (N)': N,
        'Shannon (H\')': shannon,
        'Simpson (1-D)': simpson,
        'Dominancia (D)': dominancia,
        'Equidad (J\')': pielou,
        'Margalef (DMg)': margalef,
        'Menhinick (DMn)': menhinick
    }

# --- 3. Calcular índices por cobertura ---
indices_diversidad2 = matriz_abundancia.apply(calcular_indices, axis=1, result_type='expand')

print("\n Índices de diversidad por cobertura:")
print(indices_diversidad2)

# assumes matriz_abundancia: rows=coberturas, cols=especies (enteros)
from skbio.diversity import alpha_diversity
import numpy as np
from scipy.stats import entropy

# preparar datos
counts = matriz_abundancia.values  # array shape (n_samples, n_species)
ids = matriz_abundancia.index.astype(str).tolist()

# calcular con scikit-bio
shannon_skbio = alpha_diversity('shannon', counts, ids=ids)   # devuelve H' (base e)
simpson_skbio = alpha_diversity('simpson', counts, ids=ids)   # ojo: devuelve D = sum p^2 en algunas versiones

# calcular con tu implementación (ejemplo para Shannon y Simpson 1-D)
shannon_manual = matriz_abundancia.apply(lambda row: entropy(row / row.sum(), base=np.e), axis=1)
simpson_manual = matriz_abundancia.apply(lambda row: 1 - np.sum((row / row.sum())**2), axis=1)

# comparar (tolerancia numérica)
print("Shannon equal:", np.allclose(shannon_manual.values, shannon_skbio.values, atol=1e-8))
# Para Simpson, comprobar si scikit-bio devuelve D o 1-D:
print("Simpson manual sample:", simpson_manual.iloc[0])
print("Simpson scikit-bio sample:", simpson_skbio.values[0])
# Si scikit-bio devuelve D = sum p^2, entonces 1 - simpson_skbio == simpson_manual







































#=========================================
#   GRAFICAR TODA LA TABLA indices_pivot
#=========================================
import matplotlib.pyplot as plt
import numpy as np
import os

# ================================
# COLORES (oro mejorado)
# ================================
color_bajo = "#d97a00"    # naranja quemado
color_medio = "#d4af37"    # ORO
color_alto = "#2e8b57"   # verde

# ================================
# RANGOS ECOLÓGICOS
# ================================
rangos_dict = {
    "Shannon": {"Baja": (0, 2), "Media": (2, 3.5), "Alta": (3.5, 100)},
    "Simpson": {"Baja": (0, 0.5), "Media": (0.5, 0.75), "Alta": (0.75, 1.01)},
    "Dominancia": {"Alta": (0.6, 1), "Media": (0.05, 0.1), "Baja": (0, 0.3)},
    "Equidad": {"Baja": (0, 0.6), "Media": (0.6, 0.8), "Alta": (0.8, 1.01)},
    "Margalef": {"Baja": (0, 3), "Media": (3, 5), "Alta": (5, 100)},
    "Menhinick": {"Baja": (0, 2), "Media": (2, 4), "Alta": (4, 100)},
    "Riqueza (S)": {"Baja": (0, 40), "Media": (40, 90), "Alta": (90, 500)},
    "Abundancia (N)": {"Baja": (0, 500), "Media": (500, 1500), "Alta": (1500, 10000)}
}

# mapeo entre nombre en tabla y clave de rangos
mapa_indices = {
    "Shannon (H')": "Shannon",
    "Simpson (1-D)": "Simpson",
    "Dominancia (D)": "Dominancia",
    "Equidad (J')": "Equidad",
    "Margalef (DMg)": "Margalef",
    "Menhinick (DMn)": "Menhinick",
    "Riqueza (S)": "Riqueza (S)",
    "Abundancia (N)": "Abundancia (N)"
}

# ========================================
# FUNCIÓN PARA GRAFICAR UN ÍNDICE
# ========================================
def graficar_indice_fila(nombre_indice, valores, carpeta_salida="D:/CORPONOR 2025/Backet/python_Proyect/Resultados"):

    if not os.path.exists(carpeta_salida):
        os.makedirs(carpeta_salida)

    sitios = valores.index
    y = valores.values.astype(float)

    clave_rango = mapa_indices.get(nombre_indice, None)

    if clave_rango not in rangos_dict:
        print(f"⚠ No hay rangos para {nombre_indice}. Usando color base.")
        colores = [color_bajo] * len(y)
    else:
        rangos = rangos_dict[clave_rango]

        def asignar_color(v):
            for categoria, (min_v, max_v) in rangos.items():
                if min_v <= v <= max_v:
                    return {
                        "Baja": color_bajo,
                        "Media": color_medio,
                        "Alta": color_alto
                    }[categoria]
            return color_bajo

        colores = [asignar_color(v) for v in y]

    # ---- Crear figura ----
    fig, ax = plt.subplots(figsize=(10, 6))
    barras = ax.bar(sitios, y, color=colores, edgecolor="black")

    # Etiquetas de valor encima
    for barra, valor in zip(barras, y):
        ax.text(
            barra.get_x() + barra.get_width() / 2,
            barra.get_height() + 0.02 * max(y),
            f"{valor:.3f}" if valor < 40 else f"{valor:.0f}",
            ha="center", va="bottom", fontsize=10, fontweight="bold"
        )

    ax.set_title(nombre_indice, fontsize=14, fontweight="bold")
    ax.set_ylabel(nombre_indice)
    plt.xticks(rotation=45, ha="right")

    # ---- Leyenda abajo centrada ----
    handles = [
        plt.Rectangle((0, 0), 1, 1, color=color_bajo),
        plt.Rectangle((0, 0), 1, 1, color=color_medio),
        plt.Rectangle((0, 0), 1, 1, color=color_alto),
    ]

    labels = ["Bajo", "Medio", "Alto"]

    fig.legend(
        handles, labels,
        title="Categorías",
        loc="lower center",
        ncol=3,
        frameon=False,
        bbox_to_anchor=(0.5, -0.05)
    )

    plt.tight_layout()

    # Guardar archivo
    nombre_archivo = ("7.1_" +
        nombre_indice.replace(" ", "_").replace("'", "").replace("(", "").replace(")", "") + ".png"
    )

    plt.savefig(os.path.join(carpeta_salida, nombre_archivo),
                dpi=300, bbox_inches="tight")
    plt.close()

    print(f"✔ Gráfico guardado: {nombre_archivo}")


# ====================================================
#   EJECUTAR GRAFICACIÓN SOBRE indices_pivot
# ====================================================
for indice in indices_pivot.index:
    graficar_indice_fila(indice, indices_pivot.loc[indice])































# ============================================================
#          GRÁFICOS DE ÍNDICES – ESTILO MINIMALISTA SET2
# ============================================================

import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
import os

# =====================================================
#      🎨  PALETA DEFINITIVA (Set2 – estilo del usuario)
# =====================================================
set2 = sns.color_palette("Paired", 12)

color_bajo  = set2[5]   # pastel 1
color_medio = set2[7]   # pastel 2
color_alto  = set2[3]   # pastel 3

# =====================================================
#        RANGOS ECOLÓGICOS (no se modifican)
# =====================================================
rangos_dict = {
    "Shannon": {"Baja": (0, 2), "Media": (2, 3.5), "Alta": (3.5, 100)},
    "Simpson": {"Baja": (0, 0.5), "Media": (0.5, 0.75), "Alta": (0.75, 1.01)},
   "Dominancia": {"Alta": (0.6, 1), "Media": (0.05, 0.1), "Baja": (0, 0.3)},
    "Equidad": {"Baja": (0, 0.6), "Media": (0.6, 0.8), "Alta": (0.8, 1.01)},
    "Margalef": {"Baja": (0, 3), "Media": (3, 5), "Alta": (5, 100)},
    "Menhinick": {"Baja": (0, 2), "Media": (2, 4), "Alta": (4, 100)},
    "Riqueza (S)": {"Baja": (0, 40), "Media": (40, 90), "Alta": (90, 500)},
    "Abundancia (N)": {"Baja": (0, 500), "Media": (500, 1500), "Alta": (1500, 10000)}
}

mapa_indices = {
    "Shannon (H')": "Shannon",
    "Simpson (1-D)": "Simpson",
    "Dominancia (D)": "Dominancia",
    "Equidad (J')": "Equidad",
    "Margalef (DMg)": "Margalef",
    "Menhinick (DMn)": "Menhinick",
    "Riqueza (S)": "Riqueza (S)",
    "Abundancia (N)": "Abundancia (N)"
}

# =====================================================
#     FUNCIÓN – GRAFICACIÓN ESTILO MINIMALISTA SET2
# =====================================================
def graficar_indice_fila(nombre_indice, valores, carpeta_salida="D:/CORPONOR 2025/Backet/python_Proyect/Resultados"):

    os.makedirs(carpeta_salida, exist_ok=True)

    sitios = valores.index
    y = valores.values.astype(float)

    clave_rango = mapa_indices.get(nombre_indice, None)

    # Asignación de colores por rango ecológico
    if clave_rango not in rangos_dict:
        colores = [color_medio] * len(y)
    else:
        rangos = rangos_dict[clave_rango]

        def asignar_color(v):
            for cat, (vmin, vmax) in rangos.items():
                if vmin <= v <= vmax:
                    return {
                        "Baja": color_bajo,
                        "Media": color_medio,
                        "Alta": color_alto
                    }[cat]
            return color_bajo

        colores = [asignar_color(v) for v in y]

    # =====================================================
    #         FIGURA – ESTILO MINIMALISTA PROFESIONAL
    # =====================================================
    sns.set_theme(style="whitegrid")

    fig, ax = plt.subplots(figsize=(10, 6))

    barras = ax.bar(
        sitios, y,
        color=colores,
        edgecolor="black",
        linewidth=0.8
    )

    # Etiquetas sobre barras
    for bar, val in zip(barras, y):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + max(y) * 0.015,
            f"{val:.3f}" if val < 40 else f"{val:.0f}",
            ha="center",
            va="bottom",
            fontsize=10,
            fontweight="bold"
        )

    # Títulos
    ax.set_title(
        nombre_indice,
        fontsize=15,
        fontweight="bold",
        pad=12
    )

    ax.set_ylabel(nombre_indice, fontsize=12)
    plt.xticks(rotation=45, ha="right", fontsize=11)
    plt.yticks(fontsize=11)

    # Bordes refinados
    for spine in ax.spines.values():
        spine.set_linewidth(0.8)
        spine.set_color("#444444")

    # =====================================================
    #                 LEYENDA MINIMALISTA
    # =====================================================
    handles = [
        plt.Rectangle((0, 0), 1, 1, color=color_bajo),
        plt.Rectangle((0, 0), 1, 1, color=color_medio),
        plt.Rectangle((0, 0), 1, 1, color=color_alto)
    ]

    labels = ["Bajo", "Medio", "Alto"]

    fig.legend(
        handles,
        labels,
        title="Categoría",
        title_fontsize=12,
        fontsize=11,
        loc="lower center",
        ncol=3,
        frameon=False,
        bbox_to_anchor=(0.5, -0.08)
    )

    plt.tight_layout()

    # Guardar archivo
    nombre_archivo = (
        "7.1.1_" +
        nombre_indice.replace(" ", "_").replace("'", "").replace("(", "").replace(")", "") +
        ".png"
    )

    plt.savefig(
        os.path.join(carpeta_salida, nombre_archivo),
        dpi=300,
        bbox_inches="tight"
    )

    plt.close()

    print(f"✔ Gráfico guardado: {nombre_archivo}")
# ====================================================
#   EJECUTAR GRAFICACIÓN SOBRE indices_pivot
# ====================================================
for indice in indices_pivot.index:
    graficar_indice_fila(indice, indices_pivot.loc[indice])


























# ============================================================
#     GRÁFICO ÚNICO DE TODOS LOS ÍNDICES ECOLÓGICOS
# ============================================================

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import os

# ============================================================
# COLORES
# ============================================================

color_bajo  = "#d97a00"   # naranja
color_medio = "#d4af37"   # oro
color_alto  = "#2e8b57"   # verde

# ============================================================
# RANGOS ECOLÓGICOS
# ============================================================

rangos_dict = {
    "Shannon": {"Baja": (0, 2), "Media": (2, 3.5), "Alta": (3.5, 100)},
    "Simpson": {"Baja": (0, 0.5), "Media": (0.5, 0.75), "Alta": (0.75, 1.01)},
    "Dominancia": {"Alta": (0.6, 1), "Media": (0.05, 0.1), "Baja": (0, 0.3)},
    "Equidad": {"Baja": (0, 0.6), "Media": (0.6, 0.8), "Alta": (0.8, 1.01)},
    "Margalef": {"Baja": (0, 3), "Media": (3, 5), "Alta": (5, 100)},
    "Menhinick": {"Baja": (0, 2), "Media": (2, 4), "Alta": (4, 100)},
    "Riqueza (S)": {"Baja": (0, 40), "Media": (40, 90), "Alta": (90, 500)},
    "Abundancia (N)": {"Baja": (0, 500), "Media": (500, 1500), "Alta": (1500, 10000)}
}

# ============================================================
# MAPEO DE NOMBRES
# ============================================================

mapa_indices = {
    "Shannon (H')": "Shannon",
    "Simpson (1-D)": "Simpson",
    "Dominancia (D)": "Dominancia",
    "Equidad (J')": "Equidad",
    "Margalef (DMg)": "Margalef",
    "Menhinick (DMn)": "Menhinick",
    "Riqueza (S)": "Riqueza (S)",
    "Abundancia (N)": "Abundancia (N)"
}

# ============================================================
# EXTRAER UNA SOLA COBERTURA
# ============================================================

# Toma la primera columna de indices_pivot
valores = indices_pivot.iloc[:, 0]

# Nombres de índices
indices = valores.index.tolist()

# Valores numéricos
y = valores.values.astype(float)

# ============================================================
# ASIGNAR COLORES SEGÚN CATEGORÍA
# ============================================================

colores = []

for nombre_indice, valor in zip(indices, y):

    clave = mapa_indices.get(nombre_indice)

    if clave not in rangos_dict:

        colores.append(color_medio)

    else:

        rangos = rangos_dict[clave]

        color_actual = color_medio

        for categoria, (vmin, vmax) in rangos.items():

            if vmin <= valor <= vmax:

                if categoria == "Baja":
                    color_actual = color_bajo

                elif categoria == "Media":
                    color_actual = color_medio

                elif categoria == "Alta":
                    color_actual = color_alto

        colores.append(color_actual)

# ============================================================
# CREAR FIGURA
# ============================================================

plt.figure(figsize=(12, 7))

barras = plt.bar(
    indices,
    y,
    color=colores,
    edgecolor="black",
    linewidth=1
)

# ============================================================
# ETIQUETAS
# ============================================================

max_y = max(y)

for barra, valor in zip(barras, y):

    plt.text(
        barra.get_x() + barra.get_width()/2,
        barra.get_height() + (max_y * 0.02),
        f"{valor:.2f}",
        ha="center",
        va="bottom",
        fontsize=10,
        fontweight="bold"
    )

# ============================================================
# ESTILO
# ============================================================

plt.title(
    "Síntesis de Índices Ecológicos",
    fontsize=18,
    fontweight="bold",
    pad=15
)

plt.ylabel("Valor del índice", fontsize=12)

plt.xticks(
    rotation=25,
    ha="right",
    fontsize=11
)

plt.yticks(fontsize=11)

# Grid elegante
plt.grid(
    axis="y",
    linestyle="--",
    alpha=0.4
)

# ============================================================
# LEYENDA
# ============================================================

handles = [
    plt.Rectangle((0, 0), 1, 1, color=color_bajo),
    plt.Rectangle((0, 0), 1, 1, color=color_medio),
    plt.Rectangle((0, 0), 1, 1, color=color_alto)
]

labels = ["Bajo", "Medio", "Alto"]

plt.legend(
    handles,
    labels,
    title="Categoría ecológica",
    loc="upper right",
    frameon=False
)

# ============================================================
# AJUSTE FINAL
# ============================================================

plt.tight_layout()

# ============================================================
# EXPORTAR
# ============================================================

carpeta_salida = "D:/CORPONOR 2025/Backet/python_Proyect/Resultados"

os.makedirs(carpeta_salida, exist_ok=True)

ruta_salida = os.path.join(
    carpeta_salida,
    "7.1_Sintesis_Indices_Ecologicos.png"
)

plt.savefig(
    ruta_salida,
    dpi=300,
    bbox_inches="tight"
)

plt.close()

print("✔ Gráfico exportado correctamente")
print(ruta_salida)






# ============================================================
#     GRÁFICO ÚNICO DE ÍNDICES ECOLÓGICOS (ESCALA LOG)
# ============================================================

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import os

# ============================================================
# COLORES
# ============================================================

color_bajo  = "#d97a00"   # naranja
color_medio = "#d4af37"   # oro
color_alto  = "#2e8b57"   # verde

# ============================================================
# RANGOS ECOLÓGICOS
# ============================================================

rangos_dict = {
    "Shannon": {"Baja": (0, 2), "Media": (2, 3.5), "Alta": (3.5, 100)},
    "Simpson": {"Baja": (0, 0.5), "Media": (0.5, 0.75), "Alta": (0.75, 1.01)},
    "Dominancia": {"Alta": (0.6, 1), "Media": (0.05, 0.1), "Baja": (0, 0.3)},
    "Equidad": {"Baja": (0, 0.6), "Media": (0.6, 0.8), "Alta": (0.8, 1.01)},
    "Margalef": {"Baja": (0, 3), "Media": (3, 5), "Alta": (5, 100)},
    "Menhinick": {"Baja": (0, 2), "Media": (2, 4), "Alta": (4, 100)},
    "Riqueza (S)": {"Baja": (0, 40), "Media": (40, 90), "Alta": (90, 500)},
    "Abundancia (N)": {"Baja": (0, 500), "Media": (500, 1500), "Alta": (1500, 10000)}
}

# ============================================================
# MAPEO DE NOMBRES
# ============================================================

mapa_indices = {
    "Shannon (H')": "Shannon",
    "Simpson (1-D)": "Simpson",
    "Dominancia (D)": "Dominancia",
    "Equidad (J')": "Equidad",
    "Margalef (DMg)": "Margalef",
    "Menhinick (DMn)": "Menhinick",
    "Riqueza (S)": "Riqueza (S)",
    "Abundancia (N)": "Abundancia (N)"
}

# ============================================================
# EXTRAER UNA SOLA COBERTURA
# ============================================================

valores = indices_pivot.iloc[:, 0]

indices = valores.index.tolist()

y_original = valores.values.astype(float)

# ============================================================
# TRANSFORMACIÓN LOGARÍTMICA
# ============================================================

# log10(x + 1)
# evita problemas con ceros

y_log = np.log10(y_original + 1)

# ============================================================
# ASIGNAR COLORES ECOLÓGICOS
# ============================================================

colores = []

for nombre_indice, valor in zip(indices, y_original):

    clave = mapa_indices.get(nombre_indice)

    if clave not in rangos_dict:

        colores.append(color_medio)

    else:

        rangos = rangos_dict[clave]

        color_actual = color_medio

        for categoria, (vmin, vmax) in rangos.items():

            if vmin <= valor <= vmax:

                if categoria == "Baja":
                    color_actual = color_bajo

                elif categoria == "Media":
                    color_actual = color_medio

                elif categoria == "Alta":
                    color_actual = color_alto

        colores.append(color_actual)

# ============================================================
# CREAR FIGURA
# ============================================================

plt.figure(figsize=(13, 7))

barras = plt.bar(
    indices,
    y_log,
    color=colores,
    edgecolor="black",
    linewidth=1
)

# ============================================================
# ETIQUETAS (MOSTRAR VALORES ORIGINALES)
# ============================================================

max_y = max(y_log)

for barra, valor_real in zip(barras, y_original):

    altura = barra.get_height()

    plt.text(
        barra.get_x() + barra.get_width()/2,
        altura + (max_y * 0.03),
        f"{valor_real:.2f}" if valor_real < 100 else f"{valor_real:.0f}",
        ha="center",
        va="bottom",
        fontsize=10,
        fontweight="bold"
    )

# ============================================================
# ESTILO
# ============================================================

plt.title(
    "Síntesis de Índices Ecológicos",
    fontsize=18,
    fontweight="bold",
    pad=15
)

plt.ylabel(
    "Valor transformado [log10(x + 1)]",
    fontsize=12
)

plt.xticks(
    rotation=25,
    ha="right",
    fontsize=11
)

plt.yticks(fontsize=11)

# Grid elegante
plt.grid(
    axis="y",
    linestyle="--",
    alpha=0.4
)

# ============================================================
# LEYENDA
# ============================================================

handles = [
    plt.Rectangle((0, 0), 1, 1, color=color_bajo),
    plt.Rectangle((0, 0), 1, 1, color=color_medio),
    plt.Rectangle((0, 0), 1, 1, color=color_alto)
]

labels = ["Bajo", "Medio", "Alto"]

plt.legend(
    handles,
    labels,
    title="Categoría ecológica",
    loc="upper right",
    frameon=False
)

# ============================================================
# BORDES MINIMALISTAS
# ============================================================

ax = plt.gca()

ax.spines["top"].set_visible(False)
ax.spines["right"].set_visible(False)

# ============================================================
# AJUSTE FINAL
# ============================================================

plt.tight_layout()

# ============================================================
# EXPORTAR
# ============================================================

carpeta_salida = "D:/CORPONOR 2025/Backet/python_Proyect/Resultados"

os.makedirs(carpeta_salida, exist_ok=True)

ruta_salida = os.path.join(
    carpeta_salida,
    "7.1.2_Sintesis_Indices_Ecologicos_LOG.png"
)

plt.savefig(
    ruta_salida,
    dpi=300,
    bbox_inches="tight"
)

plt.close()

print("✔ Gráfico exportado correctamente")
print(ruta_salida)